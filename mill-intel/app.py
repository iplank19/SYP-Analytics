"""
SYP Mill Intelligence Platform
Aggregated mill pricing, interactive map, market intelligence, smart quoting
"""
from flask import Flask, send_from_directory, request, jsonify
from flask_cors import CORS
import requests as req
import os, time, sqlite3, json, tempfile, math
from datetime import datetime, timedelta

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'mill_intel.db')

# State map for region assignment
STATE_REGIONS = {
    'TX':'west','LA':'west','AR':'west','OK':'west','NM':'west',
    'MS':'central','AL':'central','TN':'central','KY':'central','MO':'central',
    'GA':'east','FL':'east','SC':'east','NC':'east','VA':'east','WV':'east',
    'OH':'east','IN':'east','IL':'central','WI':'central','MI':'east','MN':'central',
    'IA':'central','PA':'east','NY':'east','NJ':'east','MD':'east','DE':'east',
    'CT':'east','MA':'east','ME':'east','NH':'east','VT':'east','RI':'east'
}

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS mills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            city TEXT,
            state TEXT,
            lat REAL,
            lon REAL,
            region TEXT,
            products TEXT DEFAULT '[]',
            notes TEXT DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_mills_name ON mills(name);
        CREATE INDEX IF NOT EXISTS idx_mills_region ON mills(region);

        CREATE TABLE IF NOT EXISTS mill_quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mill_id INTEGER NOT NULL,
            mill_name TEXT NOT NULL,
            product TEXT NOT NULL,
            price REAL NOT NULL,
            length TEXT DEFAULT 'RL',
            volume REAL DEFAULT 0,
            tls INTEGER DEFAULT 0,
            ship_window TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            date TEXT NOT NULL,
            trader TEXT NOT NULL,
            source TEXT DEFAULT 'manual',
            raw_text TEXT DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (mill_id) REFERENCES mills(id)
        );
        CREATE INDEX IF NOT EXISTS idx_mq_mill ON mill_quotes(mill_id);
        CREATE INDEX IF NOT EXISTS idx_mq_product ON mill_quotes(product);
        CREATE INDEX IF NOT EXISTS idx_mq_date ON mill_quotes(date);
        CREATE INDEX IF NOT EXISTS idx_mq_trader ON mill_quotes(trader);
        CREATE INDEX IF NOT EXISTS idx_mq_composite ON mill_quotes(mill_name, product, date);

        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            destination TEXT,
            lat REAL,
            lon REAL,
            trader TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS rl_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            product TEXT NOT NULL,
            region TEXT NOT NULL,
            price REAL NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_rl_date ON rl_prices(date);
        CREATE UNIQUE INDEX IF NOT EXISTS idx_rl_unique ON rl_prices(date, product, region);

        CREATE TABLE IF NOT EXISTS lanes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            origin TEXT NOT NULL,
            dest TEXT NOT NULL,
            miles INTEGER NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(origin, dest)
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    ''')
    conn.commit()
    conn.close()

init_db()

# ===== GEOCODING =====
geo_cache = {}

def geocode_location(location):
    if not location:
        return None
    cache_key = location.lower().strip()
    if cache_key in geo_cache:
        return geo_cache[cache_key]
    # Also check DB for previously geocoded mills
    conn = get_db()
    row = conn.execute("SELECT lat, lon FROM mills WHERE LOWER(city)=? AND lat IS NOT NULL", (cache_key,)).fetchone()
    conn.close()
    if row:
        coords = {'lat': row['lat'], 'lon': row['lon']}
        geo_cache[cache_key] = coords
        return coords
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {'q': location, 'format': 'json', 'limit': 1, 'countrycodes': 'us'}
        headers = {'User-Agent': 'SYP-Mill-Intel/1.0'}
        resp = req.get(url, params=params, headers=headers, timeout=10)
        results = resp.json()
        if results:
            coords = {'lat': float(results[0]['lat']), 'lon': float(results[0]['lon'])}
            geo_cache[cache_key] = coords
            return coords
        return None
    except Exception as e:
        print(f"Geocode error for {location}: {e}")
        return None

def get_distance(origin_coords, dest_coords):
    try:
        coords_str = f"{origin_coords['lon']},{origin_coords['lat']};{dest_coords['lon']},{dest_coords['lat']}"
        url = f"https://router.project-osrm.org/route/v1/driving/{coords_str}"
        resp = req.get(url, params={'overview': 'false'}, timeout=10)
        data = resp.json()
        if data.get('code') == 'Ok' and data.get('routes'):
            return round(data['routes'][0]['distance'] / 1609.34)
        return None
    except Exception as e:
        print(f"Distance error: {e}")
        return None

def extract_state(location):
    """Extract 2-letter state code from location string like 'DeQuincy, LA'"""
    if not location:
        return None
    parts = location.strip().rstrip('.').split(',')
    if len(parts) >= 2:
        st = parts[-1].strip().upper()[:2]
        if len(st) == 2 and st.isalpha():
            return st
    return None

def get_region(state_code):
    return STATE_REGIONS.get(state_code, 'central')

# ===== ROUTES =====

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

# ----- MILLS -----

@app.route('/api/mills', methods=['GET'])
def list_mills():
    conn = get_db()
    rows = conn.execute("SELECT * FROM mills ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mills', methods=['POST'])
def create_mill():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Mill name required'}), 400
    city = data.get('city', '')
    state = data.get('state', '') or extract_state(city) or ''
    region = data.get('region', '') or get_region(state)
    lat = data.get('lat')
    lon = data.get('lon')
    # Auto-geocode if we have city but no coords
    if city and (lat is None or lon is None):
        coords = geocode_location(city)
        if coords:
            lat, lon = coords['lat'], coords['lon']
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO mills (name, city, state, lat, lon, region, products, notes) VALUES (?,?,?,?,?,?,?,?)",
            (name, city, state, lat, lon, region, json.dumps(data.get('products', [])), data.get('notes', ''))
        )
        conn.commit()
        mill = conn.execute("SELECT * FROM mills WHERE name=?", (name,)).fetchone()
        conn.close()
        return jsonify(dict(mill)), 201
    except sqlite3.IntegrityError:
        # Mill exists, update it
        conn.execute(
            "UPDATE mills SET city=COALESCE(NULLIF(?,''),city), state=COALESCE(NULLIF(?,''),state), lat=COALESCE(?,lat), lon=COALESCE(?,lon), region=COALESCE(NULLIF(?,''),region), updated_at=CURRENT_TIMESTAMP WHERE name=?",
            (city, state, lat, lon, region, name)
        )
        conn.commit()
        mill = conn.execute("SELECT * FROM mills WHERE name=?", (name,)).fetchone()
        conn.close()
        return jsonify(dict(mill)), 200

@app.route('/api/mills/<int:mill_id>', methods=['GET'])
def get_mill(mill_id):
    conn = get_db()
    mill = conn.execute("SELECT * FROM mills WHERE id=?", (mill_id,)).fetchone()
    if not mill:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    quotes = conn.execute("SELECT * FROM mill_quotes WHERE mill_id=? ORDER BY date DESC LIMIT 100", (mill_id,)).fetchall()
    conn.close()
    result = dict(mill)
    result['quotes'] = [dict(q) for q in quotes]
    return jsonify(result)

@app.route('/api/mills/<int:mill_id>', methods=['PUT'])
def update_mill(mill_id):
    data = request.json
    conn = get_db()
    fields = []
    vals = []
    for k in ['name', 'city', 'state', 'lat', 'lon', 'region', 'notes']:
        if k in data:
            fields.append(f"{k}=?")
            vals.append(data[k])
    if 'products' in data:
        fields.append("products=?")
        vals.append(json.dumps(data['products']))
    if fields:
        fields.append("updated_at=CURRENT_TIMESTAMP")
        vals.append(mill_id)
        conn.execute(f"UPDATE mills SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    mill = conn.execute("SELECT * FROM mills WHERE id=?", (mill_id,)).fetchone()
    conn.close()
    return jsonify(dict(mill)) if mill else (jsonify({'error': 'Not found'}), 404)

@app.route('/api/mills/geocode', methods=['POST'])
def geocode_mill():
    data = request.json
    location = data.get('location', '')
    if not location:
        return jsonify({'error': 'Location required'}), 400
    coords = geocode_location(location)
    if coords:
        return jsonify(coords)
    return jsonify({'error': f'Could not geocode: {location}'}), 404

# ----- MILL QUOTES -----

@app.route('/api/quotes', methods=['GET'])
def list_quotes():
    conn = get_db()
    conditions = ["1=1"]
    params = []
    for k, col in [('mill', 'mill_name'), ('product', 'product'), ('trader', 'trader')]:
        v = request.args.get(k)
        if v:
            conditions.append(f"{col}=?")
            params.append(v)
    since = request.args.get('since')
    if since:
        conditions.append("date>=?")
        params.append(since)
    until = request.args.get('until')
    if until:
        conditions.append("date<=?")
        params.append(until)
    limit = int(request.args.get('limit', 500))
    sql = f"SELECT * FROM mill_quotes WHERE {' AND '.join(conditions)} ORDER BY date DESC, created_at DESC LIMIT ?"
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/quotes', methods=['POST'])
def submit_quotes():
    """Submit one or more quotes. Auto-creates mills and geocodes new ones."""
    data = request.json
    quotes = data if isinstance(data, list) else [data]
    conn = get_db()
    created = []
    for q in quotes:
        mill_name = q.get('mill', '').strip()
        if not mill_name or not q.get('product') or not q.get('price'):
            continue
        # Validate price is positive number
        try:
            price_val = float(q['price'])
            if price_val <= 0:
                continue
        except (ValueError, TypeError):
            continue
        # Find or create mill
        mill = conn.execute("SELECT * FROM mills WHERE name=?", (mill_name,)).fetchone()
        if not mill:
            city = q.get('city', '')
            state = extract_state(city) if city else ''
            region = get_region(state) if state else 'central'
            lat, lon = None, None
            if city:
                time.sleep(0.3)  # Rate limit
                coords = geocode_location(city)
                if coords:
                    lat, lon = coords['lat'], coords['lon']
            conn.execute(
                "INSERT OR IGNORE INTO mills (name, city, state, lat, lon, region) VALUES (?,?,?,?,?,?)",
                (mill_name, city, state, lat, lon, region)
            )
            conn.commit()
            mill = conn.execute("SELECT * FROM mills WHERE name=?", (mill_name,)).fetchone()
        mill_id = mill['id']
        # Update mill products list
        existing_products = json.loads(mill['products'] or '[]')
        product = q['product']
        if product not in existing_products:
            existing_products.append(product)
            conn.execute("UPDATE mills SET products=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                         (json.dumps(existing_products), mill_id))
        # Insert quote
        conn.execute(
            """INSERT INTO mill_quotes (mill_id, mill_name, product, price, length, volume, tls,
               ship_window, notes, date, trader, source, raw_text)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (mill_id, mill_name, product, price_val,
             q.get('length', 'RL'), float(q.get('volume', 0)), int(q.get('tls', 0)),
             q.get('shipWindow', q.get('ship_window', '')), q.get('notes', ''),
             q.get('date', datetime.now().strftime('%Y-%m-%d')),
             q.get('trader', 'Unknown'), q.get('source', 'manual'), q.get('raw_text', ''))
        )
        created.append(q)
    conn.commit()
    conn.close()
    return jsonify({'created': len(created), 'quotes': created}), 201

@app.route('/api/quotes/<int:quote_id>', methods=['DELETE'])
def delete_quote(quote_id):
    conn = get_db()
    conn.execute("DELETE FROM mill_quotes WHERE id=?", (quote_id,))
    conn.commit()
    conn.close()
    return jsonify({'deleted': quote_id})

@app.route('/api/quotes/latest', methods=['GET'])
def latest_quotes():
    """Latest price per mill+product combination"""
    conn = get_db()
    product = request.args.get('product')
    region = request.args.get('region')
    sql = """
        SELECT mq.*, m.lat, m.lon, m.region, m.city
        FROM mill_quotes mq
        JOIN mills m ON mq.mill_id = m.id
        WHERE mq.id IN (
            SELECT id FROM mill_quotes mq2
            WHERE mq2.mill_name = mq.mill_name AND mq2.product = mq.product
            ORDER BY mq2.date DESC, mq2.created_at DESC
            LIMIT 1
        )
    """
    params = []
    if product:
        sql += " AND mq.product=?"
        params.append(product)
    if region:
        sql += " AND m.region=?"
        params.append(region)
    sql += " ORDER BY mq.product, mq.price"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/quotes/matrix', methods=['GET'])
def quote_matrix():
    """Mills x Products price grid. Use ?detail=length for product+length granularity.
    Use ?product=2x4%232 to filter to one product's lengths."""
    detail = request.args.get('detail', '')
    filter_product = request.args.get('product', '')
    conn = get_db()

    if detail == 'length':
        # Granular: group by mill + product + length
        sql = """
            SELECT mq.mill_name, mq.product, mq.length, mq.price, mq.date, mq.volume,
                   mq.ship_window, mq.tls, mq.trader, m.lat, m.lon, m.region, m.city
            FROM mill_quotes mq
            JOIN mills m ON mq.mill_id = m.id
            WHERE mq.id IN (
                SELECT MAX(id) FROM mill_quotes GROUP BY mill_name, product, length
            )
        """
        params = []
        if filter_product:
            sql += " AND mq.product = ?"
            params.append(filter_product)
        sql += " ORDER BY mq.mill_name, mq.product, mq.length"
        rows = conn.execute(sql, params).fetchall()
        conn.close()

        matrix = {}
        mills = set()
        columns = set()  # "product length" combos
        best_by_col = {}

        # Define sort order for lengths
        def length_sort_key(l):
            if not l or l == 'RL':
                return 999
            try:
                return float(str(l).replace("'", "").split('-')[0])
            except:
                return 998

        for r in rows:
            r = dict(r)
            mill = r['mill_name']
            prod = r['product']
            length = r['length'] or 'RL'
            col_key = f"{prod} {length}'" if length != 'RL' else f"{prod} RL"
            mills.add(mill)
            columns.add(col_key)
            if mill not in matrix:
                matrix[mill] = {}
            matrix[mill][col_key] = {
                'price': r['price'], 'date': r['date'], 'volume': r['volume'],
                'ship_window': r['ship_window'], 'tls': r['tls'], 'trader': r['trader'],
                'product': prod, 'length': length,
                'lat': r['lat'], 'lon': r['lon'], 'region': r['region'], 'city': r['city']
            }
            if col_key not in best_by_col or r['price'] < best_by_col[col_key]:
                best_by_col[col_key] = r['price']

        # Sort columns: by product then by length
        def col_sort(c):
            parts = c.rsplit(' ', 1)
            prod = parts[0]
            length = parts[1].replace("'", "") if len(parts) > 1 else 'RL'
            return (prod, length_sort_key(length))

        sorted_cols = sorted(columns, key=col_sort)

        # Get unique products for filter dropdown
        unique_products = sorted(set(c.rsplit(' ', 1)[0] for c in columns))

        return jsonify({
            'matrix': matrix,
            'mills': sorted(mills),
            'columns': sorted_cols,
            'products': unique_products,
            'best_by_col': best_by_col,
            'detail': 'length'
        })
    else:
        # Original: group by mill + product only
        rows = conn.execute("""
            SELECT mq.mill_name, mq.product, mq.price, mq.date, mq.volume, mq.ship_window,
                   mq.tls, mq.trader, m.lat, m.lon, m.region, m.city
            FROM mill_quotes mq
            JOIN mills m ON mq.mill_id = m.id
            WHERE mq.id IN (
                SELECT MAX(id) FROM mill_quotes GROUP BY mill_name, product
            )
            ORDER BY mq.mill_name, mq.product
        """).fetchall()
        conn.close()

        matrix = {}
        mills = set()
        products = set()
        best_by_product = {}

        for r in rows:
            r = dict(r)
            mill = r['mill_name']
            prod = r['product']
            mills.add(mill)
            products.add(prod)
            if mill not in matrix:
                matrix[mill] = {}
            matrix[mill][prod] = {
                'price': r['price'], 'date': r['date'], 'volume': r['volume'],
                'ship_window': r['ship_window'], 'tls': r['tls'], 'trader': r['trader'],
                'lat': r['lat'], 'lon': r['lon'], 'region': r['region'], 'city': r['city']
            }
            if prod not in best_by_product or r['price'] < best_by_product[prod]:
                best_by_product[prod] = r['price']

        return jsonify({
            'matrix': matrix,
            'mills': sorted(mills),
            'products': sorted(products),
            'best_by_product': best_by_product
        })

@app.route('/api/quotes/history', methods=['GET'])
def quote_history():
    mill = request.args.get('mill')
    product = request.args.get('product')
    days = int(request.args.get('days', 90))
    cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    conn = get_db()
    conditions = ["date >= ?"]
    params = [cutoff]
    if mill:
        conditions.append("mill_name=?")
        params.append(mill)
    if product:
        conditions.append("product=?")
        params.append(product)
    rows = conn.execute(
        f"SELECT * FROM mill_quotes WHERE {' AND '.join(conditions)} ORDER BY date ASC, created_at ASC",
        params
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ----- INTELLIGENCE ENGINE -----

@app.route('/api/intel/signals', methods=['GET'])
def intel_signals():
    product_filter = request.args.get('product')
    conn = get_db()
    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    d7 = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    d14 = (now - timedelta(days=14)).strftime('%Y-%m-%d')
    d30 = (now - timedelta(days=30)).strftime('%Y-%m-%d')

    # Get all products
    if product_filter:
        products = [product_filter]
    else:
        products = [r['product'] for r in conn.execute("SELECT DISTINCT product FROM mill_quotes").fetchall()]

    all_signals = {}
    for product in products:
        signals = []

        # 1. Supply Pressure
        mills_7d = conn.execute(
            "SELECT COUNT(DISTINCT mill_name) as cnt, SUM(volume) as vol FROM mill_quotes WHERE product=? AND date>=?",
            (product, d7)
        ).fetchone()
        mills_30d = conn.execute(
            "SELECT COUNT(DISTINCT mill_name) as cnt, SUM(volume) as vol, COUNT(*) as quotes FROM mill_quotes WHERE product=? AND date>=?",
            (product, d30)
        ).fetchone()
        m7 = mills_7d['cnt'] or 0
        v7 = mills_7d['vol'] or 0
        m30 = mills_30d['cnt'] or 0
        v30 = mills_30d['vol'] or 0
        q30 = mills_30d['quotes'] or 0
        avg_weekly_mills = m30 / 4.3 if m30 else 0
        avg_weekly_vol = v30 / 4.3 if v30 else 0

        if avg_weekly_mills > 0:
            mill_ratio = m7 / avg_weekly_mills
            vol_ratio = v7 / avg_weekly_vol if avg_weekly_vol > 0 else 1
            if mill_ratio > 1.2 or vol_ratio > 1.3:
                direction = 'bearish'
                strength = 'strong' if mill_ratio > 1.5 or vol_ratio > 1.5 else 'moderate'
            elif mill_ratio < 0.8 or vol_ratio < 0.7:
                direction = 'bullish'
                strength = 'strong' if mill_ratio < 0.5 else 'moderate'
            else:
                direction = 'neutral'
                strength = 'weak'
            signals.append({
                'signal': 'supply_pressure',
                'mills_offering_7d': m7,
                'volume_7d': round(v7, 1),
                'mills_avg_weekly': round(avg_weekly_mills, 1),
                'volume_avg_weekly': round(avg_weekly_vol, 1),
                'direction': direction,
                'strength': strength,
                'explanation': f"{m7} mills offering {product} this week ({round(v7)} MBF), vs {round(avg_weekly_mills,1)} mills avg. {'More supply = potential to short.' if direction=='bearish' else 'Tighter supply = consider buying.' if direction=='bullish' else 'Supply steady.'}"
            })

        # 2. Price Momentum
        prices_14d = conn.execute(
            "SELECT date, AVG(price) as avg_price FROM mill_quotes WHERE product=? AND date>=? GROUP BY date ORDER BY date",
            (product, d14)
        ).fetchall()
        prices_30d = conn.execute(
            "SELECT date, AVG(price) as avg_price FROM mill_quotes WHERE product=? AND date>=? GROUP BY date ORDER BY date",
            (product, d30)
        ).fetchall()

        def calc_slope(prices):
            if len(prices) < 2:
                return 0
            n = len(prices)
            x_sum = n * (n - 1) / 2
            x2_sum = n * (n - 1) * (2 * n - 1) / 6
            y_sum = sum(p['avg_price'] for p in prices)
            xy_sum = sum(i * p['avg_price'] for i, p in enumerate(prices))
            denom = n * x2_sum - x_sum * x_sum
            if denom == 0:
                return 0
            return (n * xy_sum - x_sum * y_sum) / denom

        slope_14d = calc_slope(prices_14d)
        slope_30d = calc_slope(prices_30d)
        current_avg = prices_14d[-1]['avg_price'] if prices_14d else 0

        if abs(slope_14d) > 0.5:
            direction = 'bullish' if slope_14d > 0 else 'bearish'
            strength = 'strong' if abs(slope_14d) > 2 else 'moderate'
        else:
            direction = 'neutral'
            strength = 'weak'
        signals.append({
            'signal': 'price_momentum',
            'current_avg': round(current_avg, 2),
            'slope_14d': round(slope_14d, 2),
            'slope_30d': round(slope_30d, 2),
            'direction': direction,
            'strength': strength,
            'explanation': f"{product} avg ${round(current_avg)}. Price {'rising' if slope_14d > 0 else 'falling'} ~${abs(round(slope_14d, 1))}/day over 14d. {'Buy before prices climb higher.' if direction=='bullish' else 'Prices softening — wait or short.' if direction=='bearish' else 'Prices stable.'}"
        })

        # 3. Print vs Street
        latest_rl = conn.execute(
            "SELECT price FROM rl_prices WHERE product=? ORDER BY date DESC LIMIT 1",
            (product,)
        ).fetchone()
        if latest_rl and current_avg > 0:
            rl_price = latest_rl['price']
            gap = rl_price - current_avg
            if gap > 10:
                direction = 'bearish'
                explanation = f"Mills offering {product} ${round(gap)} below RL print (${round(rl_price)}). Street is cheaper than print — market may be weaker than published."
            elif gap < -10:
                direction = 'bullish'
                explanation = f"Mills charging ${round(abs(gap))} above RL print (${round(rl_price)}). Genuine tightness — print hasn't caught up."
            else:
                direction = 'neutral'
                explanation = f"Mill prices tracking close to RL print (${round(rl_price)} vs ${round(current_avg)} street). Print and street aligned."
            signals.append({
                'signal': 'print_vs_street',
                'rl_price': round(rl_price, 2),
                'avg_street': round(current_avg, 2),
                'gap': round(gap, 2),
                'direction': direction,
                'strength': 'strong' if abs(gap) > 20 else 'moderate' if abs(gap) > 10 else 'weak',
                'explanation': explanation
            })

        # 4. Regional Arbitrage
        regional_prices = conn.execute("""
            SELECT m.region, MIN(mq.price) as best_price, mq.mill_name
            FROM mill_quotes mq JOIN mills m ON mq.mill_id = m.id
            WHERE mq.product=? AND mq.date>=?
            GROUP BY m.region
        """, (product, d7)).fetchall()
        if len(regional_prices) >= 2:
            rp = {r['region']: {'price': r['best_price'], 'mill': r['mill_name']} for r in regional_prices}
            opps = []
            regions = list(rp.keys())
            for i in range(len(regions)):
                for j in range(i+1, len(regions)):
                    spread = abs(rp[regions[i]]['price'] - rp[regions[j]]['price'])
                    if spread > 10:
                        cheaper = regions[i] if rp[regions[i]]['price'] < rp[regions[j]]['price'] else regions[j]
                        opps.append({
                            'from_region': cheaper,
                            'to_region': regions[j] if cheaper == regions[i] else regions[i],
                            'spread': round(spread, 2),
                            'cheaper_mill': rp[cheaper]['mill'],
                            'cheaper_price': rp[cheaper]['price']
                        })
            if opps:
                signals.append({
                    'signal': 'regional_arbitrage',
                    'opportunities': opps,
                    'direction': 'opportunity',
                    'strength': 'strong' if any(o['spread'] > 25 for o in opps) else 'moderate',
                    'explanation': f"Regional spread on {product}: " + ', '.join(
                        f"${o['spread']} between {o['from_region']} and {o['to_region']} ({o['cheaper_mill']} at ${o['cheaper_price']})"
                        for o in opps[:3]
                    )
                })

        # 5. Offering Velocity
        daily_counts = conn.execute(
            "SELECT date, COUNT(*) as cnt FROM mill_quotes WHERE product=? AND date>=? GROUP BY date",
            (product, d30)
        ).fetchall()
        if daily_counts:
            avg_daily = sum(r['cnt'] for r in daily_counts) / max(len(daily_counts), 1)
            recent_daily = [r['cnt'] for r in daily_counts if r['date'] >= d7]
            recent_avg = sum(recent_daily) / max(len(recent_daily), 1) if recent_daily else 0
            vel_ratio = recent_avg / avg_daily if avg_daily > 0 else 1
            if vel_ratio > 1.3:
                direction = 'bearish'
                explanation = f"Above-average quoting activity on {product} ({round(recent_avg,1)} vs {round(avg_daily,1)} daily avg). Mills actively pushing inventory."
            elif vel_ratio < 0.7:
                direction = 'bullish'
                explanation = f"Below-average activity on {product} ({round(recent_avg,1)} vs {round(avg_daily,1)} daily avg). Quiet market suggests tightening."
            else:
                direction = 'neutral'
                explanation = f"Normal quoting velocity on {product}."
            signals.append({
                'signal': 'offering_velocity',
                'recent_avg_daily': round(recent_avg, 1),
                'avg_daily_30d': round(avg_daily, 1),
                'velocity_ratio': round(vel_ratio, 2),
                'direction': direction,
                'strength': 'strong' if abs(vel_ratio - 1) > 0.5 else 'moderate' if abs(vel_ratio - 1) > 0.3 else 'weak',
                'explanation': explanation
            })

        # 6. Volume Trend
        weekly_vol = conn.execute("""
            SELECT strftime('%%W', date) as week, SUM(volume) as vol
            FROM mill_quotes WHERE product=? AND date>=? AND volume > 0
            GROUP BY week ORDER BY week
        """, (product, d30)).fetchall()
        if len(weekly_vol) >= 2:
            vols = [r['vol'] for r in weekly_vol]
            avg_vol = sum(vols) / len(vols)
            latest_vol = vols[-1]
            change = ((latest_vol - avg_vol) / avg_vol * 100) if avg_vol > 0 else 0
            if change > 15:
                direction = 'bearish'
            elif change < -15:
                direction = 'bullish'
            else:
                direction = 'neutral'
            signals.append({
                'signal': 'volume_trend',
                'latest_week_mbf': round(latest_vol, 1),
                'avg_week_mbf': round(avg_vol, 1),
                'change_pct': round(change, 1),
                'direction': direction,
                'strength': 'strong' if abs(change) > 30 else 'moderate' if abs(change) > 15 else 'weak',
                'explanation': f"{product} volume {'up' if change > 0 else 'down'} {round(abs(change))}% vs avg ({round(latest_vol)} MBF this week vs {round(avg_vol)} avg). {'More wood available — bearish.' if direction=='bearish' else 'Less wood showing up — bullish.' if direction=='bullish' else 'Steady volume.'}"
            })

        all_signals[product] = signals

    conn.close()
    return jsonify(all_signals)

@app.route('/api/intel/recommendations', methods=['GET'])
def intel_recommendations():
    # Get signals first
    product_filter = request.args.get('product')
    conn = get_db()
    if product_filter:
        products = [product_filter]
    else:
        products = [r['product'] for r in conn.execute("SELECT DISTINCT product FROM mill_quotes").fetchall()]

    # Use the signals endpoint logic internally
    import json as json_mod
    with app.test_request_context(f'/api/intel/signals{"?product=" + product_filter if product_filter else ""}'):
        sig_response = intel_signals()
        all_signals = json_mod.loads(sig_response.get_data())

    recommendations = []
    for product in products:
        signals = all_signals.get(product, [])
        score = 0
        reasons = []

        weights = {
            'supply_pressure': 2,
            'price_momentum': 3,
            'print_vs_street': 1.5,
            'offering_velocity': 1,
            'volume_trend': 1.5,
            'regional_arbitrage': 0  # informational, not directional
        }

        for sig in signals:
            w = weights.get(sig['signal'], 1)
            if sig['direction'] == 'bullish':
                score += w * (2 if sig['strength'] == 'strong' else 1)
            elif sig['direction'] == 'bearish':
                score -= w * (2 if sig['strength'] == 'strong' else 1)
            if sig.get('explanation'):
                reasons.append(sig['explanation'])

        if score >= 4:
            action = 'BUY NOW'
        elif score >= 2:
            action = 'LEAN BUY'
        elif score <= -4:
            action = 'SHORT / SELL'
        elif score <= -2:
            action = 'LEAN SHORT'
        else:
            action = 'HOLD / NEUTRAL'

        # Suggested margin based on market conditions
        if score >= 4:
            margin_range = [35, 50]
        elif score >= 2:
            margin_range = [28, 40]
        elif score <= -4:
            margin_range = [15, 22]
        elif score <= -2:
            margin_range = [18, 28]
        else:
            margin_range = [22, 35]

        # Best source
        best = conn.execute("""
            SELECT mq.mill_name, mq.price, m.city, m.region
            FROM mill_quotes mq JOIN mills m ON mq.mill_id = m.id
            WHERE mq.product=? AND mq.date >= ?
            ORDER BY mq.price ASC LIMIT 1
        """, (product, (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))).fetchone()

        recommendations.append({
            'product': product,
            'action': action,
            'score': round(score, 1),
            'confidence': min(abs(score) / 8, 1.0),
            'margin_range': margin_range,
            'best_source': dict(best) if best else None,
            'reasons': reasons,
            'signal_count': len(signals)
        })

    conn.close()
    return jsonify(sorted(recommendations, key=lambda r: abs(r['score']), reverse=True))

# ----- SIGNAL TRENDS -----

@app.route('/api/intel/trends', methods=['GET'])
def intel_trends():
    """Daily avg price + mill count + volume per product over time for trend charts."""
    product_filter = request.args.get('product')
    days = int(request.args.get('days', 90))
    conn = get_db()
    since = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

    sql = """
        SELECT date, product,
               ROUND(AVG(price),1) as avg_price,
               ROUND(MIN(price),1) as min_price,
               ROUND(MAX(price),1) as max_price,
               COUNT(DISTINCT mill_name) as mill_count,
               ROUND(SUM(COALESCE(volume,0)),1) as total_volume,
               COUNT(*) as quote_count
        FROM mill_quotes
        WHERE date >= ?
    """
    params = [since]
    if product_filter:
        sql += " AND product = ?"
        params.append(product_filter)
    sql += " GROUP BY date, product ORDER BY date"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    # Group by product
    trends = {}
    for r in rows:
        p = r['product']
        if p not in trends:
            trends[p] = []
        trends[p].append({
            'date': r['date'],
            'avg_price': r['avg_price'],
            'min_price': r['min_price'],
            'max_price': r['max_price'],
            'mill_count': r['mill_count'],
            'volume': r['total_volume'],
            'quotes': r['quote_count']
        })

    return jsonify(trends)


# ----- FILE PARSING -----

@app.route('/api/parse-excel', methods=['POST'])
def parse_excel():
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        all_sheets = []
        all_rows = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheet_rows = []
            for row in ws.iter_rows(values_only=True):
                sheet_rows.append([str(cell) if cell is not None else '' for cell in row])
            all_sheets.append({'name': name, 'rows': sheet_rows, 'count': len(sheet_rows)})
            # Add sheet header + rows to combined output
            all_rows.append([f'=== SHEET: {name} ==='])
            all_rows.extend(sheet_rows)
            all_rows.append([])  # blank separator
        wb.close()
        return jsonify({
            'rows': all_rows,
            'count': len(all_rows),
            'sheets': all_sheets,
            'sheet_count': len(all_sheets)
        })
    except Exception as e:
        return jsonify({'error': f'Failed to parse Excel: {str(e)}'}), 400

@app.route('/api/parse-pdf', methods=['POST'])
def parse_pdf():
    try:
        import pdfplumber
    except ImportError:
        return jsonify({'error': 'pdfplumber not installed'}), 500
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        pages_text = []
        tables = []
        with pdfplumber.open(tmp_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ''
                if text.strip():
                    pages_text.append(text)
                page_tables = page.extract_tables()
                for table in page_tables:
                    cleaned = []
                    for row in table:
                        cleaned.append([str(cell).strip() if cell else '' for cell in row])
                    if cleaned:
                        tables.append({'page': i + 1, 'rows': cleaned})
        os.unlink(tmp_path)
        return jsonify({
            'text': '\n\n'.join(pages_text),
            'tables': tables,
            'pages': len(pages_text),
            'table_count': len(tables)
        })
    except Exception as e:
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass
        return jsonify({'error': f'Failed to parse PDF: {str(e)}'}), 400

# ----- GEOCODING & MILEAGE -----

@app.route('/api/geocode', methods=['POST'])
def geocode():
    data = request.json
    location = data.get('location', '')
    if not location:
        return jsonify({'error': 'Missing location'}), 400
    coords = geocode_location(location)
    if coords:
        return jsonify(coords)
    return jsonify({'error': f'Could not geocode: {location}'}), 404

@app.route('/api/mileage', methods=['POST'])
def mileage_lookup():
    data = request.json
    origin = data.get('origin', '')
    dest = data.get('dest', '')
    if not origin or not dest:
        return jsonify({'error': 'Missing origin or dest'}), 400
    # Check lane cache
    conn = get_db()
    lane = conn.execute("SELECT miles FROM lanes WHERE origin=? AND dest=?", (origin, dest)).fetchone()
    if lane:
        conn.close()
        return jsonify({'miles': lane['miles'], 'origin': origin, 'dest': dest})
    conn.close()
    origin_coords = geocode_location(origin)
    if not origin_coords:
        return jsonify({'error': f'Could not geocode origin: {origin}'}), 404
    time.sleep(0.5)
    dest_coords = geocode_location(dest)
    if not dest_coords:
        return jsonify({'error': f'Could not geocode destination: {dest}'}), 404
    miles = get_distance(origin_coords, dest_coords)
    if miles is None:
        return jsonify({'error': 'Could not calculate route'}), 404
    # Cache the lane
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO lanes (origin, dest, miles) VALUES (?,?,?)", (origin, dest, miles))
    conn.commit()
    conn.close()
    return jsonify({'miles': miles, 'origin': origin, 'dest': dest})

# ----- CUSTOMERS -----

@app.route('/api/customers', methods=['GET'])
def list_customers():
    conn = get_db()
    rows = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/customers', methods=['POST'])
def create_customer():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    dest = data.get('destination', '')
    lat, lon = data.get('lat'), data.get('lon')
    if dest and (lat is None or lon is None):
        coords = geocode_location(dest)
        if coords:
            lat, lon = coords['lat'], coords['lon']
    conn = get_db()
    conn.execute("INSERT INTO customers (name, destination, lat, lon, trader) VALUES (?,?,?,?,?)",
                 (name, dest, lat, lon, data.get('trader', '')))
    conn.commit()
    cust = conn.execute("SELECT * FROM customers WHERE id=last_insert_rowid()").fetchone()
    conn.close()
    return jsonify(dict(cust)), 201

# ----- RL PRICES -----

@app.route('/api/rl', methods=['GET'])
def list_rl():
    conn = get_db()
    rows = conn.execute("SELECT * FROM rl_prices ORDER BY date DESC LIMIT 200").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/rl', methods=['POST'])
def add_rl():
    data = request.json
    entries = data if isinstance(data, list) else [data]
    conn = get_db()
    count = 0
    for e in entries:
        if not e.get('date') or not e.get('product') or not e.get('region'):
            continue
        try:
            price = float(e['price'])
            if price <= 0:
                continue
        except (ValueError, TypeError):
            continue
        conn.execute(
            "INSERT OR REPLACE INTO rl_prices (date, product, region, price) VALUES (?,?,?,?)",
            (e['date'], e['product'], e['region'], price)
        )
        count += 1
    conn.commit()
    conn.close()
    return jsonify({'created': count}), 201

# ----- LANES -----

@app.route('/api/lanes', methods=['GET'])
def list_lanes():
    conn = get_db()
    rows = conn.execute("SELECT * FROM lanes ORDER BY origin").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/lanes', methods=['POST'])
def add_lane():
    data = request.json
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO lanes (origin, dest, miles) VALUES (?,?,?)",
                 (data['origin'], data['dest'], data['miles']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

# ----- SETTINGS -----

@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    rows = conn.execute("SELECT * FROM settings").fetchall()
    conn.close()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['PUT'])
def update_settings():
    data = request.json
    conn = get_db()
    for k, v in data.items():
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (k, str(v)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=True)
