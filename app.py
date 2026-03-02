"""
SYP Analytics - Flask Server
Handles mileage API proxy, CRM, and static file serving
"""
from flask import Flask, send_from_directory, request, jsonify, g
from flask_cors import CORS
from functools import wraps
import requests
import os
import re
import time
import sqlite3
import json
import hashlib
import secrets
from datetime import datetime, timedelta, timezone
import tempfile
import math
import gzip
import csv
import statistics
from collections import defaultdict
from entity_resolution import EntityResolver


def business_day_cutoff(biz_days):
    """Return a date string N business days ago (Mon-Fri only)."""
    d = datetime.now()
    count = 0
    while count < biz_days:
        d -= timedelta(days=1)
        if d.weekday() < 5:  # Mon=0 .. Fri=4
            count += 1
    return d.strftime('%Y-%m-%d')


try:
    import jwt
except ImportError:
    jwt = None

app = Flask(__name__, static_folder='.', static_url_path='')

# --- CORS: restrict to known origins ---
ALLOWED_ORIGINS = [
    origin.strip() for origin in
    os.environ.get('ALLOWED_ORIGINS', 'http://localhost:5000,http://localhost:5001').split(',')
]
# In production (Heroku), ALLOWED_ORIGINS env var should include the app domain
CORS(app, origins=ALLOWED_ORIGINS + ['https://trade.fctg.com'], supports_credentials=True)

# --- JWT Auth configuration ---
JWT_SECRET = os.environ.get('JWT_SECRET', secrets.token_hex(32))
JWT_EXPIRY_HOURS = int(os.environ.get('JWT_EXPIRY_HOURS', '24'))

# Trader credentials: loaded from TRADER_CREDENTIALS env var as "user:hash,user:hash"
# To generate a hash: python -c "import hashlib; print(hashlib.sha256('password'.encode()).hexdigest())"
# Default: single admin user (override via env var in production)
_DEFAULT_ADMIN_HASH = hashlib.sha256('changeme'.encode()).hexdigest()
TRADER_CREDENTIALS = {}
_cred_str = os.environ.get('TRADER_CREDENTIALS', '')
if _cred_str:
    for pair in _cred_str.split(','):
        if ':' in pair:
            user, pw_hash = pair.strip().split(':', 1)
            TRADER_CREDENTIALS[user.strip().lower()] = pw_hash.strip()
if not TRADER_CREDENTIALS:
    TRADER_CREDENTIALS['admin'] = _DEFAULT_ADMIN_HASH

# Admin users who can access destructive endpoints
ADMIN_USERS = [u.strip().lower() for u in os.environ.get('ADMIN_USERS', 'admin,ian').split(',')]

# Pricing portal login attempt tracking (simple in-memory rate limiter)
_pricing_login_attempts = {}  # ip -> {'count': int, 'lockout_until': float}
PRICING_MAX_ATTEMPTS = 5
PRICING_LOCKOUT_SECONDS = 300  # 5 minutes


def _get_token_from_request():
    """Extract JWT token from Authorization header or query param."""
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        return auth[7:]
    return request.args.get('token')


def _decode_token(token):
    """Decode and validate a JWT token. Returns payload dict or None."""
    if not jwt or not token:
        return None
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        return None


def login_required(f):
    """Decorator: require a valid JWT token for the endpoint."""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = _get_token_from_request()
        payload = _decode_token(token)
        if not payload:
            return jsonify({'error': 'Authentication required'}), 401
        g.user = payload.get('user', '')
        g.trader = payload.get('trader', '')
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Decorator: require a valid JWT token AND admin privileges."""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = _get_token_from_request()
        payload = _decode_token(token)
        if not payload:
            return jsonify({'error': 'Authentication required'}), 401
        user = payload.get('user', '').lower()
        if user not in ADMIN_USERS:
            return jsonify({'error': 'Admin privileges required'}), 403
        g.user = payload.get('user', '')
        g.trader = payload.get('trader', '')
        return f(*args, **kwargs)
    return decorated

# CRM Database Setup
CRM_DB_PATH = os.path.join(os.path.dirname(__file__), 'crm.db')

def get_crm_db():
    conn = sqlite3.connect(CRM_DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def db_execute_with_retry(conn, sql, params=(), max_retries=3):
    """Execute SQL with retry on database locked errors."""
    for attempt in range(max_retries):
        try:
            result = conn.execute(sql, params)
            return result
        except sqlite3.OperationalError as e:
            if 'database is locked' in str(e) and attempt < max_retries - 1:
                import time
                time.sleep(0.1 * (attempt + 1))
            else:
                raise

def validate_number(val, field_name, min_val=None, max_val=None, allow_none=True):
    """Validate and coerce a numeric value. Returns (value, error_msg)."""
    if val is None or val == '':
        return (None, None) if allow_none else (None, f'{field_name} is required')
    try:
        num = float(val)
    except (ValueError, TypeError):
        return None, f'{field_name} must be a number'
    if math.isnan(num) or math.isinf(num):
        return None, f'{field_name} must be a finite number'
    if min_val is not None and num < min_val:
        return None, f'{field_name} must be >= {min_val}'
    if max_val is not None and num > max_val:
        return None, f'{field_name} must be <= {max_val}'
    return num, None

def require_json():
    """Get JSON body or return 400."""
    data = request.get_json()
    if data is None:
        return None, (jsonify({'error': 'Request body must be JSON'}), 400)
    return data, None

def init_crm_db():
    conn = get_crm_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS prospects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name TEXT NOT NULL,
            contact_name TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            notes TEXT,
            status TEXT DEFAULT 'prospect' CHECK(status IN ('prospect', 'qualified', 'converted', 'lost')),
            source TEXT,
            trader TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS contact_touches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prospect_id INTEGER NOT NULL,
            touch_type TEXT CHECK(touch_type IN ('call', 'email', 'meeting', 'note')),
            notes TEXT,
            products_discussed TEXT,
            follow_up_date DATE,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS prospect_product_interest (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prospect_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            interest_level TEXT CHECK(interest_level IN ('high', 'medium', 'low')),
            volume_estimate TEXT,
            notes TEXT,
            FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_touches_prospect ON contact_touches(prospect_id);
        CREATE INDEX IF NOT EXISTS idx_touches_follow_up ON contact_touches(follow_up_date);
        CREATE INDEX IF NOT EXISTS idx_prospects_status ON prospects(status);
        CREATE INDEX IF NOT EXISTS idx_prospects_trader ON prospects(trader);

        -- Customers table (cloud-based)
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact TEXT,
            phone TEXT,
            email TEXT,
            destination TEXT,
            locations TEXT,
            notes TEXT,
            trader TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        -- Mills table (universal — used by both CRM and Mill Intel)
        CREATE TABLE IF NOT EXISTS mills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            contact TEXT,
            phone TEXT,
            email TEXT,
            location TEXT,
            city TEXT,
            state TEXT,
            region TEXT,
            lat REAL,
            lon REAL,
            products TEXT,
            notes TEXT,
            trader TEXT NOT NULL DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_customers_trader ON customers(trader);
        CREATE INDEX IF NOT EXISTS idx_mills_trader ON mills(trader);
        CREATE INDEX IF NOT EXISTS idx_mills_name ON mills(name);
        CREATE INDEX IF NOT EXISTS idx_mills_region ON mills(region);

        -- Audit trail
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL DEFAULT (datetime('now')),
            user TEXT,
            action TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            entity_id TEXT,
            entity_name TEXT,
            old_value TEXT,
            new_value TEXT,
            details TEXT,
            ip_address TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_audit_timestamp ON audit_log(timestamp);
        CREATE INDEX IF NOT EXISTS idx_audit_entity ON audit_log(entity_type, entity_id);
        CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_log(user);

        -- Trade status workflow
        CREATE TABLE IF NOT EXISTS trade_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trade_id TEXT NOT NULL,
            trade_type TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            assigned_to TEXT,
            approved_by TEXT,
            approved_at TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            notes TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_trade_status_trade ON trade_status(trade_id);
        CREATE INDEX IF NOT EXISTS idx_trade_status_status ON trade_status(status);
        CREATE INDEX IF NOT EXISTS idx_trade_status_assigned ON trade_status(assigned_to);

        -- Credit management
        CREATE TABLE IF NOT EXISTS credit_limits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_name TEXT NOT NULL UNIQUE,
            credit_limit REAL DEFAULT 0,
            current_exposure REAL DEFAULT 0,
            payment_terms TEXT DEFAULT 'Net 30',
            last_payment_date TEXT,
            notes TEXT,
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE INDEX IF NOT EXISTS idx_credit_customer ON credit_limits(customer_name);

        -- Offering profiles: configured customer offering preferences
        CREATE TABLE IF NOT EXISTS offering_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            customer_name TEXT NOT NULL,
            destination TEXT NOT NULL,
            products TEXT NOT NULL,
            margin_target REAL DEFAULT 25,
            frequency TEXT DEFAULT 'weekly',
            preferred_mills TEXT DEFAULT '',
            day_of_week INTEGER DEFAULT 1,
            active INTEGER DEFAULT 1,
            notes TEXT DEFAULT '',
            trader TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE INDEX IF NOT EXISTS idx_offering_profiles_customer ON offering_profiles(customer_id);
        CREATE INDEX IF NOT EXISTS idx_offering_profiles_trader ON offering_profiles(trader);

        -- Offerings: generated draft offerings for review/approval
        CREATE TABLE IF NOT EXISTS offerings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL,
            customer_id INTEGER NOT NULL,
            customer_name TEXT NOT NULL,
            destination TEXT NOT NULL,
            status TEXT DEFAULT 'draft',
            products TEXT NOT NULL,
            margin_target REAL,
            total_margin REAL,
            generated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            approved_at DATETIME,
            approved_by TEXT,
            sent_at DATETIME,
            expires_at DATETIME,
            edit_notes TEXT DEFAULT '',
            trader TEXT NOT NULL,
            FOREIGN KEY (profile_id) REFERENCES offering_profiles(id),
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE INDEX IF NOT EXISTS idx_offerings_status ON offerings(status);
        CREATE INDEX IF NOT EXISTS idx_offerings_customer ON offerings(customer_id);
        CREATE INDEX IF NOT EXISTS idx_offerings_trader ON offerings(trader);
        CREATE INDEX IF NOT EXISTS idx_offerings_profile ON offerings(profile_id);
    ''')
    # Add locations column if missing (migration)
    try:
        conn.execute("ALTER TABLE mills ADD COLUMN locations TEXT DEFAULT '[]'")
    except:
        pass

    # ── Entity Resolution tables (migration-safe) ──────────────────
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS entity_canonical (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('mill','customer')),
            canonical_name TEXT NOT NULL,
            canonical_id TEXT NOT NULL UNIQUE,
            normalized_key TEXT,
            metadata TEXT DEFAULT '{}',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_ec_type_key ON entity_canonical(type, normalized_key);

        CREATE TABLE IF NOT EXISTS entity_alias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            canonical_id TEXT NOT NULL REFERENCES entity_canonical(canonical_id),
            variant TEXT NOT NULL,
            variant_normalized TEXT,
            source TEXT DEFAULT 'manual',
            score REAL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(canonical_id, variant)
        );
        CREATE INDEX IF NOT EXISTS idx_ea_variant ON entity_alias(variant_normalized);
        CREATE INDEX IF NOT EXISTS idx_ea_canonical ON entity_alias(canonical_id);

        CREATE TABLE IF NOT EXISTS entity_review (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            input_name TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            candidates TEXT,
            resolved_id TEXT,
            source_context TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    # Add canonical_id columns to existing tables (migration-safe)
    for tbl in ('customers', 'mills'):
        try:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN canonical_id TEXT")
        except:
            pass
    conn.commit()
    conn.close()

# Initialize CRM database on startup
init_crm_db()

def find_or_create_crm_mill(name, city='', state='', region='', lat=None, lon=None, trader=''):
    """Find mill by company name, or create. Adds location to locations array if new."""
    company = extract_company_name(name)
    conn = get_crm_db()
    try:
        # Look up by company name (case-insensitive)
        mill = conn.execute("SELECT * FROM mills WHERE UPPER(name)=?", (company.upper(),)).fetchone()

        city_clean = city.split(',')[0].strip() if city else ''
        if not state and city:
            state = mi_extract_state(city)
        if not region and state:
            region = MI_STATE_REGIONS.get(state.upper(), 'central')

        if mill:
            # Add location to locations array if not already present
            try:
                locations = json.loads(mill['locations'] or '[]')
            except (json.JSONDecodeError, TypeError):
                locations = []
            loc_exists = any(
                l.get('city', '').lower() == city_clean.lower() and l.get('state', '').upper() == (state or '').upper()
                for l in locations
            ) if city_clean else True  # Skip if no city

            updates = []
            vals = []
            if city_clean and not loc_exists:
                locations.append({
                    'city': city_clean, 'state': state or '',
                    'lat': lat, 'lon': lon, 'name': name
                })
                updates.append("locations=?"); vals.append(json.dumps(locations))
            # Update primary geo if missing
            if not mill['city'] and city_clean:
                updates.append("city=?"); vals.append(city_clean)
            if not mill['state'] and state:
                updates.append("state=?"); vals.append(state)
            if not mill['region'] and region:
                updates.append("region=?"); vals.append(region)
            if mill['lat'] is None and lat is not None:
                updates.append("lat=?"); vals.append(lat)
            if mill['lon'] is None and lon is not None:
                updates.append("lon=?"); vals.append(lon)
            if updates:
                vals.append(mill['id'])
                conn.execute(f"UPDATE mills SET {', '.join(updates)}, updated_at=CURRENT_TIMESTAMP WHERE id=?", vals)
                conn.commit()
                mill = conn.execute("SELECT * FROM mills WHERE id=?", (mill['id'],)).fetchone()
            return dict(mill)

        # Create new company-level mill
        location_str = f"{city_clean}, {state}".strip(', ') if city_clean or state else ''
        locations = []
        if city_clean:
            locations.append({'city': city_clean, 'state': state or '', 'lat': lat, 'lon': lon, 'name': name})

        conn.execute(
            """INSERT INTO mills (name, location, city, state, region, lat, lon, locations, products, notes, trader)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (company, location_str, city_clean, state, region, lat, lon,
             json.dumps(locations), '[]', '', trader)
        )
        conn.commit()
        mill = conn.execute("SELECT * FROM mills WHERE id=?", (conn.execute("SELECT last_insert_rowid()").fetchone()[0],)).fetchone()
        return dict(mill)
    finally:
        conn.close()

def sync_mill_to_mi(crm_mill, mi_conn=None):
    """Ensure a CRM mill exists in the MI mills table (for JOINs). Uses same ID."""
    own_conn = mi_conn is None
    conn = mi_conn or get_mi_db()
    existing = conn.execute("SELECT id FROM mills WHERE id=?", (crm_mill['id'],)).fetchone()
    if not existing:
        conn.execute(
            "INSERT OR REPLACE INTO mills (id, name, city, state, lat, lon, region, locations, products, notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (crm_mill['id'], crm_mill['name'], crm_mill.get('city', ''), crm_mill.get('state', ''),
             crm_mill.get('lat'), crm_mill.get('lon'), crm_mill.get('region', ''),
             crm_mill.get('locations', '[]'), crm_mill.get('products', '[]'), crm_mill.get('notes', ''))
        )
    else:
        conn.execute(
            "UPDATE mills SET name=?, city=?, state=?, lat=?, lon=?, region=?, locations=?, products=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (crm_mill['name'], crm_mill.get('city', ''), crm_mill.get('state', ''),
             crm_mill.get('lat'), crm_mill.get('lon'), crm_mill.get('region', ''),
             crm_mill.get('locations', '[]'), crm_mill.get('products', '[]'), crm_mill.get('notes', ''),
             crm_mill['id'])
        )
    if own_conn:
        conn.commit()
        conn.close()

# ===== MILL INTEL DATABASE =====
MI_DB_PATH = os.path.join(os.path.dirname(__file__), 'mill-intel', 'mill_intel.db')

MI_STATE_REGIONS = {
    'TX':'west','LA':'west','AR':'west','OK':'west',
    'MS':'central','AL':'central','TN':'central','KY':'central','MO':'central',
    'GA':'east','FL':'east','SC':'east','NC':'east','VA':'east','WV':'east'
}

# Canonical mill directory (matches MILL_DIRECTORY in state.js)
MILL_DIRECTORY = {
    'Canfor - DeQuincy': ('DeQuincy', 'LA'), 'Canfor - Urbana': ('Urbana', 'AR'),
    'Canfor - Fulton': ('Fulton', 'AL'), 'Canfor - Axis': ('Axis', 'AL'),
    'Canfor - El Dorado': ('El Dorado', 'AR'), 'Canfor - Thomasville': ('Thomasville', 'GA'),
    'Canfor - Moultrie': ('Moultrie', 'GA'), 'Canfor - DeRidder': ('Deridder', 'LA'),
    'Canfor - Camden SC': ('Camden', 'SC'), 'Canfor - Conway': ('Conway', 'SC'),
    'Canfor - Marion': ('Marion', 'SC'), 'Canfor - Graham': ('Graham', 'NC'),
    'West Fraser - Huttig': ('Huttig', 'AR'), 'West Fraser - Leola': ('Leola', 'AR'),
    'West Fraser - Opelika': ('Opelika', 'AL'), 'West Fraser - Russellville': ('Russellville', 'AR'),
    'West Fraser - Blackshear': ('Blackshear', 'GA'), 'West Fraser - Dudley GA': ('Dudley', 'GA'), 'West Fraser - Dudley': ('Dudley', 'GA'),
    'West Fraser - Fitzgerald': ('Fitzgerald', 'GA'), 'West Fraser - New Boston': ('New Boston', 'TX'),
    'West Fraser - Henderson': ('Henderson', 'TX'), 'West Fraser - Lufkin': ('Lufkin', 'TX'),
    'West Fraser - Joyce': ('Joyce', 'LA'),
    'West Fraser - Maplesville': ('Maplesville', 'AL'), 'West Fraser - Mcdavid': ('Mcdavid', 'FL'),
    'GP - Clarendon': ('Clarendon', 'NC'), 'GP - Camden': ('Camden', 'TX'),
    'GP - Talladega': ('Talladega', 'AL'), 'GP - Frisco City': ('Frisco City', 'AL'),
    'GP - Gurdon': ('Gurdon', 'AR'), 'GP - Albany': ('Albany', 'GA'),
    'GP - Warrenton': ('Warrenton', 'GA'), 'GP - Taylorsville': ('Taylorsville', 'MS'),
    'GP - Dudley NC': ('Dudley', 'NC'), 'GP - Diboll': ('Diboll', 'TX'),
    'GP - Pineland': ('Pineland', 'TX'), 'GP - Prosperity': ('Prosperity', 'SC'),
    'GP - Rome': ('Rome', 'GA'), 'GP - Rocky Creek': ('Frisco City', 'AL'),
    'Weyerhaeuser - Dierks': ('Dierks', 'AR'), 'Weyerhaeuser - Millport': ('Millport', 'AL'),
    'Weyerhaeuser - Dodson': ('Dodson', 'LA'), 'Weyerhaeuser - Holden': ('Holden', 'LA'),
    'Weyerhaeuser - Philadelphia': ('Philadelphia', 'MS'), 'Weyerhaeuser - Bruce': ('Bruce', 'MS'),
    'Weyerhaeuser - Magnolia': ('Magnolia', 'MS'), 'Weyerhaeuser - Grifton': ('Grifton', 'NC'),
    'Weyerhaeuser - Plymouth': ('Plymouth', 'NC'), 'Weyerhaeuser - Idabel': ('Idabel', 'OK'),
    'Interfor - Monticello': ('Monticello', 'AR'), 'Interfor - Georgetown': ('Georgetown', 'SC'),
    'Interfor - Fayette': ('Fayette', 'AL'), 'Interfor - DeQuincy': ('DeQuincy', 'LA'),
    'Interfor - Preston': ('Preston', 'GA'), 'Interfor - Perry': ('Perry', 'GA'),
    'Interfor - Baxley': ('Baxley', 'GA'), 'Interfor - Swainsboro': ('Swainsboro', 'GA'),
    'Interfor - Thomaston': ('Thomaston', 'GA'), 'Interfor - Eatonton': ('Eatonton', 'GA'),
    'PotlatchDeltic - Warren': ('Warren', 'AR'), 'PotlatchDeltic - Ola': ('Ola', 'AR'),
    'PotlatchDeltic - Waldo': ('Waldo', 'AR'),
    'Rex Lumber - Bristol': ('Bristol', 'FL'), 'Rex Lumber - Graceville': ('Graceville', 'FL'),
    'Rex Lumber - Troy': ('Troy', 'AL'), 'Rex Lumber - Brookhaven': ('Brookhaven', 'MS'),
    'Tolko - Leland': ('Leland', 'MS'),
    'Idaho Forest Group - Lumberton': ('Lumberton', 'MS'),
    'Hunt Forest Products - Winnfield': ('Winnfield', 'LA'),
    'Biewer - Newton': ('Newton', 'MS'), 'Biewer - Winona': ('Winona', 'MS'),
    'Anthony Timberlands - Bearden': ('Bearden', 'AR'), 'Anthony Timberlands - Malvern': ('Malvern', 'AR'),
    'T.R. Miller - Brewton': ('Brewton', 'AL'),
    'Lincoln Lumber - Jasper': ('Jasper', 'TX'), 'Lincoln Lumber - Conroe': ('Conroe', 'TX'),
    'Barge Forest Products - Macon': ('Macon', 'MS'),
    'Scotch Lumber - Fulton': ('Fulton', 'AL'),
    'Binderholz - Live Oak': ('Live Oak', 'FL'), 'Binderholz - Enfield': ('Enfield', 'NC'),
    'Hood Industries - Beaumont': ('Beaumont', 'MS'), 'Hood Industries - Waynesboro': ('Waynesboro', 'MS'),
    'Mid-South Lumber - Booneville': ('Booneville', 'MS'),
    'Murray Lumber - Murray': ('Murray', 'KY'),
    'Langdale Forest Products - Valdosta': ('Valdosta', 'GA'),
    'LaSalle Lumber - Urania': ('Urania', 'LA'),
    'Big River Forest Products - Gloster': ('Gloster', 'MS'),
    'Big River Forest Products - Vicksburg': ('Vicksburg', 'MS'),
    'Hankins Lumber - Grenada': ('Grenada', 'MS'),
    'Westervelt Lumber - Moundville': ('Moundville', 'AL'), 'Westervelt Lumber - Tuscaloosa': ('Tuscaloosa', 'AL'),
    'Jordan Lumber - Mt. Gilead': ('Mt. Gilead', 'NC'),
    'Mid-South Lumber - Meridian': ('Meridian', 'MS'),
    'Mid-South Lumber': ('Booneville', 'MS'),
    'WM Sheppard Lumber - Brooklet': ('Brooklet', 'GA'),
    'Harrigan Lumber': ('Monroeville', 'AL'),
    'Harrigan Lumber - Monroeville': ('Monroeville', 'AL'),
    'Big River Forest Products - Unknown': ('Gloster', 'MS'),
    'Jordan Lumber': ('Mt. Gilead', 'NC'),
    'Roseburg Forest Products - Weldon': ('Weldon', 'NC'),
    'Langdale Forest Products - Barnesville': ('Barnesville', 'GA'),
    'Two Rivers Lumber': ('Troy', 'AL'),
}

# Company alias mapping (mirrors _MILL_COMPANY_ALIASES in state.js)
MILL_COMPANY_ALIASES = {
    # Canfor
    'canfor': 'Canfor', 'canfor southern pine': 'Canfor', 'canfor southern pine inc': 'Canfor',
    'canfor southern': 'Canfor', 'csp': 'Canfor',
    # West Fraser
    'west fraser': 'West Fraser', 'wf': 'West Fraser', 'west fraser inc': 'West Fraser',
    'west fraser timber': 'West Fraser',
    # Georgia-Pacific
    'georgia-pacific': 'GP', 'georgia pacific': 'GP', 'gp': 'GP', 'georgia pacific llc': 'GP',
    # Weyerhaeuser
    'weyerhaeuser': 'Weyerhaeuser', 'wey': 'Weyerhaeuser', 'weyer': 'Weyerhaeuser',
    'weyerhaeuser company': 'Weyerhaeuser', 'weyerhaeuser nr company': 'Weyerhaeuser',
    # Interfor
    'interfor': 'Interfor', 'interfor pacific': 'Interfor', 'interfor pacific inc': 'Interfor',
    'interfor pacific, inc.': 'Interfor',
    # PotlatchDeltic
    'potlatchdeltic': 'PotlatchDeltic', 'potlatch': 'PotlatchDeltic', 'potlatch deltic': 'PotlatchDeltic',
    'pld': 'PotlatchDeltic', 'pd': 'PotlatchDeltic', 'potlatchdeltic ola': 'PotlatchDeltic',
    'potlatch - warren': 'PotlatchDeltic', 'potlatchdeltic waldo': 'PotlatchDeltic',
    'potlatchdeltic - waldo': 'PotlatchDeltic', 'potlatchdeltic - warren': 'PotlatchDeltic',
    'potlatchdeltic - ola': 'PotlatchDeltic', 'waldo': 'PotlatchDeltic',
    # Rex Lumber
    'rex': 'Rex Lumber', 'rex lumber': 'Rex Lumber', 'rex lumber bristol llc': 'Rex Lumber',
    'rex lumber bristol': 'Rex Lumber',
    # Tolko
    'tolko': 'Tolko',
    # Idaho Forest Group
    'idaho forest group': 'Idaho Forest Group', 'ifg': 'Idaho Forest Group',
    'idaho forest': 'Idaho Forest Group', 'ida forest': 'Idaho Forest Group',
    'lumberton': 'Idaho Forest Group', 'lumberton lumber': 'Idaho Forest Group',
    'idaho forest group - lumberton': 'Idaho Forest Group',
    # Hunt Forest Products
    'hunt': 'Hunt Forest Products', 'hunt forest': 'Hunt Forest Products',
    'hunt forest products': 'Hunt Forest Products',
    # Biewer
    'biewer': 'Biewer', 'biewer lumber': 'Biewer',
    # Anthony Timberlands
    'anthony': 'Anthony Timberlands', 'anthony timberlands': 'Anthony Timberlands',
    'anthony timber': 'Anthony Timberlands',
    # T.R. Miller
    't.r. miller': 'T.R. Miller', 'tr miller': 'T.R. Miller', 't r miller': 'T.R. Miller',
    # Lincoln Lumber
    'lincoln': 'Lincoln Lumber', 'lincoln lumber': 'Lincoln Lumber',
    # Barge Forest Products
    'barge': 'Barge Forest Products', 'barge forest': 'Barge Forest Products',
    'barge forest products': 'Barge Forest Products',
    # Scotch Lumber
    'scotch': 'Scotch Lumber', 'scotch lumber': 'Scotch Lumber',
    # Binderholz / Klausner
    'klausner': 'Binderholz', 'klausner lumber': 'Binderholz',
    'binderholz': 'Binderholz', 'binderholz timber': 'Binderholz', 'binderholz timber llc': 'Binderholz',
    # Hood Industries
    'hood': 'Hood Industries', 'hood industries': 'Hood Industries',
    # Mid-South Lumber
    'mid-south': 'Mid-South Lumber', 'mid south': 'Mid-South Lumber', 'mid-south lumber': 'Mid-South Lumber',
    'mid south lumber': 'Mid-South Lumber', 'mid south lumber company': 'Mid-South Lumber',
    'midsouth': 'Mid-South Lumber', 'midsouth lumber': 'Mid-South Lumber',
    # Murray Lumber
    'murray': 'Murray Lumber', 'murray lumber': 'Murray Lumber',
    # Langdale Forest Products
    'langdale': 'Langdale Forest Products', 'langdale forest': 'Langdale Forest Products',
    'langdale forest products': 'Langdale Forest Products',
    # LaSalle Lumber
    'lasalle': 'LaSalle Lumber', 'lasalle lumber': 'LaSalle Lumber',
    # Big River Forest Products
    'big river': 'Big River Forest Products', 'big river forest': 'Big River Forest Products',
    'big river forest products': 'Big River Forest Products',
    # Hankins Lumber
    'hankins': 'Hankins Lumber', 'hankins lumber': 'Hankins Lumber', 'harrigan lumber co': 'Harrigan Lumber',
    # Westervelt Lumber
    'westervelt': 'Westervelt Lumber', 'westervelt lumber': 'Westervelt Lumber',
    # Beasley Forest Products
    'beasley': 'Beasley Forest Products', 'beasley forest': 'Beasley Forest Products',
    'beasley forest products': 'Beasley Forest Products',
    # Charles Ingram Lumber
    'charles ingram': 'Charles Ingram Lumber', 'charles ingram lumber': 'Charles Ingram Lumber',
    'charles ingram lumber co': 'Charles Ingram Lumber',
    # Grayson Lumber
    'grayson': 'Grayson Lumber', 'grayson lumber': 'Grayson Lumber', 'grayson lumber corp': 'Grayson Lumber',
    # Great South Timber
    'great south': 'Great South Timber', 'great south timber': 'Great South Timber',
    'great south timber & lbr': 'Great South Timber',
    # Green Bay Packaging
    'green bay': 'Green Bay Packaging', 'green bay packaging': 'Green Bay Packaging',
    'green bay packaging inc': 'Green Bay Packaging', 'green bay packaging inc.': 'Green Bay Packaging',
    # Hardy Technologies (→ Idaho Forest Group)
    'hardy': 'Idaho Forest Group', 'hardy technologies': 'Idaho Forest Group',
    'hardy technologies llc': 'Idaho Forest Group',
    # Resolute FP
    'resolute': 'Resolute FP', 'resolute fp': 'Resolute FP', 'resolute fp us': 'Resolute FP',
    'resolute fp us inc': 'Resolute FP',
    # DuPont Pine Products
    'dupont pine': 'DuPont Pine Products', 'dupont pine products': 'DuPont Pine Products',
    # Two Rivers Lumber
    'two rivers': 'Two Rivers Lumber', 'two rivers lumber': 'Two Rivers Lumber',
    'two rivers lumber co llc': 'Two Rivers Lumber',
    # Vicksburg Forest Products
    'vicksburg': 'Vicksburg Forest Products', 'vicksburg forest': 'Vicksburg Forest Products',
    'vicksburg forest products': 'Vicksburg Forest Products',
    # Harrigan Lumber
    'harrigan': 'Harrigan Lumber', 'harrigan lumber': 'Harrigan Lumber',
    # WM Sheppard Lumber
    'wm sheppard': 'WM Sheppard Lumber', 'wm sheppard lumber': 'WM Sheppard Lumber',
    'wm sheppard lumber co inc': 'WM Sheppard Lumber', 'wm shepard': 'WM Sheppard Lumber',
    'wm shepard lumber': 'WM Sheppard Lumber',
    # Jordan Lumber
    'jordan': 'Jordan Lumber', 'jordan lumber': 'Jordan Lumber',
}

def extract_company_name(mill_name):
    """Extract company name from 'Company - City' format or via alias lookup.
    Handles em-dash/en-dash like the frontend normalizer does."""
    if not mill_name:
        return mill_name
    name = mill_name.strip()
    # Direct "Company - City" format (handle regular dash, en-dash, em-dash)
    if ' - ' in name:
        return name.split(' - ')[0].strip()
    if ' \u2013 ' in name:  # en-dash
        return name.split(' \u2013 ')[0].strip()
    if ' \u2014 ' in name:  # em-dash
        return name.split(' \u2014 ')[0].strip()
    # Alias lookup (longest-first for greedy match)
    # Normalize dashes/underscores to spaces (mirrors frontend behavior)
    lower = re.sub(r'[_\-\u2013\u2014]+', ' ', name.lower()).strip()
    lower = re.sub(r'\s+', ' ', lower)
    for alias, canonical in sorted(MILL_COMPANY_ALIASES.items(), key=lambda x: -len(x[0])):
        if lower == alias:
            return canonical
    # Partial prefix match (require word boundary to avoid false positives)
    for alias, canonical in sorted(MILL_COMPANY_ALIASES.items(), key=lambda x: -len(x[0])):
        if lower.startswith(alias + ' '):
            return canonical
    return name

def normalize_mill_name(name):
    """Title-case mill names while preserving known abbreviations."""
    if not name:
        return name
    # Known abbreviations to preserve uppercase
    _PRESERVE = {'GP', 'WM', 'LLC', 'INC', 'CO', 'LP', 'LTD', 'FP', 'TEX', 'SC', 'NC', 'AL', 'GA', 'AR', 'MS', 'FL', 'TX', 'LA', 'OK', 'VA'}
    words = name.strip().split()
    result = []
    for w in words:
        upper = w.upper().rstrip('.,')
        if upper in _PRESERVE:
            result.append(w.upper().rstrip('.,') + w[len(upper):])  # Keep trailing punctuation
        else:
            result.append(w.capitalize())
    return ' '.join(result)


# Customer alias mapping (mirrors _CUSTOMER_ALIASES in state.js)
CUSTOMER_ALIASES = {
    'power truss': 'Power Truss and Lumber', 'power truss and lumber': 'Power Truss and Lumber',
    'power truss & lumber': 'Power Truss and Lumber',
    'protec panel and truss': 'ProTec Panel and Truss', 'protec panel & truss': 'ProTec Panel and Truss',
    'craters and freighters': 'Craters & Freighters', 'craters & freighters': 'Craters & Freighters',
    'craters & freighters (stl custom crating llc)': 'Craters & Freighters',
    'precision truss': 'Precision Truss & Metal', 'precision truss and metal': 'Precision Truss & Metal',
    'precision truss & metal': 'Precision Truss & Metal', 'precision truss & walls': 'Precision Truss & Metal',
    'precision truss and walls': 'Precision Truss & Metal',
    'rehkemper and sons': 'Rehkemper & Sons', 'rehkemper & sons': 'Rehkemper & Sons',
    'rehkemper sons': 'Rehkemper & Sons', 'rehkemper & son inc': 'Rehkemper & Sons',
    'rehkemper & son': 'Rehkemper & Sons', 'rehkemper and son inc': 'Rehkemper & Sons',
    'power truss inc': 'Power Truss and Lumber',
    'atwood forest prods inc': 'Atwood Forest Products Inc', 'atwood forest prods': 'Atwood Forest Products Inc',
    'g proulx bldg products': 'G Proulx Building Products',
    'jefferson home bldrs inc': 'Jefferson Home Builders Inc', 'jefferson home bldrs': 'Jefferson Home Builders Inc',
    'nvr building prdts co': 'NVR Building Products Co', 'nvr building prdts': 'NVR Building Products Co',
    'raymond bldg sy llc uslbm': 'Raymond Building Supply LLC USLBM',
}

# Common corporate suffixes for stripping during matching
_CORP_SUFFIXES_RE = re.compile(
    r'\b(inc\.?|llc\.?|co\.?|corp\.?|corporation|company|enterprises|ltd\.?|limited|'
    r'group|holdings|lumber|timber|forest products|building products|distribution|supply)\s*\.?\s*$',
    re.IGNORECASE
)

def normalize_customer_name(name):
    """Normalize a customer name by stripping common suffixes and matching aliases."""
    if not name:
        return name
    trimmed = name.strip()
    if not trimmed:
        return trimmed

    # Fetch customer names once from a single DB connection
    try:
        conn = get_crm_db()
        rows = conn.execute('SELECT DISTINCT name FROM customers').fetchall()
        conn.close()
    except Exception:
        rows = []

    # 1. Alias dictionary lookup
    lower = re.sub(r'[_\-\u2013\u2014]+', ' ', trimmed.lower()).strip()
    lower = re.sub(r'\s+', ' ', lower)
    sorted_aliases = sorted(CUSTOMER_ALIASES.items(), key=lambda x: -len(x[0]))
    for alias, canonical in sorted_aliases:
        if lower == alias:
            # Check if an existing customer maps to this same canonical form
            for row in rows:
                if row['name']:
                    row_lower = re.sub(r'[_\-\u2013\u2014]+', ' ', row['name'].lower()).strip()
                    row_lower = re.sub(r'\s+', ' ', row_lower)
                    for a2, can2 in sorted_aliases:
                        if row_lower == a2 and can2 == canonical:
                            return row['name']
            return canonical

    # 2. Fuzzy match: normalize "&" <-> "and" and check existing customers
    fuzzy_lower = re.sub(r'\s*&\s*', ' and ', lower)
    fuzzy_lower = re.sub(r'\s+', ' ', fuzzy_lower).strip()
    for row in rows:
        if row['name']:
            row_fuzzy = re.sub(r'[_\-\u2013\u2014]+', ' ', row['name'].lower()).strip()
            row_fuzzy = re.sub(r'\s*&\s*', ' and ', row_fuzzy)
            row_fuzzy = re.sub(r'\s+', ' ', row_fuzzy).strip()
            if fuzzy_lower == row_fuzzy:
                return row['name']

    # 3. Check existing customers in DB for suffix-stripped match
    stripped = _CORP_SUFFIXES_RE.sub('', lower).strip()
    if stripped:
        for row in rows:
            if row['name']:
                db_stripped = _CORP_SUFFIXES_RE.sub('', row['name'].lower()).strip()
                if stripped == db_stripped:
                    return row['name']

    # 4. No match - return trimmed original
    return trimmed

def get_mi_db():
    conn = sqlite3.connect(MI_DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_mi_db():
    conn = get_mi_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS mills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            city TEXT,
            state TEXT,
            lat REAL,
            lon REAL,
            region TEXT,
            products TEXT DEFAULT '[]',
            notes TEXT DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_mills_name ON mills(name);
        CREATE INDEX IF NOT EXISTS idx_mills_region ON mills(region);

        CREATE TABLE IF NOT EXISTS mill_quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mill_id INTEGER NOT NULL,
            mill_name TEXT NOT NULL,
            product TEXT NOT NULL,
            price REAL NOT NULL,
            length TEXT DEFAULT 'RL',
            volume REAL DEFAULT 0,
            tls INTEGER DEFAULT 0,
            ship_window TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            date TEXT NOT NULL,
            trader TEXT NOT NULL,
            source TEXT DEFAULT 'manual',
            raw_text TEXT DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (mill_id) REFERENCES mills(id)
        );
        CREATE INDEX IF NOT EXISTS idx_mq_mill ON mill_quotes(mill_id);
        CREATE INDEX IF NOT EXISTS idx_mq_product ON mill_quotes(product);
        CREATE INDEX IF NOT EXISTS idx_mq_date ON mill_quotes(date);
        CREATE INDEX IF NOT EXISTS idx_mq_trader ON mill_quotes(trader);
        CREATE INDEX IF NOT EXISTS idx_mq_composite ON mill_quotes(mill_name, product, date);
        CREATE INDEX IF NOT EXISTS idx_mq_matrix ON mill_quotes(mill_name, product, length, id DESC);

        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            destination TEXT,
            lat REAL,
            lon REAL,
            trader TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS rl_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            region TEXT NOT NULL,
            product TEXT NOT NULL,
            length TEXT NOT NULL DEFAULT 'RL',
            price REAL NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_rl_date ON rl_prices(date);
        CREATE INDEX IF NOT EXISTS idx_rl_product_region ON rl_prices(product, region);
        CREATE UNIQUE INDEX IF NOT EXISTS idx_rl_unique ON rl_prices(date, region, product, length);

        CREATE TABLE IF NOT EXISTS lanes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            origin TEXT NOT NULL,
            dest TEXT NOT NULL,
            miles INTEGER NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(origin, dest)
        );

        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );

        CREATE TABLE IF NOT EXISTS mill_price_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mill_id INTEGER NOT NULL,
            mill_name TEXT NOT NULL,
            product TEXT NOT NULL,
            length TEXT DEFAULT 'RL',
            old_price REAL,
            new_price REAL NOT NULL,
            change REAL,
            pct_change REAL,
            date TEXT NOT NULL,
            prev_date TEXT,
            source TEXT DEFAULT 'manual',
            trader TEXT DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (mill_id) REFERENCES mills(id)
        );
        CREATE INDEX IF NOT EXISTS idx_mpc_mill_product ON mill_price_changes(mill_name, product);
        CREATE INDEX IF NOT EXISTS idx_mpc_date ON mill_price_changes(date);
    ''')
    # Add locations column if missing (migration)
    try:
        conn.execute("ALTER TABLE mills ADD COLUMN locations TEXT DEFAULT '[]'")
    except:
        pass
    # Add length column to rl_prices if missing (migration from old schema)
    try:
        conn.execute("ALTER TABLE rl_prices ADD COLUMN length TEXT NOT NULL DEFAULT 'RL'")
    except:
        pass
    # Recreate unique index to include length (drop old 3-col index, create 4-col)
    try:
        conn.execute("DROP INDEX IF EXISTS idx_rl_unique")
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_rl_unique ON rl_prices(date, region, product, length)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_rl_product_region ON rl_prices(product, region)")
    except:
        pass
    # Add canonical_id to MI tables (migration-safe)
    for tbl in ('mills', 'mill_quotes'):
        col = 'canonical_id' if tbl == 'mills' else 'canonical_mill_id'
        try:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} TEXT")
        except:
            pass
    conn.commit()
    conn.close()

init_mi_db()

# ── Entity Resolution engine ──────────────────────────────────────
_entity_resolver = EntityResolver(CRM_DB_PATH, MILL_COMPANY_ALIASES)

# Sync CRM mills → MI mills table on startup (keeps JOINs working)
def sync_crm_mills_to_mi():
    crm_conn = get_crm_db()
    crm_mills = crm_conn.execute("SELECT * FROM mills").fetchall()
    crm_conn.close()
    mi_conn = get_mi_db()
    for m in crm_mills:
        md = dict(m)
        existing = mi_conn.execute("SELECT id FROM mills WHERE id=?", (md['id'],)).fetchone()
        existing_name = mi_conn.execute("SELECT id FROM mills WHERE name=? AND id!=?", (md['name'], md['id'])).fetchone() if not existing else None
        if existing_name:
            # Name exists with different ID — delete old entry to avoid UNIQUE conflict
            mi_conn.execute("DELETE FROM mills WHERE id=?", (existing_name['id'],))
        if existing:
            mi_conn.execute(
                "UPDATE mills SET name=?, city=?, state=?, lat=?, lon=?, region=?, locations=?, products=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (md['name'], md.get('city', ''), md.get('state', ''),
                 md.get('lat'), md.get('lon'), md.get('region', ''),
                 md.get('locations', '[]'), md.get('products', '[]'), md.get('notes', ''),
                 md['id'])
            )
        else:
            mi_conn.execute(
                "INSERT INTO mills (id, name, city, state, lat, lon, region, locations, products, notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (md['id'], md['name'], md.get('city', ''), md.get('state', ''),
                 md.get('lat'), md.get('lon'), md.get('region', ''),
                 md.get('locations', '[]'), md.get('products', '[]'), md.get('notes', ''))
            )
    mi_conn.commit()
    mi_conn.close()

# Seed CRM mills from MILL_DIRECTORY, grouped by company
def seed_crm_mills():
    conn = get_crm_db()
    existing = {row['name'].upper() for row in conn.execute("SELECT name FROM mills").fetchall()}
    # Group MILL_DIRECTORY by company
    companies = {}
    for full_name, (city, state) in MILL_DIRECTORY.items():
        company = full_name.split(' - ')[0]
        if company not in companies:
            companies[company] = []
        companies[company].append({'city': city, 'state': state, 'name': full_name})
    added = 0
    for company, locs in companies.items():
        if company.upper() not in existing:
            # Use first location as primary
            primary = locs[0]
            region = MI_STATE_REGIONS.get(primary['state'].upper(), 'central')
            location = f"{primary['city']}, {primary['state']}"
            locations_json = json.dumps([
                {'city': l['city'], 'state': l['state'], 'lat': None, 'lon': None, 'name': l['name']}
                for l in locs
            ])
            conn.execute(
                "INSERT INTO mills (name, location, city, state, region, locations, products, notes, trader) VALUES (?,?,?,?,?,?,?,?,?)",
                (company, location, primary['city'], primary['state'], region, locations_json, '[]', '', '')
            )
            added += 1
    if added:
        conn.commit()
        print(f"Seeded {added} company mills from MILL_DIRECTORY into CRM")
    conn.close()

seed_crm_mills()
sync_crm_mills_to_mi()

# Seed Mill Intel SQLite from Supabase cloud data on startup
# This ensures Railway (ephemeral filesystem) always has mill quotes after deploy
def seed_mi_from_supabase():
    supa_url = os.environ.get('SUPABASE_URL', '')
    supa_key = os.environ.get('SUPABASE_ANON_KEY', '')
    if not supa_url or not supa_key:
        print("Supabase not configured — skipping MI cloud seed")
        return

    # Only seed if mill_quotes table is empty (fresh deploy)
    mi_conn = get_mi_db()
    count = mi_conn.execute("SELECT COUNT(*) FROM mill_quotes").fetchone()[0]
    if count > 0:
        mi_conn.close()
        print(f"MI already has {count} quotes — skipping cloud seed")
        return
    mi_conn.close()

    print("MI tables empty — seeding from Supabase cloud...")
    try:
        res = requests.get(
            f"{supa_url}/rest/v1/syp_data?select=data&limit=1",
            headers={'apikey': supa_key, 'Authorization': f'Bearer {supa_key}'},
            timeout=15
        )
        if not res.ok:
            print(f"Supabase fetch failed: {res.status_code}")
            return
        rows = res.json()
        if not rows or not rows[0].get('data'):
            print("No cloud data found")
            return

        d = rows[0]['data']

        # Seed mills into CRM → MI
        cloud_mills = d.get('mills', [])
        if cloud_mills:
            crm_conn = get_crm_db()
            existing = {r['name'].upper() for r in crm_conn.execute("SELECT name FROM mills").fetchall()}
            added = 0
            for m in cloud_mills:
                name = m.get('name', '').strip()
                if not name or name.upper() in existing:
                    continue
                city = m.get('city', '')
                state = m.get('state', '')
                region = m.get('region', 'central')
                location = m.get('location', '')
                if not location and city:
                    location = f"{city}, {state}" if state else city
                locations = m.get('locations', '[]')
                if isinstance(locations, list):
                    locations = json.dumps(locations)
                products = m.get('products', '[]')
                if isinstance(products, list):
                    products = json.dumps(products)
                lat = m.get('lat')
                lon = m.get('lon')
                try:
                    crm_conn.execute(
                        "INSERT INTO mills (name, location, city, state, region, lat, lon, locations, products, notes, trader) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (name, location, city, state, region, lat, lon, locations, products, m.get('notes', ''), m.get('trader', ''))
                    )
                    existing.add(name.upper())
                    added += 1
                except Exception:
                    pass
            if added:
                crm_conn.commit()
                print(f"  Seeded {added} mills from cloud into CRM")
            crm_conn.close()
            # Re-sync CRM → MI so mill IDs are available for quotes
            sync_crm_mills_to_mi()

        # Seed mill quotes
        cloud_quotes = d.get('millQuotes', [])
        if cloud_quotes:
            mi_conn = get_mi_db()
            crm_conn = get_crm_db()
            inserted = 0
            for q in cloud_quotes:
                mill_name = (q.get('mill') or q.get('mill_name') or '').strip()
                product = (q.get('product') or '').strip()
                price = q.get('price') or q.get('fob')
                if not mill_name or not product or not price:
                    continue
                try:
                    price_val = float(price)
                    if price_val <= 0:
                        continue
                except (ValueError, TypeError):
                    continue

                # Find mill_id from CRM
                mill_row = crm_conn.execute("SELECT id FROM mills WHERE UPPER(name)=?", (mill_name.upper(),)).fetchone()
                if not mill_row:
                    # Create mill stub
                    city = q.get('city') or q.get('location') or ''
                    state_code = mi_extract_state(city) if city else ''
                    region = mi_get_region(state_code) if state_code else 'central'
                    crm_conn.execute(
                        "INSERT INTO mills (name, location, city, state, region, products, notes, trader) VALUES (?,?,?,?,?,?,?,?)",
                        (mill_name, city, city.split(',')[0].strip() if city else '', state_code, region, '[]', '', q.get('trader', ''))
                    )
                    crm_conn.commit()
                    mill_row = crm_conn.execute("SELECT id FROM mills WHERE UPPER(name)=?", (mill_name.upper(),)).fetchone()
                    # Sync new mill to MI
                    new_mill = crm_conn.execute("SELECT * FROM mills WHERE id=?", (mill_row['id'],)).fetchone()
                    sync_mill_to_mi(dict(new_mill), mi_conn=mi_conn)

                mill_id = mill_row['id']
                length = q.get('length', 'RL') or 'RL'
                date = q.get('date') or datetime.now().strftime('%Y-%m-%d')
                trader = q.get('trader', 'Unknown')
                ship_window = q.get('shipWindow') or q.get('ship_window') or q.get('ship') or ''

                mi_conn.execute(
                    """INSERT INTO mill_quotes (mill_id, mill_name, product, price, length, volume, tls,
                       ship_window, notes, date, trader, source)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (mill_id, mill_name, product, price_val, length,
                     q.get('volume', 0) or 0, q.get('tls', 0) or 0,
                     ship_window, q.get('notes', ''), date, trader, 'cloud_seed')
                )
                inserted += 1

            mi_conn.commit()
            mi_conn.close()
            crm_conn.close()
            print(f"  Seeded {inserted} mill quotes from cloud")

        # Seed customers
        cloud_customers = d.get('customers', [])
        if cloud_customers:
            crm_conn = get_crm_db()
            existing = {r['name'].upper() for r in crm_conn.execute("SELECT name FROM customers").fetchall()} if crm_conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0] > 0 else set()
            added = 0
            for c in cloud_customers:
                name = (c.get('name') or '').strip()
                if not name or name.upper() in existing:
                    continue
                dest = c.get('destination') or ''
                locations = c.get('locations', '[]')
                if isinstance(locations, list):
                    locations = json.dumps(locations)
                try:
                    crm_conn.execute(
                        "INSERT INTO customers (name, destination, locations, trader) VALUES (?,?,?,?)",
                        (name, dest, locations, c.get('trader', ''))
                    )
                    existing.add(name.upper())
                    added += 1
                except Exception:
                    pass
            if added:
                crm_conn.commit()
                print(f"  Seeded {added} customers from cloud")
            crm_conn.close()

        # Recompute price changes from the seeded quotes
        recompute_price_changes()

        print("Cloud seed complete!")
    except requests.exceptions.Timeout:
        print("Cloud seed TIMEOUT — Supabase did not respond within 15s")
    except requests.exceptions.ConnectionError:
        print("Cloud seed CONNECTION ERROR — cannot reach Supabase")
    except Exception as e:
        print(f"Cloud seed error: {type(e).__name__}: {e}")

def recompute_price_changes():
    """Recompute mill_price_changes from mill_quotes data.
    Deduplicates by taking one price per mill+product+length+date (latest entry wins),
    then generates change records wherever price differs between consecutive dates.
    Called after seed_mi_from_supabase() and syncMillQuotesToMillIntel().
    """
    conn = get_mi_db()
    try:
        # Clear existing price changes (will rebuild from scratch)
        conn.execute("DELETE FROM mill_price_changes")

        # Get deduplicated quotes: one price per mill+product+length+date (latest id wins)
        rows = conn.execute(
            """SELECT mill_id, mill_name, product, length, price, date, trader, source
               FROM mill_quotes
               WHERE id IN (
                   SELECT MAX(id) FROM mill_quotes
                   GROUP BY UPPER(mill_name), UPPER(product), UPPER(COALESCE(length,'RL')), date
               )
               ORDER BY UPPER(mill_name), UPPER(product), UPPER(COALESCE(length,'RL')), date ASC"""
        ).fetchall()

        if not rows:
            conn.commit()
            conn.close()
            return

        changes_added = 0
        prev = None
        for r in rows:
            key = (r['mill_name'].upper(), r['product'].upper(), (r['length'] or 'RL').upper())
            curr_price = r['price']

            if prev and prev['key'] == key and prev['date'] != r['date'] and abs(curr_price - prev['price']) > 0.001:
                change_val = round(curr_price - prev['price'], 2)
                pct_val = round((change_val / prev['price']) * 100, 2) if prev['price'] else None
                conn.execute(
                    """INSERT INTO mill_price_changes (mill_id, mill_name, product, length,
                       old_price, new_price, change, pct_change, date, prev_date, source, trader)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (r['mill_id'], r['mill_name'], r['product'], r['length'] or 'RL',
                     prev['price'], curr_price, change_val, pct_val,
                     r['date'], prev['date'], r['source'] or 'recompute', r['trader'] or '')
                )
                changes_added += 1

            prev = {'key': key, 'price': curr_price, 'date': r['date']}

        conn.commit()
        conn.close()
        print(f"  Recomputed {changes_added} mill price changes from {len(rows)} quotes")
    except Exception as e:
        conn.close()
        print(f"  Price change recomputation error: {e}")

def mi_extract_state(location):
    if not location:
        return None
    parts = location.strip().rstrip('.').split(',')
    if len(parts) >= 2:
        st = parts[-1].strip().upper()[:2]
        if len(st) == 2 and st.isalpha():
            return st
    return None

def mi_get_region(state_code):
    return MI_STATE_REGIONS.get(state_code, 'central')

# Now that mi_extract_state and mi_get_region are defined, run the cloud seed
seed_mi_from_supabase()

# ----- Seed rl_prices from gzipped CSV (historical Random Lengths data) -----

def seed_rl_from_csv():
    """Seed rl_prices table from data/rl_prices.csv.gz on startup if empty."""
    csv_path = os.path.join(os.path.dirname(__file__), 'data', 'rl_prices.csv.gz')
    if not os.path.exists(csv_path):
        print("RL CSV not found — skipping historical seed")
        return

    conn = get_mi_db()
    count = conn.execute("SELECT COUNT(*) FROM rl_prices").fetchone()[0]
    if count > 0:
        conn.close()
        print(f"RL already has {count} prices — skipping CSV seed")
        return
    conn.close()

    print(f"Seeding rl_prices from {csv_path}...")
    try:
        rows = []
        with gzip.open(csv_path, 'rt') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    price = float(row['price'])
                    if price <= 0:
                        continue
                    rows.append((row['date'], row['region'], row['product'], row['length'], price))
                except (ValueError, KeyError):
                    continue

        if rows:
            conn = get_mi_db()
            conn.executemany(
                "INSERT OR IGNORE INTO rl_prices (date, region, product, length, price) VALUES (?,?,?,?,?)",
                rows
            )
            conn.commit()
            conn.close()
            print(f"  Seeded {len(rows)} RL prices from CSV")
    except Exception as e:
        print(f"RL CSV seed error: {type(e).__name__}: {e}")

def seed_rl_from_supabase():
    """Backfill recent RL entries from Supabase cloud (captures weekly uploads since CSV was committed)."""
    supa_url = os.environ.get('SUPABASE_URL', '')
    supa_key = os.environ.get('SUPABASE_ANON_KEY', '')
    if not supa_url or not supa_key:
        return

    try:
        res = requests.get(
            f"{supa_url}/rest/v1/syp_data?select=data&limit=1",
            headers={'apikey': supa_key, 'Authorization': f'Bearer {supa_key}'},
            timeout=15
        )
        if not res.ok:
            return
        data = res.json()
        if not data or not data[0].get('data'):
            return

        cloud_rl = data[0]['data'].get('rl', [])
        if not cloud_rl:
            return

        rows = []
        for entry in cloud_rl:
            date = entry.get('date')
            if not date:
                continue
            # Composite prices (west/central/east → product keys like "2x4#2")
            for region in ['west', 'central', 'east']:
                for product, price in (entry.get(region) or {}).items():
                    if isinstance(price, (int, float)) and price > 0:
                        rows.append((date, region, product, 'RL', float(price)))
            # Specified lengths
            for region in ['west', 'central', 'east']:
                region_data = (entry.get('specified_lengths') or {}).get(region, {})
                for product, lengths in region_data.items():
                    if not isinstance(lengths, dict):
                        continue
                    for length, price in lengths.items():
                        if isinstance(price, (int, float)) and price > 0:
                            rows.append((date, region, product, str(length), float(price)))

        if rows:
            conn = get_mi_db()
            conn.executemany(
                "INSERT OR IGNORE INTO rl_prices (date, region, product, length, price) VALUES (?,?,?,?,?)",
                rows
            )
            conn.commit()
            inserted = conn.total_changes
            conn.close()
            if inserted:
                print(f"  Backfilled {inserted} RL prices from Supabase cloud")
    except Exception as e:
        print(f"RL Supabase seed error: {type(e).__name__}: {e}")

seed_rl_from_csv()
seed_rl_from_supabase()

def mi_geocode_location(location):
    """Geocode using shared geo_cache, with DB fallback then Nominatim."""
    if not location:
        return None
    cache_key = location.lower().strip()
    if cache_key in geo_cache:
        return geo_cache[cache_key]
    # Check DB for stored coords
    try:
        conn = get_mi_db()
        row = conn.execute("SELECT lat, lon FROM mills WHERE LOWER(city)=? AND lat IS NOT NULL", (cache_key,)).fetchone()
        conn.close()
        if row:
            coords = {'lat': row['lat'], 'lon': row['lon']}
            geo_cache[cache_key] = coords
            return coords
    except:
        pass
    return geocode_location(location)

def mi_get_distance(origin_coords, dest_coords):
    """Delegate to shared get_distance (with caching)."""
    return get_distance(origin_coords, dest_coords)

# Shared geocode + distance caches (populated from CRM mills on startup)
geo_cache = {}
distance_cache = {}

# Matrix response cache (short TTL to handle concurrent requests)
_matrix_cache = {}
_matrix_cache_ttl = 30  # seconds

def get_cached_matrix(cache_key):
    """Get cached matrix response if still valid."""
    if cache_key in _matrix_cache:
        data, timestamp = _matrix_cache[cache_key]
        if time.time() - timestamp < _matrix_cache_ttl:
            return data
        del _matrix_cache[cache_key]
    return None

def set_cached_matrix(cache_key, data):
    """Cache matrix response."""
    _matrix_cache[cache_key] = (data, time.time())
    # Cleanup old entries (keep cache size reasonable)
    if len(_matrix_cache) > 20:
        oldest_key = min(_matrix_cache.keys(), key=lambda k: _matrix_cache[k][1])
        del _matrix_cache[oldest_key]

def invalidate_matrix_cache():
    """Clear matrix cache (call when quotes are added/updated)."""
    global _matrix_cache
    _matrix_cache = {}

# RL price cache (data changes weekly, so 1-hour TTL is fine)
_rl_cache = {}
_rl_cache_ttl = 3600  # 1 hour

def get_rl_cached(cache_key):
    """Get cached RL response if still valid."""
    if cache_key in _rl_cache:
        data, timestamp = _rl_cache[cache_key]
        if time.time() - timestamp < _rl_cache_ttl:
            return data
        del _rl_cache[cache_key]
    return None

def set_rl_cache(cache_key, data):
    """Cache RL response."""
    _rl_cache[cache_key] = (data, time.time())
    if len(_rl_cache) > 50:
        oldest_key = min(_rl_cache.keys(), key=lambda k: _rl_cache[k][1])
        del _rl_cache[oldest_key]

def invalidate_rl_cache():
    """Clear RL cache (call when new RL data is saved)."""
    global _rl_cache
    _rl_cache = {}

def warm_geo_cache():
    """Pre-load geo_cache from CRM mills that have lat/lon stored."""
    try:
        conn = get_crm_db()
        rows = conn.execute("SELECT city, state, location, lat, lon FROM mills WHERE lat IS NOT NULL AND lon IS NOT NULL").fetchall()
        conn.close()
        for r in rows:
            coords = {'lat': r['lat'], 'lon': r['lon']}
            # Cache by "city, state" and by "location" field
            if r['city'] and r['state']:
                geo_cache[f"{r['city']}, {r['state']}".lower().strip()] = coords
            if r['location']:
                geo_cache[r['location'].lower().strip()] = coords
        print(f"Geo cache warmed: {len(geo_cache)} entries from CRM mills")
    except Exception as e:
        print(f"Geo cache warm failed: {e}")

warm_geo_cache()

# Serve main app
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

def geocode_location(location):
    """Convert city, state to coordinates - prefers cities over counties"""
    if not location:
        return None

    # Check cache first
    cache_key = location.lower().strip()
    if cache_key in geo_cache:
        return geo_cache[cache_key]

    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': location,
            'format': 'json',
            'limit': 5,  # Get multiple results to prefer cities over counties
            'countrycodes': 'us'
        }
        headers = {'User-Agent': 'SYP-Analytics/1.0'}

        resp = requests.get(url, params=params, headers=headers, timeout=10)
        results = resp.json()

        if results:
            # Prefer city/town/village over county - counties often have same name as cities
            city_types = ['city', 'town', 'village', 'hamlet', 'suburb', 'neighbourhood']
            best = None
            for r in results:
                if r.get('type') in city_types or r.get('addresstype') in city_types:
                    best = r
                    break
            if not best:
                best = results[0]

            coords = {
                'lat': float(best['lat']),
                'lon': float(best['lon'])
            }
            geo_cache[cache_key] = coords
            return coords
        return None
    except Exception as e:
        print(f"Geocode error for {location}: {e}")
        return None

def get_distance(origin_coords, dest_coords):
    """Get driving distance in miles between two coordinate pairs (cached)."""
    # Round coords to 3 decimals for cache key (~100m precision, plenty for cities)
    cache_key = (round(origin_coords['lat'],3), round(origin_coords['lon'],3),
                 round(dest_coords['lat'],3), round(dest_coords['lon'],3))
    if cache_key in distance_cache:
        return distance_cache[cache_key]
    try:
        coords_str = f"{origin_coords['lon']},{origin_coords['lat']};{dest_coords['lon']},{dest_coords['lat']}"
        url = f"https://router.project-osrm.org/route/v1/driving/{coords_str}"
        params = {'overview': 'false'}

        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()

        if data.get('code') == 'Ok' and data.get('routes'):
            meters = data['routes'][0]['distance']
            miles = round(meters / 1609.34)
            distance_cache[cache_key] = miles
            return miles
        return None
    except Exception as e:
        print(f"Distance error: {e}")
        return None

# Single mileage lookup
@app.route('/api/mileage', methods=['POST'])
def mileage_lookup():
    data = request.get_json() or {}
    origin = data.get('origin', '')
    dest = data.get('dest', '')
    
    if not origin or not dest:
        return jsonify({'error': 'Missing origin or dest'}), 400
    
    # Geocode both locations
    origin_coords = geocode_location(origin)
    if not origin_coords:
        return jsonify({'error': f'Could not geocode origin: {origin}'}), 404
    
    # Rate limit for Nominatim's 1 req/s policy. This blocks the worker thread;
    # acceptable for low-traffic internal tool. For high concurrency, consider
    # a token-bucket rate limiter or async approach.
    time.sleep(0.5)

    dest_coords = geocode_location(dest)
    if not dest_coords:
        return jsonify({'error': f'Could not geocode destination: {dest}'}), 404
    
    # Get distance
    miles = get_distance(origin_coords, dest_coords)
    if miles is None:
        return jsonify({'error': 'Could not calculate route'}), 404
    
    return jsonify({'miles': miles, 'origin': origin, 'dest': dest})

# Bulk mileage lookup
@app.route('/api/mileage/bulk', methods=['POST'])
def mileage_bulk():
    data = request.get_json() or {}
    lanes = data.get('lanes', [])

    if not lanes:
        return jsonify({'error': 'No lanes provided'}), 400
    if len(lanes) > 50:
        return jsonify({'error': 'Maximum 50 lanes per request'}), 400

    results = []
    # Pre-geocode shared destination (all lanes in a quote typically share the same dest)
    dest_cache = {}
    need_nominatim = False

    for lane in lanes:
        origin = lane.get('origin', '')
        dest = lane.get('dest', '')

        if not origin or not dest:
            results.append({'origin': origin, 'dest': dest, 'miles': None, 'error': 'Missing data'})
            continue

        # Geocode origin (only hit Nominatim rate limit if not cached)
        origin_cached = origin.lower().strip() in geo_cache
        if need_nominatim and not origin_cached:
            time.sleep(0.3)  # Nominatim 1 req/s; blocking is acceptable for internal tool
        origin_coords = geocode_location(origin)
        if not origin_coords:
            results.append({'origin': origin, 'dest': dest, 'miles': None, 'error': f'Could not geocode: {origin}'})
            continue
        if not origin_cached:
            need_nominatim = True

        # Geocode dest (cache across lanes — usually the same destination)
        if dest not in dest_cache:
            dest_cached = dest.lower().strip() in geo_cache
            if need_nominatim and not dest_cached:
                time.sleep(0.3)  # Nominatim 1 req/s; blocking is acceptable for internal tool
            dest_cache[dest] = geocode_location(dest)
            if not dest_cached:
                need_nominatim = True
        dest_coords = dest_cache[dest]
        if not dest_coords:
            results.append({'origin': origin, 'dest': dest, 'miles': None, 'error': f'Could not geocode: {dest}'})
            continue

        # Get distance (also cached)
        miles = get_distance(origin_coords, dest_coords)
        if miles:
            results.append({'origin': origin, 'dest': dest, 'miles': miles})
        else:
            results.append({'origin': origin, 'dest': dest, 'miles': None, 'error': 'Route not found'})

    return jsonify({'results': results})

# Geocode endpoint (for debugging)
@app.route('/api/geocode', methods=['POST'])
def geocode():
    data = request.get_json()
    location = data.get('location', '')
    
    if not location:
        return jsonify({'error': 'Missing location'}), 400
    
    coords = geocode_location(location)
    if coords:
        return jsonify(coords)
    return jsonify({'error': 'Location not found'}), 404

# ==================== CRM API ENDPOINTS ====================

# List prospects
@app.route('/api/crm/prospects', methods=['GET'])
def list_prospects():
    try:
        conn = get_crm_db()
        status = request.args.get('status')
        trader = request.args.get('trader')
        search = request.args.get('search')

        query = 'SELECT * FROM prospects WHERE 1=1'
        params = []

        if status:
            query += ' AND status = ?'
            params.append(status)
        if trader:
            query += ' AND trader = ?'
            params.append(trader)
        if search:
            query += ' AND (company_name LIKE ? OR contact_name LIKE ? OR phone LIKE ?)'
            params.extend([f'%{search}%'] * 3)

        query += ' ORDER BY updated_at DESC'
        prospects = conn.execute(query, params).fetchall()
        conn.close()
        return jsonify([dict(p) for p in prospects])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Get prospect with touches
@app.route('/api/crm/prospects/<int:id>', methods=['GET'])
def get_prospect(id):
    try:
        conn = get_crm_db()
        prospect = conn.execute('SELECT * FROM prospects WHERE id = ?', (id,)).fetchone()
        if not prospect:
            conn.close()
            return jsonify({'error': 'Prospect not found'}), 404

        touches = conn.execute(
            'SELECT * FROM contact_touches WHERE prospect_id = ? ORDER BY created_at DESC',
            (id,)
        ).fetchall()

        interests = conn.execute(
            'SELECT * FROM prospect_product_interest WHERE prospect_id = ?',
            (id,)
        ).fetchall()

        conn.close()
        result = dict(prospect)
        result['touches'] = [dict(t) for t in touches]
        result['interests'] = [dict(i) for i in interests]
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Create prospect
@app.route('/api/crm/prospects', methods=['POST'])

def create_prospect():
    try:
        data = request.get_json() or {}
        company_name = (data.get('company_name') or '').strip()
        if not company_name:
            return jsonify({'error': 'company_name is required'}), 400
        conn = get_crm_db()
        normalized_name = normalize_customer_name(company_name)

        # Duplicate detection: check if prospect with same name already exists
        existing_prospect = conn.execute(
            'SELECT * FROM prospects WHERE UPPER(company_name) = UPPER(?)',
            (normalized_name,)
        ).fetchone()
        if existing_prospect:
            conn.close()
            result = dict(existing_prospect)
            result['existing'] = True
            result['existing_type'] = 'prospect'
            return jsonify(result), 200

        # Check if company already exists as a customer
        existing_customer = conn.execute(
            'SELECT * FROM customers WHERE UPPER(name) = UPPER(?)',
            (normalized_name,)
        ).fetchone()
        if existing_customer:
            conn.close()
            return jsonify({
                'warning': 'already_customer',
                'customer': dict(existing_customer),
                'message': f'{normalized_name} already exists as a customer'
            }), 200

        cursor = conn.execute('''
            INSERT INTO prospects (company_name, contact_name, phone, email, address, notes, status, source, trader)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            normalized_name,
            data.get('contact_name'),
            data.get('phone'),
            data.get('email'),
            data.get('address'),
            data.get('notes'),
            data.get('status', 'prospect'),
            data.get('source'),
            data.get('trader')
        ))
        prospect_id = cursor.lastrowid
        conn.commit()

        prospect = conn.execute('SELECT * FROM prospects WHERE id = ?', (prospect_id,)).fetchone()
        conn.close()
        _log_audit(
            getattr(g, 'user', data.get('trader', 'unknown')),
            'prospect_create', 'prospect', prospect_id, normalized_name,
            ip_address=request.remote_addr
        )
        return jsonify(dict(prospect)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Update prospect
@app.route('/api/crm/prospects/<int:id>', methods=['PUT'])

def update_prospect(id):
    try:
        data = request.get_json() or {}
        conn = get_crm_db()
        # Only update fields that are present in the request (partial update support)
        # SECURITY: field names come from the hardcoded allowlist below, never from user input
        allowed = ['company_name', 'contact_name', 'phone', 'email', 'address', 'notes', 'status', 'source', 'trader']
        fields = [f for f in allowed if f in data]
        if not fields:
            return jsonify({'error': 'No fields to update'}), 400
        if not all(f in allowed for f in fields):
            return jsonify({'error': 'Invalid field'}), 400
        set_clause = ', '.join(f'{f} = ?' for f in fields) + ', updated_at = CURRENT_TIMESTAMP'
        values = [data[f] for f in fields] + [id]
        conn.execute(f'UPDATE prospects SET {set_clause} WHERE id = ?', values)
        conn.commit()

        prospect = conn.execute('SELECT * FROM prospects WHERE id = ?', (id,)).fetchone()
        conn.close()
        if not prospect:
            return jsonify({'error': 'Prospect not found'}), 404
        _log_audit(
            getattr(g, 'user', data.get('trader', 'unknown')),
            'prospect_update', 'prospect', id, prospect['company_name'],
            details=f'Updated fields: {", ".join(fields)}',
            ip_address=request.remote_addr
        )
        return jsonify(dict(prospect))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Delete prospect
@app.route('/api/crm/prospects/<int:id>', methods=['DELETE'])

def delete_prospect(id):
    try:
        conn = get_crm_db()
        old = conn.execute('SELECT company_name FROM prospects WHERE id = ?', (id,)).fetchone()
        conn.execute('DELETE FROM prospects WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        _log_audit(
            getattr(g, 'user', 'unknown'),
            'prospect_delete', 'prospect', id,
            old['company_name'] if old else None,
            ip_address=request.remote_addr
        )
        return jsonify({'message': 'Prospect deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Create touch
@app.route('/api/crm/touches', methods=['POST'])

def create_touch():
    try:
        data = request.get_json()
        conn = get_crm_db()
        products = json.dumps(data.get('products_discussed')) if data.get('products_discussed') else None

        cursor = conn.execute('''
            INSERT INTO contact_touches (prospect_id, touch_type, notes, products_discussed, follow_up_date)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data.get('prospect_id'),
            data.get('touch_type'),
            data.get('notes'),
            products,
            data.get('follow_up_date')
        ))
        touch_id = cursor.lastrowid

        # Update prospect's updated_at
        conn.execute('UPDATE prospects SET updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                    (data.get('prospect_id'),))
        conn.commit()

        touch = conn.execute('SELECT * FROM contact_touches WHERE id = ?', (touch_id,)).fetchone()
        conn.close()
        return jsonify(dict(touch)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# List touches (with follow-up filter)
@app.route('/api/crm/touches', methods=['GET'])
def list_touches():
    try:
        conn = get_crm_db()
        prospect_id = request.args.get('prospect_id')
        follow_up = request.args.get('follow_up')

        query = '''
            SELECT t.*, p.company_name, p.contact_name
            FROM contact_touches t
            JOIN prospects p ON t.prospect_id = p.id
            WHERE 1=1
        '''
        params = []

        if prospect_id:
            query += ' AND t.prospect_id = ?'
            params.append(prospect_id)
        if follow_up == 'today':
            query += ' AND t.follow_up_date = DATE("now")'
        elif follow_up == 'overdue':
            query += ' AND t.follow_up_date < DATE("now") AND t.follow_up_date IS NOT NULL'
        elif follow_up == 'upcoming':
            query += ' AND t.follow_up_date >= DATE("now")'

        query += ' ORDER BY t.created_at DESC LIMIT 100'
        touches = conn.execute(query, params).fetchall()
        conn.close()
        return jsonify([dict(t) for t in touches])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Add product interest
@app.route('/api/crm/prospects/<int:id>/interests', methods=['POST'])

def add_interest(id):
    try:
        data = request.get_json()
        conn = get_crm_db()

        # Check if interest already exists
        existing = conn.execute(
            'SELECT id FROM prospect_product_interest WHERE prospect_id = ? AND product = ?',
            (id, data.get('product'))
        ).fetchone()

        if existing:
            conn.execute('''
                UPDATE prospect_product_interest SET interest_level = ?, volume_estimate = ?, notes = ?
                WHERE prospect_id = ? AND product = ?
            ''', (data.get('interest_level'), data.get('volume_estimate'), data.get('notes'), id, data.get('product')))
        else:
            conn.execute('''
                INSERT INTO prospect_product_interest (prospect_id, product, interest_level, volume_estimate, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (id, data.get('product'), data.get('interest_level'), data.get('volume_estimate'), data.get('notes')))

        conn.commit()
        interests = conn.execute('SELECT * FROM prospect_product_interest WHERE prospect_id = ?', (id,)).fetchall()
        conn.close()
        return jsonify([dict(i) for i in interests])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Remove product interest
@app.route('/api/crm/interests/<int:id>', methods=['DELETE'])

def remove_interest(id):
    try:
        conn = get_crm_db()
        conn.execute('DELETE FROM prospect_product_interest WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Interest removed'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# CRM Dashboard stats with stale prospect tracking
@app.route('/api/crm/dashboard', methods=['GET'])
def crm_dashboard():
    try:
        conn = get_crm_db()
        trader = request.args.get('trader')

        # Build trader filter fragments for prospect-only and joined queries
        # Using explicit table aliases avoids fragile string replacement
        prospect_filter = ' AND trader = ?' if trader and trader != 'Admin' else ''
        joined_filter = ' AND p.trader = ?' if trader and trader != 'Admin' else ''
        params = [trader] if trader and trader != 'Admin' else []

        today = datetime.now().strftime('%Y-%m-%d')

        stats = {
            'total_prospects': conn.execute(f'SELECT COUNT(*) FROM prospects WHERE 1=1{prospect_filter}', params).fetchone()[0],
            'new_prospects': conn.execute(f"SELECT COUNT(*) FROM prospects WHERE status='prospect'{prospect_filter}", params).fetchone()[0],
            'qualified': conn.execute(f"SELECT COUNT(*) FROM prospects WHERE status='qualified'{prospect_filter}", params).fetchone()[0],
            'converted': conn.execute(f"SELECT COUNT(*) FROM prospects WHERE status='converted'{prospect_filter}", params).fetchone()[0],
            'touches_today': conn.execute(f'''
                SELECT COUNT(*) FROM contact_touches t
                JOIN prospects p ON t.prospect_id = p.id
                WHERE DATE(t.created_at) = DATE('now'){joined_filter}
            ''', params).fetchone()[0],
        }

        follow_ups_today = conn.execute(f'''
            SELECT t.*, p.company_name, p.contact_name FROM contact_touches t
            JOIN prospects p ON t.prospect_id = p.id
            WHERE t.follow_up_date = ?{joined_filter}
            ORDER BY t.created_at DESC
        ''', [today] + params).fetchall()

        overdue = conn.execute(f'''
            SELECT t.*, p.company_name, p.contact_name FROM contact_touches t
            JOIN prospects p ON t.prospect_id = p.id
            WHERE t.follow_up_date < ? AND t.follow_up_date IS NOT NULL{joined_filter}
            ORDER BY t.follow_up_date ASC LIMIT 10
        ''', [today] + params).fetchall()

        recent = conn.execute(f'''
            SELECT t.*, p.company_name, p.contact_name FROM contact_touches t
            JOIN prospects p ON t.prospect_id = p.id
            WHERE 1=1{joined_filter}
            ORDER BY t.created_at DESC LIMIT 10
        ''', params).fetchall()

        # Stale prospects - no touch in X days (configurable thresholds)
        # Critical: 14+ days, Warning: 7-13 days, for active prospects only
        stale_critical = conn.execute(f'''
            SELECT p.*,
                   MAX(t.created_at) as last_touch,
                   CAST(julianday('now') - julianday(COALESCE(MAX(t.created_at), p.created_at)) AS INTEGER) as days_since_touch
            FROM prospects p
            LEFT JOIN contact_touches t ON p.id = t.prospect_id
            WHERE p.status IN ('prospect', 'qualified'){joined_filter}
            GROUP BY p.id
            HAVING days_since_touch >= 14
            ORDER BY days_since_touch DESC
            LIMIT 10
        ''', params).fetchall()

        stale_warning = conn.execute(f'''
            SELECT p.*,
                   MAX(t.created_at) as last_touch,
                   CAST(julianday('now') - julianday(COALESCE(MAX(t.created_at), p.created_at)) AS INTEGER) as days_since_touch
            FROM prospects p
            LEFT JOIN contact_touches t ON p.id = t.prospect_id
            WHERE p.status IN ('prospect', 'qualified'){joined_filter}
            GROUP BY p.id
            HAVING days_since_touch >= 7 AND days_since_touch < 14
            ORDER BY days_since_touch DESC
            LIMIT 10
        ''', params).fetchall()

        # Never contacted - prospects with zero touches
        never_contacted = conn.execute(f'''
            SELECT p.*,
                   CAST(julianday('now') - julianday(p.created_at) AS INTEGER) as days_since_created
            FROM prospects p
            LEFT JOIN contact_touches t ON p.id = t.prospect_id
            WHERE p.status IN ('prospect', 'qualified'){joined_filter}
            GROUP BY p.id
            HAVING COUNT(t.id) = 0
            ORDER BY p.created_at ASC
            LIMIT 10
        ''', params).fetchall()

        # Add stale counts to stats
        stats['stale_critical'] = len(stale_critical)
        stats['stale_warning'] = len(stale_warning)
        stats['never_contacted'] = len(never_contacted)

        conn.close()
        return jsonify({
            'stats': stats,
            'follow_ups_today': [dict(f) for f in follow_ups_today],
            'overdue': [dict(o) for o in overdue],
            'recent_touches': [dict(r) for r in recent],
            'stale_critical': [dict(s) for s in stale_critical],
            'stale_warning': [dict(s) for s in stale_warning],
            'never_contacted': [dict(n) for n in never_contacted]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Seed mock data for testing
@app.route('/api/crm/seed-mock', methods=['POST'])
@admin_required
def seed_mock_data():
    try:
        conn = get_crm_db()

        # Clear existing data
        conn.execute('DELETE FROM contact_touches')
        conn.execute('DELETE FROM prospect_product_interest')
        conn.execute('DELETE FROM prospects')

        # Mock prospects with various states
        mock_prospects = [
            # Active prospects needing attention
            ('ABC Lumber Supply', 'John Smith', '555-0101', 'john@abclumber.com', 'Atlanta, GA', 'Large regional distributor', 'prospect', 'Trade show', 'Ian'),
            ('Southeast Building Materials', 'Sarah Johnson', '555-0102', 'sarah@sebm.com', 'Charlotte, NC', 'Interested in bulk orders', 'qualified', 'Referral', 'Ian'),
            ('Delta Construction Co', 'Mike Brown', '555-0103', 'mike@deltacon.com', 'Birmingham, AL', 'New construction focus', 'prospect', 'Cold call', 'Ian'),
            ('Gulf Coast Builders', 'Lisa Davis', '555-0104', 'lisa@gulfcoast.com', 'Mobile, AL', 'Coastal projects', 'qualified', 'Website', 'Ian'),
            ('Tennessee Timber', 'Bob Wilson', '555-0105', 'bob@tntimber.com', 'Nashville, TN', 'Regular buyer potential', 'prospect', 'Trade show', 'Ian'),

            # Stale prospects (will backdate)
            ('Midwest Framing Inc', 'Tom Harris', '555-0106', 'tom@midwestframe.com', 'Memphis, TN', 'High volume potential', 'prospect', 'Referral', 'Ian'),
            ('Southern Pine Distributors', 'Amy Clark', '555-0107', 'amy@southpine.com', 'Jackson, MS', 'Regional focus', 'qualified', 'Cold call', 'Ian'),
            ('Coastal Lumber Yard', 'Dan Miller', '555-0108', 'dan@coastallumber.com', 'Pensacola, FL', 'Retail and wholesale', 'prospect', 'Website', 'Ian'),

            # Never contacted
            ('Texas Wood Works', 'Chris Lee', '555-0109', 'chris@txwood.com', 'Houston, TX', 'Just added, needs first contact', 'prospect', 'Trade show', 'Ian'),
            ('Arkansas Building Supply', 'Pat Moore', '555-0110', 'pat@arkbs.com', 'Little Rock, AR', 'Warm lead from partner', 'prospect', 'Referral', 'Ian'),

            # Converted (success stories)
            ('Premium Lumber Co', 'Steve Taylor', '555-0111', 'steve@premiumlumber.com', 'Dallas, TX', 'Now a regular customer!', 'converted', 'Trade show', 'Ian'),
            ('Quality Builders Supply', 'Nancy White', '555-0112', 'nancy@qbs.com', 'New Orleans, LA', 'Converted after 3 months', 'converted', 'Referral', 'Ian'),
        ]

        prospect_ids = []
        for p in mock_prospects:
            cursor = conn.execute('''
                INSERT INTO prospects (company_name, contact_name, phone, email, address, notes, status, source, trader)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', p)
            prospect_ids.append(cursor.lastrowid)

        # Backdate some prospects to simulate stale ones
        # Prospects 6-8 (index 5-7) - make them stale (created 20 days ago)
        for i in range(5, 8):
            conn.execute('''
                UPDATE prospects SET created_at = datetime('now', '-20 days'), updated_at = datetime('now', '-20 days')
                WHERE id = ?
            ''', (prospect_ids[i],))

        # Never contacted prospects (9-10, index 8-9) - created 5 days ago, no touches
        for i in range(8, 10):
            conn.execute('''
                UPDATE prospects SET created_at = datetime('now', '-5 days'), updated_at = datetime('now', '-5 days')
                WHERE id = ?
            ''', (prospect_ids[i],))

        # Add touches for active prospects
        touches = [
            # Recent touches for prospects 1-5
            (prospect_ids[0], 'call', 'Discussed 2x4 pricing, very interested. Wants quote for 50 MBF.', '["2x4#2","2x6#2"]', 2),
            (prospect_ids[0], 'email', 'Sent pricing sheet and availability.', None, 1),
            (prospect_ids[1], 'meeting', 'Met at their office. Toured facility. Ready to place first order.', '["2x4#2","2x6#2","2x8#2"]', 0),
            (prospect_ids[1], 'call', 'Initial discovery call. Large operation, 200+ MBF/month potential.', None, 7),
            (prospect_ids[2], 'call', 'Left voicemail, will try again.', None, 3),
            (prospect_ids[3], 'email', 'Responded to inquiry. Scheduling call for next week.', '["2x4#2"]', 1),
            (prospect_ids[4], 'call', 'Good conversation. Needs time to review current supplier contract.', '["2x6#2","2x8#2"]', 5),

            # Old touches for stale prospects (these were 15-20 days ago)
            (prospect_ids[5], 'call', 'Initial call went well. Said to follow up in a week.', '["2x4#2"]', 18),
            (prospect_ids[6], 'email', 'Sent intro email with company info.', None, 15),
            (prospect_ids[7], 'call', 'Spoke briefly, was busy. Call back later.', None, 20),

            # Touches for converted prospects
            (prospect_ids[10], 'call', 'Closed the deal! First order: 100 MBF 2x4#2', '["2x4#2","2x6#2"]', 30),
            (prospect_ids[11], 'meeting', 'Signed contract. Great partnership ahead!', '["2x4#2","2x6#2","2x8#2"]', 45),
        ]

        for t in touches:
            days_ago = t[4]
            conn.execute(f'''
                INSERT INTO contact_touches (prospect_id, touch_type, notes, products_discussed, created_at)
                VALUES (?, ?, ?, ?, datetime('now', '-{days_ago} days'))
            ''', (t[0], t[1], t[2], t[3]))

        # Add follow-up dates
        # Overdue follow-ups
        conn.execute('''
            UPDATE contact_touches SET follow_up_date = date('now', '-3 days')
            WHERE prospect_id = ? AND touch_type = 'call'
        ''', (prospect_ids[5],))

        conn.execute('''
            UPDATE contact_touches SET follow_up_date = date('now', '-5 days')
            WHERE prospect_id = ? AND touch_type = 'email'
        ''', (prospect_ids[6],))

        # Today's follow-ups
        conn.execute('''
            UPDATE contact_touches SET follow_up_date = date('now')
            WHERE prospect_id = ? AND touch_type = 'email'
        ''', (prospect_ids[3],))

        # Future follow-ups
        conn.execute('''
            UPDATE contact_touches SET follow_up_date = date('now', '+3 days')
            WHERE prospect_id = ? AND touch_type = 'call'
        ''', (prospect_ids[4],))

        conn.commit()
        conn.close()

        return jsonify({
            'message': 'Mock data seeded successfully',
            'prospects_created': len(mock_prospects),
            'touches_created': len(touches)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Convert prospect to customer (update status)
@app.route('/api/crm/prospects/<int:id>/convert', methods=['POST'])

def convert_prospect(id):
    try:
        conn = get_crm_db()
        prospect = conn.execute('SELECT * FROM prospects WHERE id = ?', (id,)).fetchone()
        if not prospect:
            conn.close()
            return jsonify({'error': 'Prospect not found'}), 404
        conn.execute('''
            UPDATE prospects SET status = 'converted', updated_at = CURRENT_TIMESTAMP WHERE id = ?
        ''', (id,))
        conn.commit()
        prospect = conn.execute('SELECT * FROM prospects WHERE id = ?', (id,)).fetchone()
        conn.close()
        return jsonify(dict(prospect))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Wipe ALL CRM data (prospects, customers, mills)
@app.route('/api/crm/wipe-all', methods=['POST'])
@admin_required
def wipe_all_crm():
    try:
        conn = get_crm_db()
        conn.execute('DELETE FROM contact_touches')
        conn.execute('DELETE FROM prospect_product_interest')
        conn.execute('DELETE FROM prospects')
        conn.execute('DELETE FROM customers')
        conn.execute('DELETE FROM mills')
        conn.commit()
        conn.close()
        return jsonify({'message': 'All CRM data wiped (prospects, customers, mills)', 'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Cleanup endpoint - wipe all CRM data except Ian's
@app.route('/api/crm/cleanup-non-ian', methods=['POST'])
@admin_required
def cleanup_non_ian():
    try:
        conn = get_crm_db()

        # Delete touches for non-Ian prospects first
        conn.execute('''
            DELETE FROM contact_touches
            WHERE prospect_id IN (SELECT id FROM prospects WHERE trader != 'Ian' OR trader IS NULL)
        ''')

        # Delete interests for non-Ian prospects
        conn.execute('''
            DELETE FROM prospect_product_interest
            WHERE prospect_id IN (SELECT id FROM prospects WHERE trader != 'Ian' OR trader IS NULL)
        ''')

        # Delete non-Ian prospects
        result = conn.execute("DELETE FROM prospects WHERE trader != 'Ian' OR trader IS NULL")
        deleted = result.rowcount

        # Get remaining count
        remaining = conn.execute("SELECT COUNT(*) FROM prospects WHERE trader = 'Ian'").fetchone()[0]

        conn.commit()
        conn.close()

        return jsonify({
            'message': f'Deleted {deleted} non-Ian prospects. {remaining} Ian prospects remain.',
            'deleted': deleted,
            'remaining': remaining
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== CUSTOMERS API ====================

@app.route('/api/crm/customers', methods=['GET'])
def list_customers():
    try:
        conn = get_crm_db()
        trader = request.args.get('trader')
        query = 'SELECT * FROM customers WHERE 1=1'
        params = []
        if trader and trader != 'Admin':
            query += ' AND trader = ?'
            params.append(trader)
        query += ' ORDER BY name ASC'
        customers = conn.execute(query, params).fetchall()
        conn.close()
        return jsonify([dict(c) for c in customers])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/crm/customers', methods=['POST'])

def create_customer():
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'name is required'}), 400
        normalized_name = normalize_customer_name(name)
        conn = get_crm_db()
        # Check if customer already exists with this normalized name
        existing = conn.execute('SELECT * FROM customers WHERE name = ?', (normalized_name,)).fetchone()
        if existing:
            conn.close()
            return jsonify(dict(existing)), 200  # Return existing customer instead of creating duplicate
        cursor = conn.execute('''
            INSERT INTO customers (name, contact, phone, email, destination, locations, notes, trader)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            normalized_name,
            data.get('contact'),
            data.get('phone'),
            data.get('email'),
            data.get('destination'),
            json.dumps(data.get('locations')) if data.get('locations') else None,
            data.get('notes'),
            data.get('trader')
        ))
        conn.commit()
        customer = conn.execute('SELECT * FROM customers WHERE id = ?', (cursor.lastrowid,)).fetchone()
        conn.close()
        _log_audit(
            getattr(g, 'user', data.get('trader', 'unknown')),
            'customer_create', 'customer', cursor.lastrowid, normalized_name,
            ip_address=request.remote_addr
        )
        return jsonify(dict(customer)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/crm/customers/<int:id>', methods=['PUT'])

def update_customer(id):
    try:
        data = request.get_json() or {}
        conn = get_crm_db()
        # Only update fields that are present in the request (partial update support)
        # SECURITY: field names come from the hardcoded allowlist below, never from user input
        allowed = ['name', 'contact', 'phone', 'email', 'destination', 'locations', 'notes', 'trader']
        fields = [f for f in allowed if f in data]
        if not fields:
            return jsonify({'error': 'No fields to update'}), 400
        if not all(f in allowed for f in fields):
            return jsonify({'error': 'Invalid field'}), 400
        set_parts = []
        values = []
        for f in fields:
            set_parts.append(f'{f} = ?')
            if f == 'locations':
                values.append(json.dumps(data[f]) if data[f] else None)
            else:
                values.append(data[f])
        set_clause = ', '.join(set_parts) + ', updated_at = CURRENT_TIMESTAMP'
        values.append(id)
        conn.execute(f'UPDATE customers SET {set_clause} WHERE id = ?', values)
        conn.commit()
        customer = conn.execute('SELECT * FROM customers WHERE id = ?', (id,)).fetchone()
        conn.close()
        if not customer:
            return jsonify({'error': 'Customer not found'}), 404
        _log_audit(
            getattr(g, 'user', 'unknown'),
            'customer_update', 'customer', id, customer['name'],
            details=f'Updated fields: {", ".join(fields)}',
            ip_address=request.remote_addr
        )
        return jsonify(dict(customer))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/crm/customers/<int:id>', methods=['DELETE'])

def delete_customer(id):
    try:
        conn = get_crm_db()
        old = conn.execute('SELECT name FROM customers WHERE id = ?', (id,)).fetchone()
        conn.execute('DELETE FROM customers WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        _log_audit(
            getattr(g, 'user', 'unknown'),
            'customer_delete', 'customer', id,
            old['name'] if old else None,
            ip_address=request.remote_addr
        )
        return jsonify({'message': 'Customer deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== MILLS API ====================

@app.route('/api/crm/mills', methods=['GET'])
def list_mills():
    try:
        conn = get_crm_db()
        trader = request.args.get('trader')
        query = 'SELECT * FROM mills WHERE 1=1'
        params = []
        if trader and trader != 'Admin':
            query += ' AND trader = ?'
            params.append(trader)
        query += ' ORDER BY name ASC'
        mills = conn.execute(query, params).fetchall()
        conn.close()
        # Enrich with last_quoted date from MI quotes
        mill_list = [dict(m) for m in mills]
        try:
            mi_conn = get_mi_db()
            last_dates = mi_conn.execute(
                "SELECT mill_id, MAX(date) as last_date, COUNT(*) as quote_count FROM mill_quotes GROUP BY mill_id"
            ).fetchall()
            mi_conn.close()
            date_map = {r['mill_id']: {'last_quoted': r['last_date'], 'quote_count': r['quote_count']} for r in last_dates}
            for m in mill_list:
                info = date_map.get(m['id'], {})
                m['last_quoted'] = info.get('last_quoted')
                m['quote_count'] = info.get('quote_count', 0)
        except Exception:
            pass  # MI db may not exist yet
        return jsonify(mill_list)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/crm/mills', methods=['POST'])

def create_mill():
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'name is required'}), 400
        # Use extract_company_name to canonicalize, then normalize casing
        company = extract_company_name(name) if name else name
        company = normalize_mill_name(company)
        conn = get_crm_db()
        # Check if company already exists
        existing = conn.execute("SELECT * FROM mills WHERE UPPER(name)=?", (company.upper(),)).fetchone()
        if existing:
            conn.close()
            return jsonify(dict(existing)), 200  # Already exists
        cursor = conn.execute('''
            INSERT INTO mills (name, contact, phone, email, location, city, state, region, locations, products, notes, trader)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            company,
            data.get('contact'),
            data.get('phone'),
            data.get('email'),
            data.get('location'),
            data.get('city', ''),
            data.get('state', ''),
            data.get('region', 'central'),
            json.dumps(data.get('locations')) if isinstance(data.get('locations'), list) else data.get('locations', '[]'),
            json.dumps(data.get('products')) if isinstance(data.get('products'), list) else data.get('products', '[]'),
            data.get('notes', ''),
            data.get('trader') or ''
        ))
        conn.commit()
        mill = conn.execute('SELECT * FROM mills WHERE id = ?', (cursor.lastrowid,)).fetchone()
        conn.close()
        # Sync to MI
        if mill:
            sync_mill_to_mi(dict(mill))
        _log_audit(
            getattr(g, 'user', data.get('trader', 'unknown')),
            'mill_create', 'mill', cursor.lastrowid, company,
            ip_address=request.remote_addr
        )
        return jsonify(dict(mill)), 201
    except Exception as e:
        print(f"[create_mill ERROR] {e} | data={request.get_json()}", flush=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/crm/mills/<int:id>', methods=['PUT'])

def update_mill(id):
    try:
        data = request.get_json() or {}
        conn = get_crm_db()
        # Only update fields that are present in the request (partial update support)
        # SECURITY: field names come from the hardcoded allowlist below, never from user input
        allowed = ['name', 'contact', 'phone', 'email', 'location', 'locations', 'city', 'state', 'region', 'products', 'notes', 'trader']
        fields = [f for f in allowed if f in data]
        if not fields:
            return jsonify({'error': 'No fields to update'}), 400
        if not all(f in allowed for f in fields):
            return jsonify({'error': 'Invalid field'}), 400
        set_parts = []
        values = []
        for f in fields:
            set_parts.append(f'{f} = ?')
            if f in ('products', 'locations'):
                values.append(json.dumps(data[f]) if isinstance(data[f], list) else data.get(f, '[]'))
            else:
                values.append(data[f])
        set_clause = ', '.join(set_parts) + ', updated_at = CURRENT_TIMESTAMP'
        values.append(id)
        # Get old name before update (for syncing mill_quotes)
        old_row = conn.execute('SELECT name FROM mills WHERE id = ?', (id,)).fetchone()
        old_name = old_row['name'] if old_row else None

        conn.execute(f'UPDATE mills SET {set_clause} WHERE id = ?', values)
        conn.commit()
        mill = conn.execute('SELECT * FROM mills WHERE id = ?', (id,)).fetchone()
        conn.close()
        if not mill:
            return jsonify({'error': 'Mill not found'}), 404

        # Sync to Mill Intel database
        mill_dict = dict(mill)
        sync_mill_to_mi(mill_dict)

        # If name changed, update mill_quotes.mill_name in MI database
        new_name = mill_dict.get('name', '')
        if old_name and new_name and old_name != new_name:
            mi_conn = get_mi_db()
            mi_conn.execute('UPDATE mill_quotes SET mill_name = ? WHERE mill_id = ?', (new_name, id))
            mi_conn.commit()
            mi_conn.close()

        _log_audit(
            getattr(g, 'user', 'unknown'),
            'mill_update', 'mill', id, mill_dict.get('name'),
            details=f'Updated fields: {", ".join(fields)}',
            ip_address=request.remote_addr
        )
        return jsonify(mill_dict)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/crm/mills/<int:id>', methods=['DELETE'])

def delete_mill(id):
    try:
        conn = get_crm_db()
        old = conn.execute('SELECT name FROM mills WHERE id = ?', (id,)).fetchone()
        conn.execute('DELETE FROM mills WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        _log_audit(
            getattr(g, 'user', 'unknown'),
            'mill_delete', 'mill', id,
            old['name'] if old else None,
            ip_address=request.remote_addr
        )
        # Cascade delete from Mill Intel database
        try:
            mi_conn = get_mi_db()
            mi_conn.execute('DELETE FROM mill_quotes WHERE mill_id = ?', (id,))
            mi_conn.execute('DELETE FROM mills WHERE id = ?', (id,))
            mi_conn.commit()
            mi_conn.close()
        except Exception:
            pass  # MI db may not exist yet
        return jsonify({'message': 'Mill deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rename customer with old name returned for frontend trade-record sweep
@app.route('/api/crm/customers/<int:id>/rename', methods=['POST'])

def rename_customer(id):
    try:
        data = request.get_json()
        new_name = normalize_customer_name(data.get('name', '').strip())
        if not new_name:
            return jsonify({'error': 'New name required'}), 400
        conn = get_crm_db()
        old_row = conn.execute('SELECT name FROM customers WHERE id = ?', (id,)).fetchone()
        if not old_row:
            conn.close()
            return jsonify({'error': 'Customer not found'}), 404
        old_name = old_row['name']
        conn.execute('UPDATE customers SET name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (new_name, id))
        conn.commit()
        customer = conn.execute('SELECT * FROM customers WHERE id = ?', (id,)).fetchone()
        conn.close()
        result = dict(customer)
        result['old_name'] = old_name
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Rename mill with old name returned for frontend trade-record sweep
@app.route('/api/crm/mills/<int:id>/rename', methods=['POST'])

def rename_mill(id):
    try:
        data = request.get_json()
        new_name = data.get('name', '').strip()
        if not new_name:
            return jsonify({'error': 'New name required'}), 400
        conn = get_crm_db()
        old_row = conn.execute('SELECT name FROM mills WHERE id = ?', (id,)).fetchone()
        if not old_row:
            conn.close()
            return jsonify({'error': 'Mill not found'}), 404
        old_name = old_row['name']
        conn.execute('UPDATE mills SET name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (new_name, id))
        conn.commit()
        mill = conn.execute('SELECT * FROM mills WHERE id = ?', (id,)).fetchone()
        conn.close()
        # Also update Mill Intel database
        try:
            mi_conn = get_mi_db()
            mi_conn.execute('UPDATE mills SET name = ? WHERE id = ?', (new_name, id))
            mi_conn.execute('UPDATE mill_quotes SET mill_name = ? WHERE mill_id = ?', (new_name, id))
            mi_conn.commit()
            mi_conn.close()
        except Exception:
            pass  # MI db may not exist yet
        result = dict(mill)
        result['old_name'] = old_name
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== END CRM API ====================

# ==================== ENTITY RESOLUTION API ====================

@app.route('/api/entity/resolve', methods=['POST'])
def entity_resolve():
    """Resolve a name to a canonical entity (fuzzy match)."""
    try:
        data = request.get_json()
        name = (data or {}).get('name', '').strip()
        entity_type = (data or {}).get('type', 'mill')
        context = (data or {}).get('context', 'manual')
        if not name:
            return jsonify({'error': 'name is required'}), 400
        result = _entity_resolver.resolve(name, entity_type, context)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/search', methods=['GET'])
def entity_search():
    """Search entities by fuzzy match."""
    try:
        q = request.args.get('q', '').strip()
        entity_type = request.args.get('type', 'mill')
        limit = int(request.args.get('limit', '10'))
        if not q:
            return jsonify([])
        results = _entity_resolver.search(q, entity_type, limit)
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/review', methods=['GET'])
def entity_review_list():
    """Get pending review items."""
    try:
        reviews = _entity_resolver.get_pending_reviews()
        return jsonify(reviews)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/review/<int:review_id>', methods=['POST'])
def entity_review_submit(review_id):
    """Submit a review decision."""
    try:
        data = request.get_json()
        choice = (data or {}).get('choice', '')
        create_new = (data or {}).get('create_new', False)
        result = _entity_resolver.submit_review(
            review_id,
            chosen_canonical_id=choice if choice and choice != 'NEW' else None,
            create_new=create_new or choice == 'NEW'
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/link', methods=['POST'])
def entity_link():
    """Manually link a variant name to a canonical entity."""
    try:
        data = request.get_json()
        canonical_id = (data or {}).get('canonical_id', '')
        variant = (data or {}).get('variant', '').strip()
        if not canonical_id or not variant:
            return jsonify({'error': 'canonical_id and variant required'}), 400
        result = _entity_resolver.link_alias(canonical_id, variant)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/<canonical_id>/unified', methods=['GET', 'POST'])
def entity_unified(canonical_id):
    """Unified view: all data for one entity across all systems."""
    try:
        trades_data = None
        if request.method == 'POST':
            trades_data = request.get_json()
        result = _entity_resolver.get_unified_view(
            canonical_id, mi_db_path=MI_DB_PATH, trades_data=trades_data
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/merge', methods=['POST'])
def entity_merge():
    """Merge two entities into one."""
    try:
        data = request.get_json()
        source_id = (data or {}).get('source_id', '')
        target_id = (data or {}).get('target_id', '')
        if not source_id or not target_id:
            return jsonify({'error': 'source_id and target_id required'}), 400
        result = _entity_resolver.merge_entities(source_id, target_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/migrate', methods=['POST'])
def entity_migrate():
    """One-time migration: seed entities from existing data."""
    try:
        # Get MILL_DIRECTORY from state.js (we need to pass it from frontend or hardcode)
        # For now, we use the server-side aliases + scan CRM/MI data
        stats = _entity_resolver.migrate_existing(
            mill_company_aliases=MILL_COMPANY_ALIASES,
            mill_directory=None,  # Frontend will pass this
            customer_aliases=CUSTOMER_ALIASES,
            mi_db_path=MI_DB_PATH
        )
        return jsonify(stats)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/migrate-with-directory', methods=['POST'])
def entity_migrate_with_directory():
    """Migration with MILL_DIRECTORY from frontend."""
    try:
        data = request.get_json() or {}
        mill_directory = data.get('mill_directory', {})
        stats = _entity_resolver.migrate_existing(
            mill_company_aliases=MILL_COMPANY_ALIASES,
            mill_directory=mill_directory,
            customer_aliases=CUSTOMER_ALIASES,
            mi_db_path=MI_DB_PATH
        )
        return jsonify(stats)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity/stats', methods=['GET'])
def entity_stats():
    """Get entity resolution statistics."""
    try:
        stats = _entity_resolver.get_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== END ENTITY RESOLUTION API ====================

# ==================== TRADE CENTRAL SYNC ====================

_tc_staging = {'data': None, 'timestamp': 0}

@app.route('/api/tc-import', methods=['POST', 'OPTIONS'])
def tc_import_receive():
    """Receive TC order data (POST from TC tab or SYP tab).
    Stores in server memory for the SYP tab to retrieve."""
    if request.method == 'OPTIONS':
        # Handle CORS preflight
        resp = app.make_default_options_response()
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp

    try:
        data = request.get_json(force=True)
        if not isinstance(data, list):
            return jsonify({'error': 'Expected array of orders'}), 400
        _tc_staging['data'] = data
        _tc_staging['timestamp'] = time.time()
        return jsonify({
            'status': 'ok',
            'count': len(data),
            'message': f'Received {len(data)} TC orders. Fetch from /api/tc-import to retrieve.'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/tc-import', methods=['GET'])
def tc_import_retrieve():
    """Retrieve staged TC data (called by SYP Analytics tab)."""
    if not _tc_staging['data']:
        return jsonify({'error': 'No TC data staged. POST data first.'}), 404
    age = time.time() - _tc_staging['timestamp']
    return jsonify({
        'data': _tc_staging['data'],
        'count': len(_tc_staging['data']),
        'age_seconds': round(age)
    })

@app.route('/api/tc-import', methods=['DELETE'])
def tc_import_clear():
    """Clear staged TC data."""
    _tc_staging['data'] = None
    _tc_staging['timestamp'] = 0
    return jsonify({'status': 'cleared'})

# ==================== END TRADE CENTRAL SYNC ====================

# ==================== FUTURES DATA PROXY ====================

# CME SYP futures months: F=Jan, H=Mar, K=May, N=Jul, U=Sep, X=Nov
SYP_MONTHS = [
    {'code': 'F', 'label': 'Jan', 'yahoo_code': 'F'},
    {'code': 'H', 'label': 'Mar', 'yahoo_code': 'H'},
    {'code': 'K', 'label': 'May', 'yahoo_code': 'K'},
    {'code': 'N', 'label': 'Jul', 'yahoo_code': 'N'},
    {'code': 'U', 'label': 'Sep', 'yahoo_code': 'U'},
    {'code': 'X', 'label': 'Nov', 'yahoo_code': 'X'},
]

futures_cache = {'data': None, 'timestamp': 0}
FUTURES_CACHE_TTL = 300  # 5 minutes

@app.route('/api/futures/quotes')
def futures_quotes():
    """Fetch delayed SYP futures quotes from Yahoo Finance"""
    now = time.time()
    if futures_cache['data'] and (now - futures_cache['timestamp']) < FUTURES_CACHE_TTL:
        return jsonify(futures_cache['data'])

    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    results = {'contracts': [], 'front': None, 'fetched_at': datetime.now().isoformat()}

    # Determine current and next year for contract symbols
    current_year = datetime.now().year
    years = [current_year % 100, (current_year + 1) % 100]

    # Fetch front-month quote
    try:
        r = requests.get(
            'https://query1.finance.yahoo.com/v8/finance/chart/SYP=F?interval=1d&range=max',
            headers=headers, timeout=10
        )
        if r.status_code == 200:
            data = r.json()
            if data.get('chart', {}).get('result'):
                meta = data['chart']['result'][0]['meta']
                results['front'] = {
                    'price': meta.get('regularMarketPrice'),
                    'previousClose': meta.get('previousClose'),
                    'name': meta.get('shortName', 'SYP Front')
                }
                ts_list = data['chart']['result'][0].get('timestamp', [])
                quote = data['chart']['result'][0].get('indicators', {}).get('quote', [{}])[0]
                closes = quote.get('close', [])
                opens = quote.get('open', [])
                highs = quote.get('high', [])
                lows = quote.get('low', [])
                volumes = quote.get('volume', [])
                results['front']['history'] = [
                    {'timestamp': ts_list[i], 'close': closes[i],
                     'open': opens[i] if i < len(opens) else None,
                     'high': highs[i] if i < len(highs) else None,
                     'low': lows[i] if i < len(lows) else None,
                     'volume': volumes[i] if i < len(volumes) else None}
                    for i in range(len(ts_list)) if i < len(closes) and closes[i] is not None
                ]
    except Exception as e:
        print(f"Front month fetch error: {e}")

    # Fetch individual contract months
    for month in SYP_MONTHS:
        for yr in years:
            symbol = f"SYP{month['yahoo_code']}{yr}.CME"
            try:
                r = requests.get(
                    f'https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=max',
                    headers=headers, timeout=8
                )
                if r.status_code == 200:
                    data = r.json()
                    if data.get('chart', {}).get('result'):
                        meta = data['chart']['result'][0]['meta']
                        price = meta.get('regularMarketPrice')
                        if price and price > 0:
                            ts_list = data['chart']['result'][0].get('timestamp', [])
                            cquote = data['chart']['result'][0].get('indicators', {}).get('quote', [{}])[0]
                            closes = cquote.get('close', [])
                            opens = cquote.get('open', [])
                            highs = cquote.get('high', [])
                            lows = cquote.get('low', [])
                            volumes = cquote.get('volume', [])
                            history = [
                                {'timestamp': ts_list[i], 'close': closes[i],
                                 'open': opens[i] if i < len(opens) else None,
                                 'high': highs[i] if i < len(highs) else None,
                                 'low': lows[i] if i < len(lows) else None,
                                 'volume': volumes[i] if i < len(volumes) else None}
                                for i in range(len(ts_list)) if i < len(closes) and closes[i] is not None
                            ]
                            results['contracts'].append({
                                'symbol': symbol,
                                'month': month['label'],
                                'code': month['code'],
                                'year': 2000 + yr,
                                'price': price,
                                'previousClose': meta.get('previousClose'),
                                'history': history
                            })
            except Exception as e:
                print(f"Contract fetch error {symbol}: {e}")
            time.sleep(0.2)  # Yahoo Finance rate limit; blocking OK since results are cached 5min

    # Sort contracts by year then month order
    month_order = {'F': 0, 'H': 1, 'K': 2, 'N': 3, 'U': 4, 'X': 5}
    results['contracts'].sort(key=lambda c: (c['year'], month_order.get(c['code'], 99)))

    futures_cache['data'] = results
    futures_cache['timestamp'] = now
    return jsonify(results)

# Excel file parser for mill pricing intake
@app.route('/api/parse-excel', methods=['GET', 'POST'])
def parse_excel():
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500

    if request.method == 'GET':
        return jsonify({'error': 'POST a .xlsx file as multipart form data'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        all_rows = []
        sheet_count = len(wb.sheetnames)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                # Skip completely empty rows
                if any(c.strip() for c in row_data):
                    sheet_rows.append(row_data)
            if sheet_rows:
                # Add sheet marker so AI can distinguish locations/tabs
                if sheet_count > 1:
                    all_rows.append([f'=== SHEET: {sheet_name} ==='])
                all_rows.extend(sheet_rows)
        wb.close()
        return jsonify({'rows': all_rows, 'count': len(all_rows), 'sheet_count': sheet_count})
    except Exception as e:
        return jsonify({'error': f'Failed to parse Excel: {str(e)}'}), 400

# PDF file parser for mill pricing intake
@app.route('/api/parse-pdf', methods=['GET', 'POST'])
def parse_pdf():
    try:
        import pdfplumber
    except ImportError:
        return jsonify({'error': 'pdfplumber not installed. Run: pip install pdfplumber'}), 500

    if request.method == 'GET':
        return jsonify({'error': 'POST a .pdf file as multipart form data'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    tmp_path = None
    try:
        import tempfile
        # Save to temp file since pdfplumber needs a file path
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name
            file.save(tmp_path)

        pages_text = []
        tables = []
        with pdfplumber.open(tmp_path) as pdf:
            for i, page in enumerate(pdf.pages):
                # Extract text
                text = page.extract_text() or ''
                if text.strip():
                    pages_text.append(text)

                # Extract tables (pdfplumber is excellent at this)
                page_tables = page.extract_tables()
                for table in page_tables:
                    # Convert to list of lists of strings
                    cleaned = []
                    for row in table:
                        cleaned.append([str(cell).strip() if cell else '' for cell in row])
                    if cleaned:
                        tables.append({'page': i + 1, 'rows': cleaned})

        # If no text/tables found, it's likely a scanned PDF — convert pages to images
        page_images = []
        if not pages_text and not tables:
            import io, base64
            with pdfplumber.open(tmp_path) as pdf:
                for i, page in enumerate(pdf.pages):
                    try:
                        img = page.to_image(resolution=200)
                        buf = io.BytesIO()
                        img.original.save(buf, format='PNG')
                        b64 = base64.b64encode(buf.getvalue()).decode()
                        page_images.append({'page': i + 1, 'data': b64, 'media_type': 'image/png'})
                    except Exception:
                        pass

        # Clean up temp file
        os.unlink(tmp_path)

        return jsonify({
            'text': '\n\n'.join(pages_text),
            'tables': tables,
            'pages': len(pages_text),
            'table_count': len(tables),
            'images': page_images
        })
    except Exception as e:
        # Clean up temp file on error
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except:
                pass
        return jsonify({'error': f'Failed to parse PDF: {str(e)}'}), 400

# Health check for Railway
@app.route('/api/po/seed')
def po_seed():
    """Serve pre-parsed PO history seed data from po-seed.json."""
    seed_path = os.path.join(os.path.dirname(__file__), 'po-seed.json')
    if os.path.exists(seed_path):
        return send_from_directory(os.path.dirname(__file__), 'po-seed.json', mimetype='application/json')
    return jsonify([])

@app.route('/health')
def health():
    return jsonify({'status': 'ok', 'cache_size': len(geo_cache)})

@app.route('/health/mi')
def health_mi():
    """Mill Intel health check — verifies SQLite has data and reports counts."""
    try:
        mi_conn = get_mi_db()
        quote_count = mi_conn.execute("SELECT COUNT(*) FROM mill_quotes").fetchone()[0]
        mill_count = mi_conn.execute("SELECT COUNT(*) FROM mills").fetchone()[0]
        latest_row = mi_conn.execute("SELECT MAX(date) as latest FROM mill_quotes").fetchone()
        latest_date = latest_row['latest'] if latest_row else None
        mi_conn.close()

        crm_conn = get_crm_db()
        crm_mill_count = crm_conn.execute("SELECT COUNT(*) FROM mills").fetchone()[0]
        crm_cust_count = crm_conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        crm_conn.close()

        status = 'ok' if quote_count > 0 else 'empty'
        return jsonify({
            'status': status,
            'mi_quotes': quote_count,
            'mi_mills': mill_count,
            'latest_quote_date': latest_date,
            'crm_mills': crm_mill_count,
            'crm_customers': crm_cust_count
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/mi/reseed', methods=['POST'])
def mi_reseed():
    """Manually trigger MI re-seed from Supabase cloud data."""
    try:
        # Clear existing quotes to force re-seed
        mi_conn = get_mi_db()
        mi_conn.execute("DELETE FROM mill_quotes")
        mi_conn.commit()
        mi_conn.close()
        seed_mi_from_supabase()
        # Report results
        mi_conn = get_mi_db()
        new_count = mi_conn.execute("SELECT COUNT(*) FROM mill_quotes").fetchone()[0]
        mi_conn.close()
        return jsonify({'status': 'ok', 'quotes_seeded': new_count})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== AUTH ENDPOINTS ====================

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    """Authenticate a trader and return a JWT token."""
    if not jwt:
        return jsonify({'error': 'JWT support not available (pip install PyJWT)'}), 500
    data = request.get_json() or {}
    username = (data.get('username') or '').strip().lower()
    password = data.get('password') or ''
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    expected = TRADER_CREDENTIALS.get(username)
    if not expected or not secrets.compare_digest(pw_hash, expected):
        return jsonify({'error': 'Invalid credentials'}), 401
    payload = {
        'user': username,
        'trader': data.get('trader', username),
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS),
        'iat': datetime.now(timezone.utc)
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm='HS256')
    return jsonify({'token': token, 'user': username, 'expires_in': JWT_EXPIRY_HOURS * 3600})

@app.route('/api/auth/verify', methods=['GET'])
def auth_verify():
    """Verify a JWT token is still valid."""
    token = _get_token_from_request()
    payload = _decode_token(token)
    if payload:
        return jsonify({'valid': True, 'user': payload.get('user', ''), 'trader': payload.get('trader', '')})
    return jsonify({'valid': False}), 401

# ==================== CONFIG ENDPOINT ====================

@app.route('/api/config', methods=['GET'])
def get_config():
    """Return Supabase credentials (behind auth so frontend doesn't hardcode them)."""
    return jsonify({
        'supabaseUrl': os.environ.get('SUPABASE_URL', ''),
        'supabaseKey': os.environ.get('SUPABASE_ANON_KEY', ''),
    })

# ==========================================
# PRICING MATRIX — standalone read-only view
# ==========================================

PRICING_PASSWORD = os.environ.get('PRICING_PASSWORD', '2026')

@app.route('/pricing')
def pricing_page():
    return send_from_directory('.', 'pricing.html')

@app.route('/api/pricing/auth', methods=['POST'])
def pricing_auth():
    if not PRICING_PASSWORD:
        return jsonify({'ok': False, 'error': 'Pricing portal not configured. Set PRICING_PASSWORD env var.'}), 503

    # Rate limit by IP
    ip = request.remote_addr or 'unknown'
    now = time.time()
    record = _pricing_login_attempts.get(ip, {'count': 0, 'lockout_until': 0})
    if record['lockout_until'] > now:
        remaining = int(record['lockout_until'] - now)
        return jsonify({'ok': False, 'error': f'Too many attempts. Try again in {remaining}s.'}), 429

    data = request.get_json() or {}
    if data.get('password') == PRICING_PASSWORD:
        # Reset on success
        _pricing_login_attempts.pop(ip, None)
        return jsonify({'ok': True})

    # Track failed attempt
    record['count'] = record.get('count', 0) + 1
    if record['count'] >= PRICING_MAX_ATTEMPTS:
        record['lockout_until'] = now + PRICING_LOCKOUT_SECONDS
        record['count'] = 0
    _pricing_login_attempts[ip] = record
    return jsonify({'ok': False, 'error': 'Invalid password'}), 401

@app.route('/api/pricing/customers', methods=['GET'])
def pricing_customers():
    """Merged customer list for portal quote builder (CRM + MI, deduplicated)."""
    seen = {}
    # CRM customers (have destination field)
    try:
        conn = get_crm_db()
        rows = conn.execute("SELECT name, destination, locations FROM customers ORDER BY name").fetchall()
        conn.close()
        for r in rows:
            name = r['name']
            dest = r['destination'] or ''
            if not dest and r['locations']:
                try:
                    locs = json.loads(r['locations']) if isinstance(r['locations'], str) else r['locations']
                    if locs and isinstance(locs, list) and len(locs) > 0:
                        loc = locs[0]
                        if isinstance(loc, dict):
                            dest = (loc.get('city', '') + ', ' + loc.get('state', '')).strip(', ')
                        elif isinstance(loc, str):
                            dest = loc
                except: pass
            key = name.lower().strip()
            if key not in seen:
                seen[key] = {'name': name, 'destination': dest}
    except: pass
    # MI customers
    try:
        conn = get_mi_db()
        rows = conn.execute("SELECT name, destination FROM customers ORDER BY name").fetchall()
        conn.close()
        for r in rows:
            key = r['name'].lower().strip()
            if key not in seen:
                seen[key] = {'name': r['name'], 'destination': r['destination'] or ''}
    except: pass
    return jsonify(sorted(seen.values(), key=lambda x: x['name'].lower()))

# Server-side matrix cutoff (shared between in-app and portal)
_matrix_cutoff = {'since': ''}

@app.route('/api/pricing/cutoff', methods=['GET'])
def get_matrix_cutoff():
    return jsonify({'since': _matrix_cutoff['since']})

@app.route('/api/pricing/cutoff', methods=['POST'])
def set_matrix_cutoff():
    data = request.get_json() or {}
    _matrix_cutoff['since'] = data.get('since', '')
    return jsonify({'since': _matrix_cutoff['since']})

# ==========================================
# MILL INTEL ROUTES — /api/mi/*
# ==========================================

# ----- MI: MILLS -----

@app.route('/api/mi/mills', methods=['GET'])
def mi_list_mills():
    """List all mills from CRM (single source of truth)."""
    conn = get_crm_db()
    rows = conn.execute("SELECT * FROM mills ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mi/mills', methods=['POST'])

def mi_create_mill():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Mill name required'}), 400
    city = data.get('city', '')
    state = data.get('state', '') or mi_extract_state(city) or ''
    region = data.get('region', '') or mi_get_region(state)
    lat = data.get('lat')
    lon = data.get('lon')
    if city and (lat is None or lon is None):
        coords = mi_geocode_location(city)
        if coords:
            lat, lon = coords['lat'], coords['lon']
    # Create in CRM and sync to MI
    crm_mill = find_or_create_crm_mill(name, city, state, region, lat, lon, data.get('trader', ''))
    sync_mill_to_mi(crm_mill)
    return jsonify(crm_mill), 201

@app.route('/api/mi/mills/<int:mill_id>', methods=['GET'])
def mi_get_mill(mill_id):
    conn = get_crm_db()
    mill = conn.execute("SELECT * FROM mills WHERE id=?", (mill_id,)).fetchone()
    conn.close()
    if not mill:
        return jsonify({'error': 'Not found'}), 404
    mi_conn = get_mi_db()
    quotes = mi_conn.execute("SELECT * FROM mill_quotes WHERE mill_id=? ORDER BY date DESC LIMIT 100", (mill_id,)).fetchall()
    mi_conn.close()
    result = dict(mill)
    result['quotes'] = [dict(q) for q in quotes]
    return jsonify(result)

@app.route('/api/mi/mills/<int:mill_id>', methods=['PUT'])

def mi_update_mill(mill_id):
    data = request.get_json()
    # Update CRM mill
    conn = get_crm_db()
    fields = []
    vals = []
    for k in ['name', 'city', 'state', 'lat', 'lon', 'region', 'notes', 'contact', 'phone', 'email', 'trader']:
        if k in data:
            fields.append(f"{k}=?")
            vals.append(data[k])
    if 'location' not in data and ('city' in data or 'state' in data):
        # Auto-update location from city/state
        city = data.get('city', '')
        state = data.get('state', '')
        if city or state:
            fields.append("location=?")
            vals.append(f"{city}, {state}".strip(', '))
    if 'products' in data:
        fields.append("products=?")
        vals.append(json.dumps(data['products']))
    if fields:
        fields.append("updated_at=CURRENT_TIMESTAMP")
        vals.append(mill_id)
        conn.execute(f"UPDATE mills SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    mill = conn.execute("SELECT * FROM mills WHERE id=?", (mill_id,)).fetchone()
    conn.close()
    if not mill:
        return jsonify({'error': 'Not found'}), 404
    # Sync to MI
    sync_mill_to_mi(dict(mill))
    return jsonify(dict(mill))

@app.route('/api/mi/mills/geocode', methods=['POST'])
def mi_geocode_mill():
    data = request.get_json()
    location = data.get('location', '')
    if not location:
        return jsonify({'error': 'Location required'}), 400
    coords = mi_geocode_location(location)
    if coords:
        return jsonify(coords)
    return jsonify({'error': f'Could not geocode: {location}'}), 404

@app.route('/api/admin/consolidate-mills', methods=['POST'])
@admin_required
def consolidate_mills():
    """One-time migration: consolidate per-location mill entries into per-company entries."""
    conn = get_crm_db()
    all_mills = [dict(m) for m in conn.execute("SELECT * FROM mills ORDER BY id").fetchall()]

    # Group by company name
    groups = {}
    for m in all_mills:
        company = extract_company_name(m['name'])
        if company not in groups:
            groups[company] = []
        groups[company].append(m)

    changes = []
    for company, entries in groups.items():
        if len(entries) <= 1 and entries[0]['name'] == company:
            continue  # Already consolidated

        # Pick survivor: prefer the one already named as company, else lowest ID
        survivor = next((e for e in entries if e['name'] == company), entries[0])
        others = [e for e in entries if e['id'] != survivor['id']]

        if not others and survivor['name'] == company:
            continue  # Single entry already correct

        # Build merged locations array
        existing_locs = json.loads(survivor.get('locations') or '[]')
        seen_cities = {(l.get('city', '').lower(), l.get('state', '').upper()) for l in existing_locs}

        for e in entries:
            city = e.get('city', '') or ''
            state = e.get('state', '') or ''
            key = (city.lower(), state.upper())
            if city and key not in seen_cities:
                existing_locs.append({
                    'city': city, 'state': state,
                    'lat': e.get('lat'), 'lon': e.get('lon'),
                    'name': e['name']
                })
                seen_cities.add(key)

        # Merge products
        all_products = set()
        for e in entries:
            try:
                prods = json.loads(e.get('products') or '[]')
                all_products.update(prods)
            except:
                pass

        # Update survivor
        conn.execute(
            "UPDATE mills SET name=?, locations=?, products=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (company, json.dumps(existing_locs), json.dumps(sorted(all_products)), survivor['id'])
        )

        # Reassign mill_quotes in MI DB
        if others:
            mi_conn = get_mi_db()
            old_ids = [e['id'] for e in others]
            placeholders = ','.join('?' * len(old_ids))
            mi_conn.execute(
                f"UPDATE mill_quotes SET mill_id=? WHERE mill_id IN ({placeholders})",
                [survivor['id']] + old_ids
            )
            mi_conn.commit()
            mi_conn.close()

            # Delete non-survivor CRM entries
            conn.execute(f"DELETE FROM mills WHERE id IN ({placeholders})", old_ids)

        changes.append({
            'company': company,
            'survivor_id': survivor['id'],
            'merged': len(others),
            'locations': len(existing_locs)
        })

    conn.commit()
    conn.close()

    # Re-sync to MI
    sync_crm_mills_to_mi()

    return jsonify({
        'consolidated': len([c for c in changes if c['merged'] > 0]),
        'total_companies': len(groups),
        'changes': changes
    })

# ----- MI: MILL QUOTES -----

@app.route('/api/mi/quotes', methods=['GET'])
def mi_list_quotes():
    conn = get_mi_db()
    conditions = ["1=1"]
    params = []
    for k, col in [('mill', 'mill_name'), ('product', 'product'), ('trader', 'trader')]:
        v = request.args.get(k)
        if v:
            conditions.append(f"{col}=?")
            params.append(v)
    since = request.args.get('since')
    if since:
        conditions.append("date>=?")
        params.append(since)
    until = request.args.get('until')
    if until:
        conditions.append("date<=?")
        params.append(until)
    try:
        limit = int(request.args.get('limit', 500))
    except (ValueError, TypeError):
        limit = 500
    sql = f"SELECT * FROM mill_quotes WHERE {' AND '.join(conditions)} ORDER BY date DESC, created_at DESC LIMIT ?"
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mi/quotes', methods=['POST'])

def mi_submit_quotes():
    data = request.get_json()
    if data is None:
        return jsonify({'error': 'Request body must be JSON'}), 400
    quotes = data if isinstance(data, list) else [data]
    conn = get_mi_db()
    created = []

    # Auto-replace: For each mill+product+length combo being uploaded, delete existing quotes
    # This ensures uploaded quotes always show as "today" even if price unchanged
    today_date = datetime.now().strftime('%Y-%m-%d')
    cleared_combos = set()
    # Capture old prices before deletion for mill_price_changes tracking
    _old_prices = {}  # key: (MILL_UPPER, PROD_UPPER, LEN_UPPER) -> {price, date, mill_id}
    for q in quotes:
        mill_name = q.get('mill', '').strip()
        product = q.get('product', '').strip()
        length = q.get('length', 'RL').strip()
        if mill_name and product:
            key = (mill_name.upper(), product.upper(), length.upper())
            if key not in cleared_combos:
                cleared_combos.add(key)
                # Capture old price before deleting
                old_row = conn.execute(
                    """SELECT price, date, mill_id FROM mill_quotes
                       WHERE UPPER(mill_name)=? AND UPPER(product)=? AND UPPER(COALESCE(length,'RL'))=?
                       ORDER BY id DESC LIMIT 1""",
                    (mill_name.upper(), product.upper(), length.upper() if length else 'RL')
                ).fetchone()
                if old_row:
                    _old_prices[key] = {'price': old_row['price'], 'date': old_row['date'], 'mill_id': old_row['mill_id']}
                deleted = conn.execute(
                    "DELETE FROM mill_quotes WHERE UPPER(mill_name)=? AND UPPER(product)=? AND UPPER(COALESCE(length,'RL'))=?",
                    (mill_name.upper(), product.upper(), length.upper() if length else 'RL')
                ).rowcount
                if deleted:
                    app.logger.info(f"Replaced existing quote for {mill_name} {product} {length}")

    # Pre-cache existing CRM mills to avoid per-quote DB lookups and geocoding
    crm_conn_pre = get_crm_db()
    _mill_cache = {}
    for row in crm_conn_pre.execute("SELECT * FROM mills").fetchall():
        _mill_cache[row['name'].upper()] = dict(row)
    crm_conn_pre.close()

    # Pre-sync cached mills to MI so mill_id is available
    for cached_mill in _mill_cache.values():
        sync_mill_to_mi(cached_mill, mi_conn=conn)

    for q in quotes:
        mill_name = q.get('mill', '').strip()
        if not mill_name or not q.get('product') or not q.get('price'):
            continue
        try:
            price_val = float(q['price'])
            if price_val <= 0:
                continue
        except (ValueError, TypeError):
            continue

        # Find or create mill in CRM — skip geocoding for known mills
        company = extract_company_name(mill_name)
        cached = _mill_cache.get(company.upper())
        if cached:
            crm_mill = cached
        else:
            city = q.get('city', '')
            state = mi_extract_state(city) if city else ''
            region = mi_get_region(state) if state else 'central'
            lat, lon = None, None
            if city:
                coords = mi_geocode_location(city)
                if coords:
                    lat, lon = coords['lat'], coords['lon']
            crm_mill = find_or_create_crm_mill(mill_name, city, state, region, lat, lon,
                                                q.get('trader', 'Unknown'))
            _mill_cache[company.upper()] = crm_mill
            sync_mill_to_mi(crm_mill, mi_conn=conn)

        mill_id = crm_mill['id']
        product = q['product']

        # Update products list on CRM mill
        existing_products = json.loads(crm_mill.get('products') or '[]')
        if product not in existing_products:
            existing_products.append(product)
            crm_conn = get_crm_db()
            crm_conn.execute("UPDATE mills SET products=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                             (json.dumps(existing_products), mill_id))
            crm_conn.commit()
            crm_conn.close()
            # Also update MI mirror
            conn.execute("UPDATE mills SET products=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                         (json.dumps(existing_products), mill_id))

        conn.execute(
            """INSERT INTO mill_quotes (mill_id, mill_name, product, price, length, volume, tls,
               ship_window, notes, date, trader, source, raw_text)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (mill_id, mill_name, product, price_val,
             q.get('length', 'RL'),
             max(0, float(q.get('volume', 0) or 0)),
             max(0, int(float(q.get('tls', 0) or 0))),
             q.get('shipWindow', q.get('ship_window', '')) or 'Prompt', q.get('notes', ''),
             q.get('date', today_date),  # Preserve original date for syncs, default to today
             q.get('trader', 'Unknown'), q.get('source', 'manual'), q.get('raw_text', ''))
        )
        created.append(q)

        # Track price changes for intelligence/mill-moves
        length_val = q.get('length', 'RL')
        combo_key = (mill_name.upper(), product.upper(), (length_val or 'RL').upper())
        old_info = _old_prices.get(combo_key)
        if old_info and abs(price_val - old_info['price']) > 0.001:
            # Check if this exact change was already recorded (prevent duplicates from re-submissions)
            existing_change = conn.execute(
                """SELECT id FROM mill_price_changes
                   WHERE mill_id=? AND product=? AND length=? AND new_price=? AND date=?
                   ORDER BY id DESC LIMIT 1""",
                (mill_id, product, length_val or 'RL', price_val, q.get('date', today_date))
            ).fetchone()
            if not existing_change:
                change_val = round(price_val - old_info['price'], 2)
                pct_val = round((change_val / old_info['price']) * 100, 2) if old_info['price'] else None
                conn.execute(
                    """INSERT INTO mill_price_changes (mill_id, mill_name, product, length,
                       old_price, new_price, change, pct_change, date, prev_date, source, trader)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (mill_id, mill_name, product, length_val or 'RL',
                     old_info['price'], price_val, change_val, pct_val,
                     q.get('date', today_date), old_info['date'],
                     q.get('source', 'manual'), q.get('trader', 'Unknown'))
                )

    conn.commit()
    conn.close()
    invalidate_matrix_cache()  # Clear cached matrix data

    # Recompute price changes if this was a bulk sync (>50 quotes = likely full sync)
    if len(created) > 50:
        recompute_price_changes()

    return jsonify({'created': len(created), 'quotes': created}), 201

@app.route('/api/mi/quotes/by-mill', methods=['DELETE'])

def mi_delete_mill_quotes():
    """Delete all quotes for a specific mill."""
    mill_name = request.args.get('mill', '').strip()
    if not mill_name:
        return jsonify({'error': 'mill parameter required'}), 400
    conn = get_mi_db()
    cur = conn.execute("DELETE FROM mill_quotes WHERE mill_name=?", (mill_name,))
    conn.commit()
    conn.close()
    invalidate_matrix_cache()
    return jsonify({'deleted': cur.rowcount, 'mill': mill_name})

@app.route('/api/mi/quotes/<int:quote_id>', methods=['DELETE'])

def mi_delete_quote(quote_id):
    conn = get_mi_db()
    conn.execute("DELETE FROM mill_quotes WHERE id=?", (quote_id,))
    conn.commit()
    conn.close()
    invalidate_matrix_cache()  # Clear cached matrix data
    return jsonify({'deleted': quote_id})

@app.route('/api/mi/quotes/rename-mill', methods=['POST'])

def mi_rename_mill_quotes():
    """Bulk rename mill_name in quotes (admin utility)."""
    data = request.get_json()
    old_name = data.get('old_name', '').strip()
    new_name = data.get('new_name', '').strip()
    if not old_name or not new_name:
        return jsonify({'error': 'old_name and new_name required'}), 400
    conn = get_mi_db()
    cur = conn.execute("UPDATE mill_quotes SET mill_name=? WHERE mill_name=?", (new_name, old_name))
    conn.commit()
    conn.close()
    return jsonify({'updated': cur.rowcount, 'old_name': old_name, 'new_name': new_name})

@app.route('/api/mi/quotes/latest', methods=['GET'])
def mi_latest_quotes():
    conn = get_mi_db()
    product = request.args.get('product')
    region = request.args.get('region')
    since = request.args.get('since')
    sql = """
        SELECT mq.*, m.lat, m.lon, m.region, m.city, m.state
        FROM mill_quotes mq
        LEFT JOIN mills m ON mq.mill_id = m.id
        WHERE mq.id IN (
            SELECT id FROM mill_quotes mq2
            WHERE mq2.mill_name = mq.mill_name AND mq2.product = mq.product
            ORDER BY mq2.date DESC, mq2.created_at DESC
            LIMIT 1
        )
    """
    params = []
    if product:
        sql += " AND LOWER(REPLACE(mq.product, ' ', ''))=LOWER(REPLACE(?, ' ', ''))"
        params.append(product)
    if region:
        sql += " AND m.region=?"
        params.append(region)
    if since:
        sql += " AND mq.date>=?"
        params.append(since)
    sql += " ORDER BY mq.product, mq.price"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mi/quotes/matrix', methods=['GET'])
def mi_quote_matrix():
    detail = request.args.get('detail', '')
    filter_product = request.args.get('product', '')
    filter_since = request.args.get('since', '')

    # Check cache first (30s TTL to handle concurrent users)
    cache_key = f"matrix:{detail}:{filter_product}:{filter_since}"
    cached = get_cached_matrix(cache_key)
    if cached:
        return jsonify(cached)

    conn = get_mi_db()

    if detail == 'length':
        # Inner subquery also respects since filter to avoid stale data
        inner_where = ""
        inner_params = []
        if filter_since:
            inner_where = " WHERE date >= ?"
            inner_params = [filter_since]
        sql = f"""
            SELECT mq.mill_name, mq.product, mq.length, mq.price, mq.date, mq.volume,
                   mq.ship_window, mq.tls, mq.trader,
                   m.lat, m.lon, m.region, m.city, m.state
            FROM mill_quotes mq
            LEFT JOIN mills m ON mq.mill_id = m.id
            WHERE mq.id IN (
                SELECT MAX(id) FROM mill_quotes{inner_where} GROUP BY mill_name, product, length
            )
        """
        params = list(inner_params)
        if filter_product:
            sql += " AND mq.product = ?"
            params.append(filter_product)
        sql += " ORDER BY mq.mill_name, mq.product, mq.length"
        rows = conn.execute(sql, params).fetchall()
        conn.close()

        matrix = {}
        mills = set()
        columns = set()
        best_by_col = {}

        def length_sort_key(l):
            if not l or l == 'RL':
                return 999
            try:
                return float(str(l).replace("'", "").split('-')[0])
            except:
                return 998

        for r in rows:
            r = dict(r)
            mill = r['mill_name']
            prod = r['product']
            length = r['length'] or 'RL'
            col_key = f"{prod} {length}'" if length != 'RL' else f"{prod} RL"
            mills.add(mill)
            columns.add(col_key)
            if mill not in matrix:
                matrix[mill] = {}
            # Use MILL_DIRECTORY for accurate city/state/region (CRM parent record may differ)
            dir_city, dir_state = MILL_DIRECTORY.get(mill, (r['city'], r['state']))
            dir_region = MI_STATE_REGIONS.get(dir_state.upper(), 'central') if dir_state else r['region']
            matrix[mill][col_key] = {
                'price': r['price'], 'date': r['date'], 'volume': r['volume'],
                'ship_window': r['ship_window'], 'tls': r['tls'], 'trader': r['trader'],
                'product': prod, 'length': length,
                'lat': r['lat'], 'lon': r['lon'], 'region': dir_region,
                'city': dir_city, 'state': dir_state
            }
            if col_key not in best_by_col or r['price'] < best_by_col[col_key]:
                best_by_col[col_key] = r['price']

        def col_sort(c):
            parts = c.rsplit(' ', 1)
            prod = parts[0]
            length = parts[1].replace("'", "") if len(parts) > 1 else 'RL'
            return (prod, length_sort_key(length))

        sorted_cols = sorted(columns, key=col_sort)
        unique_products = sorted(set(c.rsplit(' ', 1)[0] for c in columns))

        result = {
            'matrix': matrix,
            'mills': sorted(mills),
            'columns': sorted_cols,
            'products': unique_products,
            'best_by_col': best_by_col,
            'detail': 'length'
        }
        set_cached_matrix(cache_key, result)
        return jsonify(result)
    else:
        inner_where2 = ""
        inner_params2 = []
        if filter_since:
            inner_where2 = " WHERE date >= ?"
            inner_params2 = [filter_since]
        sql = f"""
            SELECT mq.mill_name, mq.product, mq.price, mq.date, mq.volume, mq.ship_window,
                   mq.tls, mq.trader, m.lat, m.lon, m.region, m.city, m.state
            FROM mill_quotes mq
            LEFT JOIN mills m ON mq.mill_id = m.id
            WHERE mq.id IN (
                SELECT MAX(id) FROM mill_quotes{inner_where2} GROUP BY mill_name, product
            )
        """
        params = list(inner_params2)
        sql += " ORDER BY mq.mill_name, mq.product"
        rows = conn.execute(sql, params).fetchall()
        conn.close()

        matrix = {}
        mills = set()
        products = set()
        best_by_product = {}

        for r in rows:
            r = dict(r)
            mill = r['mill_name']
            prod = r['product']
            mills.add(mill)
            products.add(prod)
            if mill not in matrix:
                matrix[mill] = {}
            dir_city, dir_state = MILL_DIRECTORY.get(mill, (r['city'], r['state']))
            dir_region = MI_STATE_REGIONS.get(dir_state.upper(), 'central') if dir_state else r['region']
            matrix[mill][prod] = {
                'price': r['price'], 'date': r['date'], 'volume': r['volume'],
                'ship_window': r['ship_window'], 'tls': r['tls'], 'trader': r['trader'],
                'lat': r['lat'], 'lon': r['lon'], 'region': dir_region,
                'city': dir_city, 'state': dir_state
            }
            if prod not in best_by_product or r['price'] < best_by_product[prod]:
                best_by_product[prod] = r['price']

        result = {
            'matrix': matrix,
            'mills': sorted(mills),
            'products': sorted(products),
            'best_by_product': best_by_product
        }
        set_cached_matrix(cache_key, result)
        return jsonify(result)

@app.route('/api/mi/quotes/history', methods=['GET'])
def mi_quote_history():
    mill = request.args.get('mill')
    product = request.args.get('product')
    try:
        days = int(request.args.get('days', 90))
    except (ValueError, TypeError):
        days = 90
    cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    conn = get_mi_db()
    conditions = ["date >= ?"]
    params = [cutoff]
    if mill:
        conditions.append("mill_name=?")
        params.append(mill)
    if product:
        conditions.append("product=?")
        params.append(product)
    rows = conn.execute(
        f"SELECT * FROM mill_quotes WHERE {' AND '.join(conditions)} ORDER BY date ASC, created_at ASC",
        params
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ----- MI: INTELLIGENCE ENGINE -----

@app.route('/api/mi/intel/signals', methods=['GET'])
def mi_intel_signals():
    product_filter = request.args.get('product')
    conn = get_mi_db()
    now = datetime.now()
    d7 = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    d14 = (now - timedelta(days=14)).strftime('%Y-%m-%d')
    d30 = (now - timedelta(days=30)).strftime('%Y-%m-%d')

    if product_filter:
        products = [product_filter]
    else:
        products = [r['product'] for r in conn.execute("SELECT DISTINCT product FROM mill_quotes").fetchall()]

    all_signals = {}
    for product in products:
        signals = []

        # 1. Supply Pressure
        mills_7d = conn.execute(
            "SELECT COUNT(DISTINCT mill_name) as cnt, SUM(volume) as vol FROM mill_quotes WHERE product=? AND date>=?",
            (product, d7)
        ).fetchone()
        mills_30d = conn.execute(
            "SELECT COUNT(DISTINCT mill_name) as cnt, SUM(volume) as vol, COUNT(*) as quotes FROM mill_quotes WHERE product=? AND date>=?",
            (product, d30)
        ).fetchone()
        m7 = mills_7d['cnt'] or 0
        v7 = mills_7d['vol'] or 0
        m30 = mills_30d['cnt'] or 0
        v30 = mills_30d['vol'] or 0
        avg_weekly_mills = m30 / 4.3 if m30 else 0
        avg_weekly_vol = v30 / 4.3 if v30 else 0

        if avg_weekly_mills > 0:
            mill_ratio = m7 / avg_weekly_mills
            vol_ratio = v7 / avg_weekly_vol if avg_weekly_vol > 0 else 1
            if mill_ratio > 1.2 or vol_ratio > 1.3:
                direction = 'bearish'
                strength = 'strong' if mill_ratio > 1.5 or vol_ratio > 1.5 else 'moderate'
            elif mill_ratio < 0.8 or vol_ratio < 0.7:
                direction = 'bullish'
                strength = 'strong' if mill_ratio < 0.5 else 'moderate'
            else:
                direction = 'neutral'
                strength = 'weak'
            signals.append({
                'signal': 'supply_pressure',
                'mills_offering_7d': m7, 'volume_7d': round(v7, 1),
                'mills_avg_weekly': round(avg_weekly_mills, 1), 'volume_avg_weekly': round(avg_weekly_vol, 1),
                'direction': direction, 'strength': strength,
                'explanation': f"{m7} mills offering {product} this week ({round(v7)} MBF), vs {round(avg_weekly_mills,1)} mills avg. {'More supply = potential to short.' if direction=='bearish' else 'Tighter supply = consider buying.' if direction=='bullish' else 'Supply steady.'}"
            })

        # 2. Price Momentum
        prices_14d = conn.execute(
            "SELECT date, AVG(price) as avg_price FROM mill_quotes WHERE product=? AND date>=? GROUP BY date ORDER BY date",
            (product, d14)
        ).fetchall()
        prices_30d = conn.execute(
            "SELECT date, AVG(price) as avg_price FROM mill_quotes WHERE product=? AND date>=? GROUP BY date ORDER BY date",
            (product, d30)
        ).fetchall()

        def calc_slope(prices):
            if len(prices) < 2:
                return 0
            n = len(prices)
            x_sum = n * (n - 1) / 2
            x2_sum = n * (n - 1) * (2 * n - 1) / 6
            y_sum = sum(p['avg_price'] for p in prices)
            xy_sum = sum(i * p['avg_price'] for i, p in enumerate(prices))
            denom = n * x2_sum - x_sum * x_sum
            if denom == 0:
                return 0
            return (n * xy_sum - x_sum * y_sum) / denom

        slope_14d = calc_slope(prices_14d)
        slope_30d = calc_slope(prices_30d)
        current_avg = prices_14d[-1]['avg_price'] if prices_14d else 0

        if abs(slope_14d) > 0.5:
            direction = 'bullish' if slope_14d > 0 else 'bearish'
            strength = 'strong' if abs(slope_14d) > 2 else 'moderate'
        else:
            direction = 'neutral'
            strength = 'weak'
        signals.append({
            'signal': 'price_momentum',
            'current_avg': round(current_avg, 2), 'slope_14d': round(slope_14d, 2), 'slope_30d': round(slope_30d, 2),
            'direction': direction, 'strength': strength,
            'explanation': f"{product} avg ${round(current_avg)}. Price {'rising' if slope_14d > 0 else 'falling'} ~${abs(round(slope_14d, 1))}/day over 14d. {'Buy before prices climb higher.' if direction=='bullish' else 'Prices softening — wait or short.' if direction=='bearish' else 'Prices stable.'}"
        })

        # 3. Print vs Street
        latest_rl = conn.execute(
            "SELECT price FROM rl_prices WHERE product=? ORDER BY date DESC LIMIT 1",
            (product,)
        ).fetchone()
        if latest_rl and current_avg > 0:
            rl_price = latest_rl['price']
            gap = rl_price - current_avg
            if gap > 10:
                direction = 'bearish'
                explanation = f"Mills offering {product} ${round(gap)} below RL print (${round(rl_price)}). Street is cheaper than print."
            elif gap < -10:
                direction = 'bullish'
                explanation = f"Mills charging ${round(abs(gap))} above RL print (${round(rl_price)}). Genuine tightness."
            else:
                direction = 'neutral'
                explanation = f"Mill prices tracking close to RL print (${round(rl_price)} vs ${round(current_avg)} street)."
            signals.append({
                'signal': 'print_vs_street',
                'rl_price': round(rl_price, 2), 'avg_street': round(current_avg, 2), 'gap': round(gap, 2),
                'direction': direction,
                'strength': 'strong' if abs(gap) > 20 else 'moderate' if abs(gap) > 10 else 'weak',
                'explanation': explanation
            })

        # 4. Regional Arbitrage
        regional_prices = conn.execute("""
            SELECT m.region, MIN(mq.price) as best_price, mq.mill_name
            FROM mill_quotes mq LEFT JOIN mills m ON mq.mill_id = m.id
            WHERE mq.product=? AND mq.date>=?
            GROUP BY m.region
        """, (product, d7)).fetchall()
        if len(regional_prices) >= 2:
            rp = {r['region']: {'price': r['best_price'], 'mill': r['mill_name']} for r in regional_prices}
            opps = []
            regions = list(rp.keys())
            for i in range(len(regions)):
                for j in range(i+1, len(regions)):
                    spread = abs(rp[regions[i]]['price'] - rp[regions[j]]['price'])
                    if spread > 10:
                        cheaper = regions[i] if rp[regions[i]]['price'] < rp[regions[j]]['price'] else regions[j]
                        opps.append({
                            'from_region': cheaper,
                            'to_region': regions[j] if cheaper == regions[i] else regions[i],
                            'spread': round(spread, 2),
                            'cheaper_mill': rp[cheaper]['mill'],
                            'cheaper_price': rp[cheaper]['price']
                        })
            if opps:
                signals.append({
                    'signal': 'regional_arbitrage',
                    'opportunities': opps,
                    'direction': 'opportunity',
                    'strength': 'strong' if any(o['spread'] > 25 for o in opps) else 'moderate',
                    'explanation': f"Regional spread on {product}: " + ', '.join(
                        f"${o['spread']} between {o['from_region']} and {o['to_region']} ({o['cheaper_mill']} at ${o['cheaper_price']})"
                        for o in opps[:3]
                    )
                })

        # 5. Offering Velocity
        daily_counts = conn.execute(
            "SELECT date, COUNT(*) as cnt FROM mill_quotes WHERE product=? AND date>=? GROUP BY date",
            (product, d30)
        ).fetchall()
        if daily_counts:
            avg_daily = sum(r['cnt'] for r in daily_counts) / max(len(daily_counts), 1)
            recent_daily = [r['cnt'] for r in daily_counts if r['date'] >= d7]
            recent_avg = sum(recent_daily) / max(len(recent_daily), 1) if recent_daily else 0
            vel_ratio = recent_avg / avg_daily if avg_daily > 0 else 1
            if vel_ratio > 1.3:
                direction = 'bearish'
                explanation = f"Above-average quoting on {product} ({round(recent_avg,1)} vs {round(avg_daily,1)} daily avg). Mills pushing inventory."
            elif vel_ratio < 0.7:
                direction = 'bullish'
                explanation = f"Below-average activity on {product}. Quiet market suggests tightening."
            else:
                direction = 'neutral'
                explanation = f"Normal quoting velocity on {product}."
            signals.append({
                'signal': 'offering_velocity',
                'recent_avg_daily': round(recent_avg, 1), 'avg_daily_30d': round(avg_daily, 1),
                'velocity_ratio': round(vel_ratio, 2),
                'direction': direction,
                'strength': 'strong' if abs(vel_ratio - 1) > 0.5 else 'moderate' if abs(vel_ratio - 1) > 0.3 else 'weak',
                'explanation': explanation
            })

        # 6. Volume Trend
        weekly_vol = conn.execute("""
            SELECT strftime('%%W', date) as week, SUM(volume) as vol
            FROM mill_quotes WHERE product=? AND date>=? AND volume > 0
            GROUP BY week ORDER BY week
        """, (product, d30)).fetchall()
        if len(weekly_vol) >= 2:
            vols = [r['vol'] for r in weekly_vol]
            avg_vol = sum(vols) / len(vols)
            latest_vol = vols[-1]
            change = ((latest_vol - avg_vol) / avg_vol * 100) if avg_vol > 0 else 0
            if change > 15:
                direction = 'bearish'
            elif change < -15:
                direction = 'bullish'
            else:
                direction = 'neutral'
            signals.append({
                'signal': 'volume_trend',
                'latest_week_mbf': round(latest_vol, 1), 'avg_week_mbf': round(avg_vol, 1),
                'change_pct': round(change, 1),
                'direction': direction,
                'strength': 'strong' if abs(change) > 30 else 'moderate' if abs(change) > 15 else 'weak',
                'explanation': f"{product} volume {'up' if change > 0 else 'down'} {round(abs(change))}% vs avg ({round(latest_vol)} MBF this week vs {round(avg_vol)} avg)."
            })

        all_signals[product] = signals

    conn.close()
    return jsonify(all_signals)

@app.route('/api/mi/intel/recommendations', methods=['GET'])
def mi_intel_recommendations():
    product_filter = request.args.get('product')
    conn = get_mi_db()
    if product_filter:
        products = [product_filter]
    else:
        products = [r['product'] for r in conn.execute("SELECT DISTINCT product FROM mill_quotes").fetchall()]

    import json as json_mod
    with app.test_request_context(f'/api/mi/intel/signals{"?product=" + product_filter if product_filter else ""}'):
        sig_response = mi_intel_signals()
        all_signals = json_mod.loads(sig_response.get_data())

    recommendations = []
    for product in products:
        signals = all_signals.get(product, [])
        score = 0
        reasons = []

        weights = {
            'supply_pressure': 2, 'price_momentum': 3, 'print_vs_street': 1.5,
            'offering_velocity': 1, 'volume_trend': 1.5, 'regional_arbitrage': 0
        }

        for sig in signals:
            w = weights.get(sig['signal'], 1)
            if sig['direction'] == 'bullish':
                score += w * (2 if sig['strength'] == 'strong' else 1)
            elif sig['direction'] == 'bearish':
                score -= w * (2 if sig['strength'] == 'strong' else 1)
            if sig.get('explanation'):
                reasons.append(sig['explanation'])

        if score >= 4: action = 'BUY NOW'
        elif score >= 2: action = 'LEAN BUY'
        elif score <= -4: action = 'SHORT / SELL'
        elif score <= -2: action = 'LEAN SHORT'
        else: action = 'HOLD / NEUTRAL'

        if score >= 4: margin_range = [35, 50]
        elif score >= 2: margin_range = [28, 40]
        elif score <= -4: margin_range = [15, 22]
        elif score <= -2: margin_range = [18, 28]
        else: margin_range = [22, 35]

        best = conn.execute("""
            SELECT mq.mill_name, mq.price, m.city, m.region
            FROM mill_quotes mq LEFT JOIN mills m ON mq.mill_id = m.id
            WHERE mq.product=? AND mq.date >= ?
            ORDER BY mq.price ASC LIMIT 1
        """, (product, (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))).fetchone()

        recommendations.append({
            'product': product, 'action': action, 'score': round(score, 1),
            'confidence': min(abs(score) / 8, 1.0), 'margin_range': margin_range,
            'best_source': dict(best) if best else None, 'reasons': reasons,
            'signal_count': len(signals)
        })

    conn.close()
    return jsonify(sorted(recommendations, key=lambda r: abs(r['score']), reverse=True))

@app.route('/api/mi/intel/trends', methods=['GET'])
def mi_intel_trends():
    product_filter = request.args.get('product')
    try:
        days = int(request.args.get('days', 90))
    except (ValueError, TypeError):
        days = 90
    conn = get_mi_db()
    since = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

    sql = """
        SELECT date, product,
               ROUND(AVG(price),1) as avg_price,
               ROUND(MIN(price),1) as min_price,
               ROUND(MAX(price),1) as max_price,
               COUNT(DISTINCT mill_name) as mill_count,
               ROUND(SUM(COALESCE(volume,0)),1) as total_volume,
               COUNT(*) as quote_count
        FROM mill_quotes
        WHERE date >= ?
    """
    params = [since]
    if product_filter:
        sql += " AND product = ?"
        params.append(product_filter)
    sql += " GROUP BY date, product ORDER BY date"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    trends = {}
    for r in rows:
        p = r['product']
        if p not in trends:
            trends[p] = []
        trends[p].append({
            'date': r['date'], 'avg_price': r['avg_price'],
            'min_price': r['min_price'], 'max_price': r['max_price'],
            'mill_count': r['mill_count'], 'volume': r['total_volume'],
            'quotes': r['quote_count']
        })

    return jsonify(trends)

# ----- MI: CUSTOMERS -----

@app.route('/api/mi/customers', methods=['GET'])
def mi_list_customers():
    conn = get_mi_db()
    rows = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mi/customers', methods=['POST'])
def mi_create_customer():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    dest = data.get('destination', '')
    lat, lon = data.get('lat'), data.get('lon')
    if dest and (lat is None or lon is None):
        coords = mi_geocode_location(dest)
        if coords:
            lat, lon = coords['lat'], coords['lon']
    conn = get_mi_db()
    conn.execute("INSERT INTO customers (name, destination, lat, lon, trader) VALUES (?,?,?,?,?)",
                 (name, dest, lat, lon, data.get('trader', '')))
    conn.commit()
    cust = conn.execute("SELECT * FROM customers WHERE id=last_insert_rowid()").fetchone()
    conn.close()
    return jsonify(dict(cust)), 201

# ----- MI: RL PRICES -----

@app.route('/api/mi/rl', methods=['GET'])
def mi_list_rl():
    conn = get_mi_db()
    rows = conn.execute("SELECT * FROM rl_prices ORDER BY date DESC LIMIT 200").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mi/rl', methods=['POST'])
def mi_add_rl():
    data = request.get_json()
    entries = data if isinstance(data, list) else [data]
    rows = []
    for e in entries:
        if not e.get('date') or not e.get('product') or not e.get('region'):
            continue
        try:
            price = float(e['price'])
            if price <= 0:
                continue
        except (ValueError, TypeError):
            continue
        length = e.get('length', 'RL') or 'RL'
        rows.append((e['date'], e['region'], e['product'], length, price))
    if rows:
        for attempt in range(3):
            try:
                conn = get_mi_db()
                conn.executemany(
                    "INSERT OR REPLACE INTO rl_prices (date, region, product, length, price) VALUES (?,?,?,?,?)",
                    rows
                )
                conn.commit()
                conn.close()
                return jsonify({'created': len(rows)}), 201
            except sqlite3.OperationalError as e:
                if 'locked' in str(e) and attempt < 2:
                    try: conn.close()
                    except: pass
                    import time; time.sleep(1)
                    continue
                raise
    return jsonify({'created': 0}), 201

# ----- RL: HISTORICAL PRICE API -----

@app.route('/api/rl/history', methods=['GET'])
def rl_history():
    """Return time series of RL prices, filtered by product/region/length/date range."""
    try:
        product = request.args.get('product')
        region = request.args.get('region')
        length = request.args.get('length')
        date_from = request.args.get('from')
        date_to = request.args.get('to')

        cache_key = f"history_{product}_{region}_{length}_{date_from}_{date_to}"
        cached = get_rl_cached(cache_key)
        if cached is not None:
            return jsonify(cached)

        sql = "SELECT date, region, product, length, price FROM rl_prices WHERE 1=1"
        params = []

        if product:
            sql += " AND product = ?"
            params.append(product)
        if region:
            sql += " AND region = ?"
            params.append(region)
        if length:
            sql += " AND length = ?"
            params.append(length)
        if date_from:
            sql += " AND date >= ?"
            params.append(date_from)
        if date_to:
            sql += " AND date <= ?"
            params.append(date_to)

        sql += " ORDER BY date"

        conn = get_mi_db()
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        result = [dict(r) for r in rows]
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rl/dates', methods=['GET'])
def rl_dates():
    """Return list of available dates with row counts."""
    try:
        conn = get_mi_db()
        rows = conn.execute(
            "SELECT date, COUNT(*) as row_count FROM rl_prices GROUP BY date ORDER BY date"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rl/entry', methods=['GET'])
def rl_entry():
    """Return full RL entry for one date, structured by region → product → length → price."""
    try:
        date = request.args.get('date')
        if not date:
            return jsonify({'error': 'date parameter required'}), 400

        conn = get_mi_db()
        rows = conn.execute(
            "SELECT region, product, length, price FROM rl_prices WHERE date = ?",
            (date,)
        ).fetchall()
        conn.close()

        result = {'date': date, 'west': {}, 'central': {}, 'east': {}}
        for r in rows:
            region = r['region']
            product = r['product']
            length = r['length']
            price = r['price']
            if region not in result:
                result[region] = {}
            if product not in result[region]:
                result[region][product] = {}
            result[region][product][length] = price

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rl/save', methods=['POST'])
def rl_save():
    """Save/upsert RL price rows for a given date."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'JSON body required'}), 400

        date = data.get('date')
        entries = data.get('rows', [])
        if not date or not entries:
            return jsonify({'error': 'date and rows required'}), 400

        rows = []
        for e in entries:
            region = e.get('region', '').strip()
            product = e.get('product', '').strip()
            length = e.get('length', 'RL').strip() or 'RL'
            try:
                price = float(e.get('price', 0))
                if price <= 0:
                    continue
            except (ValueError, TypeError):
                continue
            if not region or not product:
                continue
            rows.append((date, region, product, length, price))

        if rows:
            conn = get_mi_db()
            conn.executemany(
                "INSERT OR REPLACE INTO rl_prices (date, region, product, length, price) VALUES (?,?,?,?,?)",
                rows
            )
            conn.commit()
            conn.close()
            invalidate_rl_cache()

        return jsonify({'saved': len(rows)}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rl/chart-batch', methods=['GET'])
def rl_chart_batch():
    """Batch endpoint: returns all 3 regions + spread data for one product in a single call."""
    try:
        product = request.args.get('product', '2x4#2')
        length = request.args.get('length', 'RL')
        date_from = request.args.get('from', '')
        date_to = request.args.get('to', '')

        cache_key = f"chart_batch_{product}_{length}_{date_from}_{date_to}"
        cached = get_rl_cached(cache_key)
        if cached is not None:
            return jsonify(cached)

        conn = get_mi_db()

        # Fetch all regions for this product
        sql = "SELECT date, region, price FROM rl_prices WHERE product=? AND length=?"
        params = [product, length]
        if date_from:
            sql += " AND date>=?"
            params.append(date_from)
        if date_to:
            sql += " AND date<=?"
            params.append(date_to)
        sql += " ORDER BY date"
        rows = conn.execute(sql, params).fetchall()

        west, central, east = [], [], []
        for r in rows:
            entry = {'date': r['date'], 'price': r['price']}
            if r['region'] == 'west':
                west.append(entry)
            elif r['region'] == 'central':
                central.append(entry)
            elif r['region'] == 'east':
                east.append(entry)

        # Compute spreads if product is #2 grade
        spread46, spread_wc = [], []
        if '#2' in product:
            # Get the companion product for 2x4/2x6 spread
            if product.startswith('2x4'):
                spread_product = product.replace('2x4', '2x6')
            elif product.startswith('2x6'):
                spread_product = product.replace('2x6', '2x4')
            else:
                spread_product = None

            if spread_product:
                sql2 = "SELECT date, region, price FROM rl_prices WHERE product=? AND length=? AND region='west'"
                params2 = [spread_product, length]
                if date_from:
                    sql2 += " AND date>=?"
                    params2.append(date_from)
                if date_to:
                    sql2 += " AND date<=?"
                    params2.append(date_to)
                sql2 += " ORDER BY date"
                spread_rows = conn.execute(sql2, params2).fetchall()
                spread_map = {r['date']: r['price'] for r in spread_rows}

                # Also get 2x4#2 west prices for the spread
                if product.startswith('2x4'):
                    w4_map = {e['date']: e['price'] for e in west}
                    w6_map = spread_map
                else:
                    w6_map = {e['date']: e['price'] for e in west}
                    sql3 = "SELECT date, price FROM rl_prices WHERE product='2x4#2' AND length=? AND region='west'"
                    params3 = [length]
                    if date_from:
                        sql3 += " AND date>=?"
                        params3.append(date_from)
                    if date_to:
                        sql3 += " AND date<=?"
                        params3.append(date_to)
                    sql3 += " ORDER BY date"
                    w4_rows = conn.execute(sql3, params3).fetchall()
                    w4_map = {r['date']: r['price'] for r in w4_rows}

                # Build 2x4/2x6 spread
                for e in west:
                    d = e['date']
                    v4 = w4_map.get(d)
                    v6 = w6_map.get(d)
                    if v4 and v6:
                        spread46.append({'date': d, 'spread': round(v6 - v4)})

            # West vs Central spread
            c_map = {e['date']: e['price'] for e in central}
            for e in west:
                d = e['date']
                w_price = e['price']
                c_price = c_map.get(d)
                if w_price and c_price:
                    spread_wc.append({'date': d, 'spread': round(w_price - c_price)})

        conn.close()

        result = {
            'west': west,
            'central': central,
            'east': east,
            'spread46': spread46,
            'spreadWC': spread_wc
        }
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rl/spreads', methods=['GET'])
def rl_spreads():
    """Batch endpoint: returns length/dimension/grade spreads with historical stats."""
    try:
        region = request.args.get('region', 'west')
        date_from = request.args.get('from', '')
        date_to = request.args.get('to', '')
        exclude_covid = request.args.get('exclude_covid', '0') == '1'

        cache_key = f"spreads_{region}_{date_from}_{date_to}_covid{int(exclude_covid)}"
        cached = get_rl_cached(cache_key)
        if cached is not None:
            return jsonify(cached)

        conn = get_mi_db()

        # Find latest two "complete" dates in range (skip partial entries with <10 rows)
        date_sql = """SELECT date, COUNT(*) as cnt FROM rl_prices WHERE region=?"""
        date_params = [region]
        if date_from:
            date_sql += " AND date>=?"
            date_params.append(date_from)
        if date_to:
            date_sql += " AND date<=?"
            date_params.append(date_to)
        date_sql += " GROUP BY date HAVING cnt >= 10 ORDER BY date DESC LIMIT 2"
        date_rows = conn.execute(date_sql, date_params).fetchall()
        latest_date = date_rows[0]['date'] if date_rows else None
        prev_date = date_rows[1]['date'] if len(date_rows) > 1 else None

        if not latest_date:
            conn.close()
            return jsonify({'length_spreads': [], 'dimension_spreads': [], 'grade_spreads': [], 'wow_changes': []})

        # Get latest prices
        latest_rows = conn.execute(
            "SELECT product, length, price FROM rl_prices WHERE region=? AND date=?",
            (region, latest_date)
        ).fetchall()
        latest = {}
        for r in latest_rows:
            latest[(r['product'], r['length'])] = r['price']

        # Get previous week prices for WoW
        prev = {}
        if prev_date:
            prev_rows = conn.execute(
                "SELECT product, length, price FROM rl_prices WHERE region=? AND date=?",
                (region, prev_date)
            ).fetchall()
            for r in prev_rows:
                prev[(r['product'], r['length'])] = r['price']

        # Get aggregates for full date range
        agg_sql = "SELECT product, length, AVG(price) as avg_price, MIN(price) as min_price, MAX(price) as max_price, COUNT(*) as cnt FROM rl_prices WHERE region=?"
        agg_params = [region]
        if date_from:
            agg_sql += " AND date>=?"
            agg_params.append(date_from)
        if date_to:
            agg_sql += " AND date<=?"
            agg_params.append(date_to)
        agg_sql += " GROUP BY product, length"
        agg_rows = conn.execute(agg_sql, agg_params).fetchall()
        agg = {}
        for r in agg_rows:
            agg[(r['product'], r['length'])] = {
                'avg': round(r['avg_price'], 2),
                'min': r['min_price'],
                'max': r['max_price'],
                'cnt': r['cnt']
            }

        # For percentile rank, get all prices per product+length
        # Only include "complete" dates (10+ rows) to avoid partial mid-week updates skewing averages
        complete_dates_sql = "SELECT date FROM rl_prices WHERE region=?"
        cd_params = [region]
        if date_from:
            complete_dates_sql += " AND date>=?"
            cd_params.append(date_from)
        if date_to:
            complete_dates_sql += " AND date<=?"
            cd_params.append(date_to)
        complete_dates_sql += " GROUP BY date HAVING COUNT(*) >= 10"
        complete_dates = set(r['date'] for r in conn.execute(complete_dates_sql, cd_params).fetchall())

        hist_sql = "SELECT date, product, length, price FROM rl_prices WHERE region=?"
        hist_params = [region]
        if date_from:
            hist_sql += " AND date>=?"
            hist_params.append(date_from)
        if date_to:
            hist_sql += " AND date<=?"
            hist_params.append(date_to)
        hist_sql += " ORDER BY date"
        hist_rows = conn.execute(hist_sql, hist_params).fetchall()
        hist = {}
        for r in hist_rows:
            if r['date'] not in complete_dates:
                continue
            key = (r['product'], r['length'])
            if key not in hist:
                hist[key] = {}
            hist[key][r['date']] = r['price']

        conn.close()

        # COVID exclusion: remove dates between 2020-03-01 and 2022-12-31
        COVID_START = '2020-03-01'
        COVID_END = '2022-12-31'
        if exclude_covid:
            for key in hist:
                hist[key] = {d: p for d, p in hist[key].items()
                             if d < COVID_START or d > COVID_END}

        def _weighted_avg(spreads_by_date, latest_dt):
            """Recency-weighted average using exponential decay (half-life = 180 days)."""
            import math
            if not spreads_by_date:
                return None
            HALF_LIFE = 180  # days
            decay = math.log(2) / HALF_LIFE
            try:
                latest_ord = datetime.strptime(latest_dt, '%Y-%m-%d').toordinal()
            except Exception:
                latest_ord = datetime.now().toordinal()
            w_sum = 0.0
            w_total = 0.0
            for d, s in spreads_by_date:
                try:
                    d_ord = datetime.strptime(d, '%Y-%m-%d').toordinal()
                except Exception:
                    continue
                age_days = latest_ord - d_ord
                w = math.exp(-decay * age_days)
                w_sum += w * s
                w_total += w
            return round(w_sum / w_total, 2) if w_total > 0 else None

        def _compute_spread_stats(hist_a, hist_b, current_spread, latest_dt):
            """Compute spread stats from two date-aligned price histories.
            Returns (avg, weighted_avg, min, max, pct, n) tuple."""
            common_dates = sorted(set(hist_a.keys()) & set(hist_b.keys()))
            if common_dates:
                hist_spreads = [(d, hist_a[d] - hist_b[d]) for d in common_dates]
                vals = [s for _, s in hist_spreads]
                avg_s = round(sum(vals) / len(vals), 2)
                wavg_s = _weighted_avg(hist_spreads, latest_dt)
                min_s = round(min(vals), 2)
                max_s = round(max(vals), 2)
                pct = round(sum(1 for v in vals if v <= current_spread) / len(vals) * 100)
                return avg_s, wavg_s, min_s, max_s, pct, len(common_dates)
            return None

        # Build length spreads (vs 16' base)
        length_spreads = []
        lengths_to_check = ['8', '10', '12', '14', '18', '20']
        products_seen = set()
        for (prod, ln), price in latest.items():
            products_seen.add(prod)

        for prod in sorted(products_seen):
            base_key = (prod, '16')
            base_price = latest.get(base_key)
            if not base_price:
                continue
            base_agg = agg.get(base_key, {})
            for ln in lengths_to_check:
                key = (prod, ln)
                price = latest.get(key)
                if not price:
                    continue
                spread = round(price - base_price, 2)
                stats = _compute_spread_stats(hist.get(key, {}), hist.get(base_key, {}), spread, latest_date)
                if stats:
                    avg_s, wavg_s, min_s, max_s, pct, n = stats
                else:
                    a_data = agg.get(key, {})
                    b_data = base_agg
                    avg_s = round(a_data.get('avg', price) - b_data.get('avg', base_price), 2) if a_data and b_data else spread
                    wavg_s = avg_s
                    min_s = spread; max_s = spread; pct = 50; n = 0
                length_spreads.append({
                    'product': prod, 'length': ln, 'base': base_price, 'price': price,
                    'spread': spread, 'avg': avg_s, 'wavg': wavg_s, 'min': min_s, 'max': max_s, 'pct': pct, 'n': n
                })

        # Build dimension spreads (vs 2x4 base)
        dimension_spreads = []
        dims_to_check = ['2x6', '2x8', '2x10', '2x12']
        for ln in ['RL', '8', '10', '12', '14', '16', '18', '20']:
            # Find 2x4 base (try #2, then #1)
            base_key_2 = ('2x4#2', ln)
            base_key_1 = ('2x4#1', ln)
            base_price = latest.get(base_key_2) or latest.get(base_key_1)
            base_key = base_key_2 if latest.get(base_key_2) else base_key_1
            if not base_price:
                continue
            for dim in dims_to_check:
                key2 = (dim + '#2', ln)
                key1 = (dim + '#1', ln)
                key = key2 if latest.get(key2) else key1
                price = latest.get(key)
                if not price:
                    continue
                spread = round(price - base_price, 2)
                stats = _compute_spread_stats(hist.get(key, {}), hist.get(base_key, {}), spread, latest_date)
                if stats:
                    avg_s, wavg_s, min_s, max_s, pct, n = stats
                else:
                    a_data = agg.get(key, {})
                    b_data = agg.get(base_key, {})
                    avg_s = round(a_data.get('avg', price) - b_data.get('avg', base_price), 2) if a_data and b_data else spread
                    wavg_s = avg_s
                    min_s = spread; max_s = spread; pct = 50; n = 0
                dimension_spreads.append({
                    'length': ln, 'dim': dim, 'base': base_price, 'price': price,
                    'spread': spread, 'avg': avg_s, 'wavg': wavg_s, 'min': min_s, 'max': max_s, 'pct': pct, 'n': n
                })

        # Build grade spreads (#1 vs #2)
        grade_spreads = []
        for dim in ['2x4', '2x6', '2x8', '2x10', '2x12']:
            for ln in ['RL', '8', '10', '12', '14', '16', '18', '20']:
                p1 = latest.get((dim + '#1', ln))
                p2 = latest.get((dim + '#2', ln))
                if not p1 or not p2:
                    continue
                premium = round(p1 - p2, 2)
                stats = _compute_spread_stats(hist.get((dim + '#1', ln), {}), hist.get((dim + '#2', ln), {}), premium, latest_date)
                if stats:
                    avg_prem, wavg_prem, min_p, max_p, pct, n = stats
                else:
                    a1 = agg.get((dim + '#1', ln), {})
                    a2 = agg.get((dim + '#2', ln), {})
                    avg_prem = round(a1.get('avg', p1) - a2.get('avg', p2), 2) if a1 and a2 else premium
                    wavg_prem = avg_prem
                    min_p = premium; max_p = premium; pct = 50; n = 0
                grade_spreads.append({
                    'dim': dim, 'length': ln, 'p1': p1, 'p2': p2,
                    'premium': premium, 'avg': avg_prem, 'wavg': wavg_prem, 'min': min_p, 'max': max_p, 'pct': pct, 'n': n
                })

        # Week-over-week changes
        wow_changes = []
        if prev:
            for (prod, ln), curr_price in latest.items():
                if ln != 'RL':
                    continue
                prev_price = prev.get((prod, ln))
                if prev_price and curr_price != prev_price:
                    wow_changes.append({
                        'product': prod, 'curr': curr_price, 'prev': prev_price,
                        'chg': round(curr_price - prev_price, 2)
                    })
            wow_changes.sort(key=lambda x: abs(x['chg']), reverse=True)

        result = {
            'latest_date': latest_date,
            'prev_date': prev_date,
            'region': region,
            'exclude_covid': exclude_covid,
            'length_spreads': length_spreads,
            'dimension_spreads': dimension_spreads,
            'grade_spreads': grade_spreads,
            'wow_changes': wow_changes
        }
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/rl/backfill', methods=['GET'])
def rl_backfill():
    """Batch endpoint: returns S.rl-shaped entries for backfilling frontend state."""
    try:
        date_from = request.args.get('from', '')
        products_param = request.args.get('products', '2x4#2,2x6#2,2x8#2,2x10#2,2x12#2,2x4#1,2x6#1')
        products = [p.strip() for p in products_param.split(',') if p.strip()]

        cache_key = f"backfill_{date_from}_{products_param}"
        cached = get_rl_cached(cache_key)
        if cached is not None:
            return jsonify(cached)

        conn = get_mi_db()
        placeholders = ','.join('?' for _ in products)
        sql = f"SELECT date, region, product, price FROM rl_prices WHERE length='RL' AND product IN ({placeholders})"
        params = list(products)
        if date_from:
            sql += " AND date>=?"
            params.append(date_from)
        sql += " ORDER BY date"
        rows = conn.execute(sql, params).fetchall()
        conn.close()

        # Group by date → S.rl-shaped entries
        by_date = {}
        for r in rows:
            d = r['date']
            if d not in by_date:
                by_date[d] = {'date': d, 'west': {}, 'central': {}, 'east': {}}
            region = r['region']
            if region in by_date[d]:
                by_date[d][region][r['product']] = r['price']

        result = sorted(by_date.values(), key=lambda x: x['date'])
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =====================================================================
# FORECAST ENDPOINTS — Seasonal, Short-term, and Pricing Models
# =====================================================================

MONTH_NAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

@app.route('/api/forecast/seasonal', methods=['GET'])
def forecast_seasonal():
    """Compute seasonal indices from historical RL prices."""
    try:
        product = request.args.get('product', '2x4#2')
        region = request.args.get('region', 'west')
        years = int(request.args.get('years', 5))

        cache_key = f"seasonal_{product}_{region}_{years}"
        cached = get_rl_cached(cache_key)
        if cached is not None:
            return jsonify(cached)

        conn = get_mi_db()
        cutoff = (datetime.now() - timedelta(days=365 * years)).strftime('%Y-%m-%d')

        # Fetch RL-length prices only (composite prices, not specified lengths)
        rows = conn.execute(
            "SELECT date, price FROM rl_prices WHERE product=? AND region=? AND length='RL' AND date>=? ORDER BY date",
            (product, region, cutoff)
        ).fetchall()
        conn.close()

        if not rows or len(rows) < 24:
            return jsonify({'error': 'Insufficient data', 'dataPoints': len(rows) if rows else 0})

        prices = [float(r['price']) for r in rows]
        dates = [r['date'] for r in rows]
        n = len(prices)

        # Linear detrend: fit y = a*x + b via least squares
        x_mean = (n - 1) / 2.0
        y_mean = sum(prices) / n
        num = sum((i - x_mean) * (prices[i] - y_mean) for i in range(n))
        den = sum((i - x_mean) ** 2 for i in range(n))
        slope = num / den if den else 0
        intercept = y_mean - slope * x_mean
        trend_vals = [slope * i + intercept for i in range(n)]

        # Group detrended prices by calendar month
        monthly_raw = {}       # raw prices by month
        monthly_detrended = {}  # detrended prices by month
        for i, (date_str, price) in enumerate(zip(dates, prices)):
            month = int(date_str[5:7])
            monthly_raw.setdefault(month, []).append(price)
            monthly_detrended.setdefault(month, []).append(price - trend_vals[i])

        baseline = round(y_mean)
        overall_avg = y_mean

        # Compute monthly factors
        factors = []
        for month in range(1, 13):
            raw = monthly_raw.get(month, [])
            if not raw:
                factors.append({'month': month, 'name': MONTH_NAMES[month-1], 'avg': 0, 'index': 1.0, 'pctRank': 50, 'volatility': 0, 'count': 0})
                continue
            month_avg = sum(raw) / len(raw)
            month_std = statistics.stdev(raw) if len(raw) > 1 else 0
            index = round(month_avg / overall_avg, 3) if overall_avg else 1.0
            # Percentile rank: how does this month's avg compare to ALL prices?
            pct_rank = round(sum(1 for p in prices if p <= month_avg) / n * 100)
            factors.append({
                'month': month,
                'name': MONTH_NAMES[month-1],
                'avg': round(month_avg),
                'index': index,
                'pctRank': pct_rank,
                'volatility': round(month_std),
                'count': len(raw)
            })

        # Current position: where is the latest price vs this month's historical norm?
        current_month = datetime.now().month
        current_factor = next((f for f in factors if f['month'] == current_month), None)
        latest_price = prices[-1] if prices else None
        current_month_prices = monthly_raw.get(current_month, [])
        if latest_price and current_month_prices:
            price_pct = round(sum(1 for p in current_month_prices if p <= latest_price) / len(current_month_prices) * 100)
            if price_pct < 25:
                signal = 'well_below_seasonal'
            elif price_pct < 40:
                signal = 'below_seasonal'
            elif price_pct > 75:
                signal = 'well_above_seasonal'
            elif price_pct > 60:
                signal = 'above_seasonal'
            else:
                signal = 'at_seasonal_norm'
        else:
            price_pct = 50
            signal = 'unknown'

        # Seasonal outlook text
        peak_months = [f for f in factors if f['index'] > 1.02]
        low_months = [f for f in factors if f['index'] < 0.98]
        peak_names = ', '.join(f['name'] for f in sorted(peak_months, key=lambda x: x['index'], reverse=True)[:3])
        low_names = ', '.join(f['name'] for f in sorted(low_months, key=lambda x: x['index'])[:3])

        result = {
            'product': product,
            'region': region,
            'baseline': baseline,
            'monthlyFactors': factors,
            'currentPosition': {
                'month': current_month,
                'monthName': MONTH_NAMES[current_month - 1],
                'latestPrice': round(latest_price) if latest_price else None,
                'seasonalAvg': current_factor['avg'] if current_factor else None,
                'pctRank': price_pct,
                'signal': signal,
                'index': current_factor['index'] if current_factor else 1.0
            },
            'outlook': {
                'peakMonths': peak_names or 'None identified',
                'lowMonths': low_names or 'None identified',
                'trend': 'up' if slope > 0.5 else 'down' if slope < -0.5 else 'flat',
                'trendPerWeek': round(slope, 2)
            },
            'dataPoints': n,
            'period': f"{dates[0]} to {dates[-1]}"
        }
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forecast/shortterm', methods=['GET'])
def forecast_shortterm():
    """Short-term price forecast using Holt exponential smoothing + seasonal adjustment."""
    try:
        product = request.args.get('product', '2x4#2')
        region = request.args.get('region', 'west')
        weeks = int(request.args.get('weeks', 8))

        cache_key = f"forecast_{product}_{region}_{weeks}"
        cached = get_rl_cached(cache_key)
        if cached is not None:
            return jsonify(cached)

        conn = get_mi_db()
        # Fetch last 104 weeks (2 years) for smoothing + volatility
        cutoff = (datetime.now() - timedelta(days=730)).strftime('%Y-%m-%d')
        rows = conn.execute(
            "SELECT date, price FROM rl_prices WHERE product=? AND region=? AND length='RL' AND date>=? ORDER BY date",
            (product, region, cutoff)
        ).fetchall()

        # Also fetch 5yr seasonal factors inline
        cutoff_5y = (datetime.now() - timedelta(days=365 * 5)).strftime('%Y-%m-%d')
        seasonal_rows = conn.execute(
            "SELECT date, price FROM rl_prices WHERE product=? AND region=? AND length='RL' AND date>=? ORDER BY date",
            (product, region, cutoff_5y)
        ).fetchall()
        conn.close()

        prices = [float(r['price']) for r in rows]
        price_dates = [r['date'] for r in rows]

        if len(prices) < 12:
            return jsonify({'error': 'Insufficient data', 'dataPoints': len(prices)})

        # Compute seasonal factors from 5yr data
        seasonal_prices = [float(r['price']) for r in seasonal_rows]
        seasonal_dates = [r['date'] for r in seasonal_rows]
        s_avg = sum(seasonal_prices) / len(seasonal_prices) if seasonal_prices else 1
        monthly_avgs = {}
        for d, p in zip(seasonal_dates, seasonal_prices):
            monthly_avgs.setdefault(int(d[5:7]), []).append(p)
        seasonal_index = {}
        for m in range(1, 13):
            vals = monthly_avgs.get(m, [])
            seasonal_index[m] = (sum(vals) / len(vals)) / s_avg if vals and s_avg else 1.0

        # Holt exponential smoothing
        alpha, beta = 0.3, 0.1
        level = prices[0]
        trend = (prices[min(11, len(prices)-1)] - prices[0]) / min(11, len(prices)-1) if len(prices) > 1 else 0

        for p in prices[1:]:
            prev_level = level
            level = alpha * p + (1 - alpha) * (level + trend)
            trend = beta * (level - prev_level) + (1 - beta) * trend

        # Rolling volatility (last 12 data points)
        recent = prices[-12:] if len(prices) >= 12 else prices
        vol = statistics.stdev(recent) if len(recent) > 1 else 10

        # Momentum: 4-week avg vs 12-week avg
        avg4 = sum(prices[-4:]) / min(4, len(prices))
        avg12 = sum(prices[-12:]) / min(12, len(prices))
        momentum = round(avg4 - avg12)

        # Generate forecast using SMOOTHED RELATIVE seasonal adjustment
        # The Holt level already reflects current seasonal conditions,
        # so we adjust by the ratio of future month's index to current month's index.
        # We interpolate between monthly indices based on day-of-month to avoid jitter.
        last_date_dt = datetime.strptime(price_dates[-1], '%Y-%m-%d')
        current_month = last_date_dt.month if price_dates else datetime.now().month
        current_si = seasonal_index.get(current_month, 1.0)

        def smoothed_seasonal(dt):
            """Interpolate seasonal index between mid-month anchor points for smooth transitions."""
            m = dt.month
            d = dt.day
            si_this = seasonal_index.get(m, 1.0)
            if d <= 15:
                # Blend with previous month (transition into this month)
                prev_m = 12 if m == 1 else m - 1
                si_prev = seasonal_index.get(prev_m, 1.0)
                t = (d + 15) / 30.0  # 0.5 at day 1, 1.0 at day 15
                return si_prev * (1 - t) + si_this * t
            else:
                # Blend with next month (transition out of this month)
                next_m = 1 if m == 12 else m + 1
                si_next = seasonal_index.get(next_m, 1.0)
                t = (d - 15) / 30.0  # 0.0 at day 15, ~0.5 at day 30
                return si_this * (1 - t) + si_next * t

        forecast = []
        for w in range(1, weeks + 1):
            forecast_date = last_date_dt + timedelta(days=7 * w)
            pred = level + w * trend
            # Smoothed relative seasonal: interpolated target vs current
            target_si = smoothed_seasonal(forecast_date)
            ratio = target_si / current_si if current_si else 1.0
            # Cap seasonal swing at ±15% to prevent extreme jumps
            ratio = max(0.85, min(1.15, ratio))
            # Ramp seasonal effect gradually — week 1 is mostly trend,
            # full seasonal influence by the end of the horizon
            ramp = w / weeks  # 0.125 at w=1, 1.0 at w=8
            effective_ratio = 1.0 + ramp * (ratio - 1.0)
            pred_adj = pred * effective_ratio
            # Confidence widens with horizon
            width = 1.96 * vol * (1 + 0.15 * (w - 1))
            forecast.append({
                'date': forecast_date.strftime('%Y-%m-%d'),
                'price': round(pred_adj),
                'low': round(pred_adj - width),
                'high': round(pred_adj + width),
                'week': w
            })

        # Actual prices for chart context (last 26 weeks)
        actuals = []
        for d, p in zip(price_dates[-26:], prices[-26:]):
            actuals.append({'date': d, 'price': round(p)})

        # Seasonal outlook text
        now_month = datetime.now().month
        next_months = [(now_month + i - 1) % 12 + 1 for i in range(1, 4)]
        upcoming_indices = [seasonal_index.get(m, 1.0) for m in next_months]
        avg_upcoming = sum(upcoming_indices) / len(upcoming_indices)
        if avg_upcoming > 1.02:
            outlook = f"Entering seasonally strong period ({', '.join(MONTH_NAMES[m-1] for m in next_months)}). Prices typically above average."
        elif avg_upcoming < 0.98:
            outlook = f"Entering seasonally weak period ({', '.join(MONTH_NAMES[m-1] for m in next_months)}). Prices typically below average."
        else:
            outlook = f"Neutral seasonal period ahead ({', '.join(MONTH_NAMES[m-1] for m in next_months)})."

        result = {
            'product': product,
            'region': region,
            'lastPrice': round(prices[-1]),
            'trend': 'up' if trend > 0.5 else 'down' if trend < -0.5 else 'flat',
            'trendPerWeek': round(trend, 1),
            'momentum': momentum,
            'volatility': round(vol),
            'forecast': forecast,
            'actuals': actuals,
            'seasonalOutlook': outlook,
            'dataPoints': len(prices),
            'method': 'Holt exponential smoothing + seasonal adjustment'
        }
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forecast/pricing', methods=['POST'])
def forecast_pricing():
    """Customer pricing recommendation: best mill + freight + seasonal margin adjustment."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'JSON body required'}), 400

        customer = data.get('customer', '')
        destination = data.get('destination', '')
        products = data.get('products', ['2x4#2'])
        target_margin = float(data.get('targetMargin', 25))

        if not destination:
            return jsonify({'error': 'destination required'}), 400

        if isinstance(products, str):
            products = [products]

        recommendations = []

        # Only consider quotes from the last 2 business days for active quoting
        max_age_days = int(data.get('maxAgeDays', 2))
        quote_cutoff = business_day_cutoff(max_age_days)

        for product in products:
            # 1. Get best mill costs from mill_quotes (only recent)
            conn = get_mi_db()
            mill_rows = conn.execute("""
                SELECT mill_name, price, date FROM mill_quotes
                WHERE product=? AND price > 0 AND date >= ?
                ORDER BY date DESC
            """, (product, quote_cutoff)).fetchall()
            conn.close()

            # Deduplicate: latest price per mill
            seen_mills = {}
            for r in mill_rows:
                mill = r['mill_name']
                if mill not in seen_mills:
                    seen_mills[mill] = {'mill': mill, 'fob': float(r['price']), 'date': r['date']}

            if not seen_mills:
                recommendations.append({
                    'product': product,
                    'error': f'No mill pricing available within {max_age_days} days'
                })
                continue

            # 2. Calculate freight for each mill to destination
            candidates = []
            for mill_name, mill_data in seen_mills.items():
                # Resolve mill origin
                origin = ''
                dir_entry = MILL_DIRECTORY.get(mill_name)
                if dir_entry:
                    origin = f"{dir_entry[0]}, {dir_entry[1]}"
                else:
                    # Try partial match
                    for k, v in MILL_DIRECTORY.items():
                        if mill_name.startswith(k.split(' - ')[0]):
                            origin = f"{v[0]}, {v[1]}"
                            break

                if not origin or not destination:
                    continue

                # Calculate freight using existing mileage infrastructure
                try:
                    # Use cached lane data or geocoding
                    freight_per_mbf = 0
                    # Simplified freight: base + miles * rate / MBF per TL
                    miles = None
                    mi_conn = get_mi_db()
                    lane = mi_conn.execute(
                        "SELECT miles FROM lanes WHERE origin=? AND destination=?",
                        (origin, destination)
                    ).fetchone()
                    mi_conn.close()

                    if lane:
                        miles = float(lane['miles'])
                    else:
                        # Try geocoding
                        try:
                            coords_o = geocode_location(origin)
                            coords_d = geocode_location(destination)
                            if coords_o and coords_d:
                                route = get_driving_distance(coords_o, coords_d)
                                if route:
                                    miles = route.get('miles', route.get('distance', 0))
                        except Exception:
                            pass

                    if miles:
                        # Freight formula: (base + miles * rate) / MBF_per_TL
                        base = 450  # default freight base
                        rate = 2.25  # default per-mile rate
                        mbf_per_tl = 23
                        freight_per_mbf = round((base + miles * rate) / mbf_per_tl)
                    else:
                        freight_per_mbf = 20  # reasonable default
                except Exception:
                    freight_per_mbf = 20

                landed = mill_data['fob'] + freight_per_mbf
                candidates.append({
                    'mill': mill_name,
                    'fob': mill_data['fob'],
                    'freight': freight_per_mbf,
                    'landed': landed,
                    'date': mill_data['date']
                })

            if not candidates:
                recommendations.append({
                    'product': product,
                    'error': 'Could not calculate freight for any mill'
                })
                continue

            # Sort by landed cost
            candidates.sort(key=lambda c: c['landed'])
            best = candidates[0]

            # 3. Get seasonal position for margin adjustment
            seasonal_adj = 0
            seasonal_note = ''
            try:
                # Quick seasonal check inline
                mi_conn = get_mi_db()
                cutoff_5y = (datetime.now() - timedelta(days=365 * 5)).strftime('%Y-%m-%d')
                s_rows = mi_conn.execute(
                    "SELECT date, price FROM rl_prices WHERE product=? AND region='west' AND length='RL' AND date>=?",
                    (product, cutoff_5y)
                ).fetchall()
                mi_conn.close()

                if s_rows:
                    s_prices = [float(r['price']) for r in s_rows]
                    current_month = datetime.now().month
                    month_prices = [float(r['price']) for r in s_rows if int(r['date'][5:7]) == current_month]
                    if month_prices:
                        latest = s_prices[-1]
                        pct = round(sum(1 for p in month_prices if p <= latest) / len(month_prices) * 100)
                        if pct < 30:
                            seasonal_adj = -5
                            seasonal_note = f"Below seasonal norm ({pct}th %ile for {MONTH_NAMES[current_month-1]}) — tighter margin, good buying window"
                        elif pct > 70:
                            seasonal_adj = 5
                            seasonal_note = f"Above seasonal norm ({pct}th %ile for {MONTH_NAMES[current_month-1]}) — wider margin, prices elevated"
                        else:
                            seasonal_note = f"Near seasonal norm ({pct}th %ile for {MONTH_NAMES[current_month-1]})"
            except Exception:
                pass

            adjusted_margin = target_margin + seasonal_adj
            recommended_sell = best['landed'] + adjusted_margin

            rec = {
                'product': product,
                'bestMill': best['mill'],
                'fob': best['fob'],
                'freight': best['freight'],
                'landed': best['landed'],
                'targetMargin': target_margin,
                'seasonalAdj': seasonal_adj,
                'adjustedMargin': adjusted_margin,
                'recommendedSell': round(recommended_sell),
                'seasonalNote': seasonal_note,
                'millDate': best['date'],
                'alternatives': [{'mill': c['mill'], 'fob': c['fob'], 'freight': c['freight'], 'landed': c['landed']} for c in candidates[1:4]]
            }
            recommendations.append(rec)

        result = {
            'customer': customer,
            'destination': destination,
            'products': recommendations,
            'generatedAt': datetime.now().isoformat()
        }
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ----- MI: LANES -----

@app.route('/api/mi/lanes', methods=['GET'])
def mi_list_lanes():
    conn = get_mi_db()
    rows = conn.execute("SELECT * FROM lanes ORDER BY origin").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mi/lanes', methods=['POST'])

def mi_add_lane():
    data = request.get_json() or {}
    origin = data.get('origin', '')
    dest = data.get('dest', '')
    miles = data.get('miles', 0)
    if not origin or not dest:
        return jsonify({'error': 'origin and dest are required'}), 400
    try:
        conn = get_mi_db()
        conn.execute("INSERT OR REPLACE INTO lanes (origin, dest, miles) VALUES (?,?,?)",
                     (origin, dest, miles))
        conn.commit()
        conn.close()
        return jsonify({'ok': True}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ----- MI: MILEAGE -----

@app.route('/api/mi/mileage', methods=['POST'])
def mi_mileage_lookup():
    data = request.get_json()
    origin = data.get('origin', '')
    dest = data.get('dest', '')
    if not origin or not dest:
        return jsonify({'error': 'Missing origin or dest'}), 400
    conn = get_mi_db()
    lane = conn.execute("SELECT miles FROM lanes WHERE origin=? AND dest=?", (origin, dest)).fetchone()
    if lane:
        conn.close()
        return jsonify({'miles': lane['miles'], 'origin': origin, 'dest': dest})
    conn.close()
    origin_coords = mi_geocode_location(origin)
    if not origin_coords:
        return jsonify({'error': f'Could not geocode origin: {origin}'}), 404
    time.sleep(0.5)  # Nominatim 1 req/s; blocking is acceptable for internal tool
    dest_coords = mi_geocode_location(dest)
    if not dest_coords:
        return jsonify({'error': f'Could not geocode destination: {dest}'}), 404
    miles = mi_get_distance(origin_coords, dest_coords)
    if miles is None:
        return jsonify({'error': 'Could not calculate route'}), 404
    conn = get_mi_db()
    conn.execute("INSERT OR IGNORE INTO lanes (origin, dest, miles) VALUES (?,?,?)", (origin, dest, miles))
    conn.commit()
    conn.close()
    return jsonify({'miles': miles, 'origin': origin, 'dest': dest})

# ----- MI: SETTINGS -----

@app.route('/api/mi/settings', methods=['GET'])
def mi_get_settings():
    conn = get_mi_db()
    rows = conn.execute("SELECT * FROM settings").fetchall()
    conn.close()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/mi/settings', methods=['PUT'])

def mi_update_settings():
    data = request.get_json()
    conn = get_mi_db()
    for k, v in data.items():
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (k, str(v)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ==================== AUDIT TRAIL ====================

def _log_audit(user, action, entity_type, entity_id=None, entity_name=None,
               old_value=None, new_value=None, details=None, ip_address=None):
    """Log an action to the audit trail."""
    try:
        conn = get_crm_db()
        conn.execute('''
            INSERT INTO audit_log (user, action, entity_type, entity_id, entity_name,
                                   old_value, new_value, details, ip_address)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            user, action, entity_type,
            str(entity_id) if entity_id is not None else None,
            entity_name,
            json.dumps(old_value) if isinstance(old_value, (dict, list)) else old_value,
            json.dumps(new_value) if isinstance(new_value, (dict, list)) else new_value,
            details,
            ip_address
        ))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[audit_log ERROR] {e}", flush=True)


@app.route('/api/audit/log', methods=['POST'])

def create_audit_entry():
    """Accept audit log entries from the frontend."""
    try:
        data = request.get_json() or {}
        _log_audit(
            user=data.get('user', g.user),
            action=data.get('action', 'unknown'),
            entity_type=data.get('entity_type', 'unknown'),
            entity_id=data.get('entity_id'),
            entity_name=data.get('entity_name'),
            old_value=data.get('old_value'),
            new_value=data.get('new_value'),
            details=data.get('details'),
            ip_address=request.remote_addr
        )
        return jsonify({'ok': True}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Intelligence Endpoints ──────────────────────────────────────────────────

@app.route('/api/intelligence/mill-moves', methods=['GET'])
def intel_mill_moves():
    """Recent mill price changes from the mill_price_changes table."""
    try:
        days = int(request.args.get('days', 30))
        product = request.args.get('product', '').strip()
        mill = request.args.get('mill', '').strip()
        conn = get_mi_db()
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        sql = "SELECT * FROM mill_price_changes WHERE date >= ?"
        params = [cutoff]
        if product:
            sql += " AND UPPER(product)=?"
            params.append(product.upper())
        if mill:
            sql += " AND UPPER(mill_name) LIKE ?"
            params.append(f"%{mill.upper()}%")
        sql += " ORDER BY date DESC, mill_name"
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        changes = [dict(r) for r in rows]
        # Summary stats
        total = len(changes)
        up_count = sum(1 for c in changes if (c.get('change') or 0) > 0)
        down_count = sum(1 for c in changes if (c.get('change') or 0) < 0)
        avg_change = round(sum(c.get('change', 0) or 0 for c in changes) / total, 2) if total else 0
        # Most active mills
        mill_counts = {}
        for c in changes:
            mn = c.get('mill_name', '')
            mill_counts[mn] = mill_counts.get(mn, 0) + 1
        most_active = sorted(mill_counts.items(), key=lambda x: -x[1])[:5]
        return jsonify({
            'changes': changes,
            'summary': {
                'total': total,
                'up': up_count,
                'down': down_count,
                'avgChange': avg_change,
                'mostActive': [{'mill': m, 'count': c} for m, c in most_active]
            },
            'days': days
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/intelligence/regime', methods=['GET'])
def intel_regime():
    """Market regime detection using ROC on RL benchmark prices."""
    try:
        region = request.args.get('region', 'west').strip()
        product = request.args.get('product', '2x4#2').strip()
        cache_key = f"regime_{region}_{product}"
        cached = get_rl_cached(cache_key)
        if cached:
            return jsonify(cached)

        conn = get_mi_db()
        # Get last 90 days of RL prices (need buffer for 8-week ROC)
        cutoff = (datetime.now() - timedelta(days=120)).strftime('%Y-%m-%d')
        rows = conn.execute(
            """SELECT date, price FROM rl_prices
               WHERE region=? AND product=? AND length='RL' AND date>=?
               ORDER BY date ASC""",
            (region, product, cutoff)
        ).fetchall()
        conn.close()

        if len(rows) < 5:
            return jsonify({'error': 'Not enough RL data for regime detection', 'regime': 'Unknown', 'confidence': 0})

        prices = [(r['date'], r['price']) for r in rows]
        current_price = prices[-1][1]
        current_date = prices[-1][0]

        # Calculate ROC at 3 horizons (approximate trading days)
        def _get_price_at_offset(prices_list, offset_days):
            target_date = (datetime.strptime(prices_list[-1][0], '%Y-%m-%d') - timedelta(days=offset_days)).strftime('%Y-%m-%d')
            # Find closest date at or before target
            best = None
            for d, p in prices_list:
                if d <= target_date:
                    best = p
            return best

        p_2wk = _get_price_at_offset(prices, 14)
        p_4wk = _get_price_at_offset(prices, 28)
        p_8wk = _get_price_at_offset(prices, 56)

        roc_2wk = round(((current_price - p_2wk) / p_2wk) * 100, 2) if p_2wk else 0
        roc_4wk = round(((current_price - p_4wk) / p_4wk) * 100, 2) if p_4wk else 0
        roc_8wk = round(((current_price - p_8wk) / p_8wk) * 100, 2) if p_8wk else 0

        chg_2wk = round(current_price - p_2wk, 2) if p_2wk else 0
        chg_4wk = round(current_price - p_4wk, 2) if p_4wk else 0
        chg_8wk = round(current_price - p_8wk, 2) if p_8wk else 0

        # Classify regime
        if roc_2wk > 2 and roc_4wk > 3:
            regime = 'Rally'
            confidence = min(100, int(40 + abs(roc_2wk) * 8 + abs(roc_4wk) * 5))
            bias = f"Prices up ${chg_4wk}/MBF over 4 weeks. Momentum supports higher prices near-term."
            trading = "Favor buying on dips — strong upward momentum."
        elif roc_2wk < 1 and roc_4wk > 2:
            regime = 'Topping'
            confidence = min(100, int(35 + abs(roc_4wk - roc_2wk) * 10))
            bias = f"Momentum fading — 2wk change slowing to {roc_2wk}% while 4wk still +{roc_4wk}%."
            trading = "Consider locking in sales at current levels. Upside may be limited."
        elif roc_2wk < -2 and roc_4wk < -2:
            regime = 'Decline'
            confidence = min(100, int(40 + abs(roc_2wk) * 8 + abs(roc_4wk) * 5))
            bias = f"Prices down ${abs(chg_4wk)}/MBF over 4 weeks. Downward pressure continues."
            trading = "Delay purchases if possible. Consider selling inventory at current levels."
        elif roc_2wk > -1 and roc_4wk < -2:
            regime = 'Bottoming'
            confidence = min(100, int(35 + abs(roc_4wk - roc_2wk) * 10))
            bias = f"Decline losing steam — 2wk change recovering to {roc_2wk}% while 4wk still {roc_4wk}%."
            trading = "Watch for buying opportunities. Market may be finding a floor."
        else:
            regime = 'Choppy'
            confidence = max(20, int(50 - abs(roc_2wk) * 5 - abs(roc_4wk) * 3))
            bias = f"No clear trend — 2wk {'+' if roc_2wk >= 0 else ''}{roc_2wk}%, 4wk {'+' if roc_4wk >= 0 else ''}{roc_4wk}%."
            trading = "Range-bound market. Trade tactically around spread opportunities."

        confidence = max(10, min(95, confidence))

        result = {
            'regime': regime,
            'confidence': confidence,
            'roc': {'2wk': roc_2wk, '4wk': roc_4wk, '8wk': roc_8wk},
            'changes': {'2wk': chg_2wk, '4wk': chg_4wk, '8wk': chg_8wk},
            'currentPrice': current_price,
            'currentDate': current_date,
            'product': product,
            'region': region,
            'context': bias,
            'tradingBias': trading,
            'priceHistory': [{'date': d, 'price': p} for d, p in prices[-20:]]  # Last 20 data points
        }
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/intelligence/spread-signals', methods=['GET'])
def intel_spread_signals():
    """Spread mean-reversion signals — flags extreme percentile spreads with reversion probability."""
    try:
        region = request.args.get('region', 'west').strip()
        spread_type = request.args.get('type', 'all').strip()  # dimension, length, grade, zone, all
        cache_key = f"spread_signals_{region}_{spread_type}"
        cached = get_rl_cached(cache_key)
        if cached:
            return jsonify(cached)

        conn = get_mi_db()
        try:
            # Get current and 5-year historical RL data for this region
            five_yr_ago = (datetime.now() - timedelta(days=5*365)).strftime('%Y-%m-%d')
            rows = conn.execute(
                """SELECT date, product, length, price FROM rl_prices
                   WHERE region=? AND date>=? ORDER BY date""",
                (region, five_yr_ago)
            ).fetchall()
        finally:
            conn.close()

        # Filter to complete dates only (≥10 rows per date)
        from collections import Counter
        date_counts = Counter(r['date'] for r in rows)
        complete_dates = set(d for d, c in date_counts.items() if c >= 10)

        # Build lookup: (product, length) -> {date: price}
        hist = {}
        for r in rows:
            if r['date'] not in complete_dates:
                continue
            key = (r['product'], r['length'])
            if key not in hist:
                hist[key] = {}
            hist[key][r['date']] = r['price']

        # Get latest date
        all_dates = sorted(complete_dates)
        if not all_dates:
            return jsonify({'signals': [], 'signalCount': 0, 'region': region})
        latest_dt = all_dates[-1]

        # Get current regime for context
        regime_data = None
        try:
            regime_cache = get_rl_cached(f"regime_{region}_2x4#2")
            if regime_cache:
                regime_data = regime_cache
        except Exception:
            pass
        current_regime = regime_data['regime'] if regime_data else 'Unknown'

        signals = []
        EXTREME_LOW = 10
        EXTREME_HIGH = 90

        def _check_spread(name, key_a, key_b, spread_category):
            """Check if a spread is at extreme percentile and compute reversion probability."""
            if key_a not in hist or key_b not in hist:
                return
            hist_a = hist[key_a]
            hist_b = hist[key_b]
            # Current spread
            if latest_dt not in hist_a or latest_dt not in hist_b:
                return
            current = round(hist_a[latest_dt] - hist_b[latest_dt], 2)

            # Historical spreads on common dates
            common = sorted(set(hist_a.keys()) & set(hist_b.keys()))
            if len(common) < 20:
                return
            hist_vals = [(d, hist_a[d] - hist_b[d]) for d in common]
            vals = [v for _, v in hist_vals]
            avg_s = round(sum(vals) / len(vals), 2)
            pct = round(sum(1 for v in vals if v <= current) / len(vals) * 100)

            if pct > EXTREME_LOW and pct < EXTREME_HIGH:
                return  # Not extreme — no signal

            # Compute reversion probability: how often did extreme spreads revert toward mean within 4 weeks?
            bucket_low = 0 if pct <= EXTREME_LOW else 90
            bucket_high = 10 if pct <= EXTREME_LOW else 100
            revert_count = 0
            total_instances = 0
            for i, (d, s) in enumerate(hist_vals):
                # Check if this historical point was in the same percentile bucket
                rank = sum(1 for v in vals if v <= s) / len(vals) * 100
                if rank >= bucket_low and rank <= bucket_high:
                    total_instances += 1
                    # Look ahead ~4 weeks (20 trading days ≈ 4-5 data points in weekly data)
                    look_ahead = min(i + 5, len(hist_vals) - 1)
                    if look_ahead > i:
                        future_s = hist_vals[look_ahead][1]
                        # Did it revert toward mean?
                        if pct <= EXTREME_LOW and future_s > s:  # Was low, moved up
                            revert_count += 1
                        elif pct >= EXTREME_HIGH and future_s < s:  # Was high, moved down
                            revert_count += 1

            reversion_prob = round((revert_count / total_instances) * 100) if total_instances > 5 else None

            direction = 'narrow' if abs(current) > abs(avg_s) else 'widen'
            if pct <= EXTREME_LOW:
                context = f"{name} spread is at {pct}th percentile (historically low)."
                if reversion_prob:
                    context += f" {reversion_prob}% chance of reverting within 4 weeks."
                actionable = f"Spread likely to {direction}. Watch for mean-reversion opportunity."
            else:
                context = f"{name} spread is at {pct}th percentile (historically high)."
                if reversion_prob:
                    context += f" {reversion_prob}% chance of reverting within 4 weeks."
                actionable = f"Spread likely to {direction}. Consider position adjustment."

            # Weighted avg
            import math
            HALF_LIFE = 180
            decay_c = math.log(2) / HALF_LIFE
            latest_ord = datetime.strptime(latest_dt, '%Y-%m-%d').toordinal()
            w_sum = 0.0; w_total = 0.0
            for d, s in hist_vals:
                try:
                    age = latest_ord - datetime.strptime(d, '%Y-%m-%d').toordinal()
                    w = math.exp(-decay_c * age)
                    w_sum += w * s; w_total += w
                except Exception:
                    pass
            wavg = round(w_sum / w_total, 2) if w_total > 0 else avg_s

            signals.append({
                'spread': name,
                'category': spread_category,
                'current': current,
                'avg': avg_s,
                'wavg': wavg,
                'percentile': pct,
                'direction': direction,
                'reversionProb': reversion_prob,
                'regime': current_regime,
                'context': context,
                'actionable': actionable,
                'n': len(common)
            })

        # Check dimension spreads (vs 2x4)
        if spread_type in ('dimension', 'all'):
            for dim in ['2x6', '2x8', '2x10', '2x12']:
                _check_spread(f"{dim} vs 2x4", (f"{dim}#2", 'RL'), ('2x4#2', 'RL'), 'dimension')

        # Check length spreads (vs 16')
        if spread_type in ('length', 'all'):
            for prod in ['2x4#2', '2x6#2']:
                for ln in ['8', '10', '12', '14', '20']:
                    if (prod, ln) in hist and (prod, '16') in hist:
                        _check_spread(f"{prod} {ln}' vs 16'", (prod, ln), (prod, '16'), 'length')

        # Check grade spreads (#1 vs #2)
        if spread_type in ('grade', 'all'):
            for dim in ['2x4', '2x6', '2x8', '2x10', '2x12']:
                _check_spread(f"{dim}#1 vs {dim}#2", (f"{dim}#1", 'RL'), (f"{dim}#2", 'RL'), 'grade')

        # Check cross-zone (inter-region) spreads
        if spread_type in ('zone', 'all'):
            # Load data for the other two regions
            other_regions = [r for r in ['west', 'central', 'east'] if r != region]
            zone_hist = {region: hist}  # reuse already-loaded data for primary region
            for oreg in other_regions:
                conn2 = get_mi_db()
                try:
                    orows = conn2.execute(
                        """SELECT date, product, length, price FROM rl_prices
                           WHERE region=? AND date>=? ORDER BY date""",
                        (oreg, five_yr_ago)
                    ).fetchall()
                finally:
                    conn2.close()
                oh = {}
                for r in orows:
                    if r['date'] not in complete_dates:
                        continue
                    key = (r['product'], r['length'])
                    if key not in oh:
                        oh[key] = {}
                    oh[key][r['date']] = r['price']
                zone_hist[oreg] = oh

            # Build cross-zone spread checks for key products
            zone_products = ['2x4#2', '2x6#2', '2x4#3', '2x6#3', '2x10#2', '2x4 MSR', '2x6 MSR']
            zone_pairs = [('west', 'central'), ('west', 'east'), ('central', 'east')]

            # Save/restore hist for _check_spread since it reads from outer `hist`
            orig_hist = hist
            for prod in zone_products:
                for reg_a, reg_b in zone_pairs:
                    h_a = zone_hist.get(reg_a, {})
                    h_b = zone_hist.get(reg_b, {})
                    key_a = (prod, 'RL')
                    key_b = (prod, 'RL')
                    if key_a not in h_a or key_b not in h_b:
                        continue
                    # Merge into hist temporarily so _check_spread can read them
                    fake_a = (f"_zone_{reg_a}_{prod}", 'RL')
                    fake_b = (f"_zone_{reg_b}_{prod}", 'RL')
                    hist[fake_a] = h_a[key_a]
                    hist[fake_b] = h_b[key_b]
                    label = f"{prod} {reg_a.title()} vs {reg_b.title()}"
                    _check_spread(label, fake_a, fake_b, 'zone')
                    # Clean up
                    del hist[fake_a]
                    del hist[fake_b]
            hist = orig_hist

        # Sort by extremity (most extreme percentile first)
        signals.sort(key=lambda s: min(s['percentile'], 100 - s['percentile']))

        result = {
            'signals': signals,
            'signalCount': len(signals),
            'region': region,
            'regime': current_regime,
            'asOf': latest_dt
        }
        set_rl_cache(cache_key, result)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── End Intelligence Endpoints ─────────────────────────────────────────────

@app.route('/api/audit/log', methods=['GET'])

def list_audit_log():
    """Paginated audit log retrieval with filters."""
    try:
        conn = get_crm_db()
        conditions = ['1=1']
        params = []

        entity_type = request.args.get('entity_type')
        if entity_type:
            conditions.append('entity_type = ?')
            params.append(entity_type)

        user = request.args.get('user')
        if user:
            conditions.append('user = ?')
            params.append(user)

        action = request.args.get('action')
        if action:
            conditions.append('action = ?')
            params.append(action)

        date_from = request.args.get('from')
        if date_from:
            conditions.append('timestamp >= ?')
            params.append(date_from)

        date_to = request.args.get('to')
        if date_to:
            conditions.append('timestamp <= ?')
            params.append(date_to)

        entity_id = request.args.get('entity_id')
        if entity_id:
            conditions.append('entity_id = ?')
            params.append(entity_id)

        try:
            page = max(1, int(request.args.get('page', 1)))
        except (ValueError, TypeError):
            page = 1
        try:
            per_page = min(200, max(1, int(request.args.get('per_page', 50))))
        except (ValueError, TypeError):
            per_page = 50

        offset = (page - 1) * per_page
        where = ' AND '.join(conditions)

        total = conn.execute(f'SELECT COUNT(*) FROM audit_log WHERE {where}', params).fetchone()[0]
        rows = conn.execute(
            f'SELECT * FROM audit_log WHERE {where} ORDER BY timestamp DESC LIMIT ? OFFSET ?',
            params + [per_page, offset]
        ).fetchall()
        conn.close()

        return jsonify({
            'entries': [dict(r) for r in rows],
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': math.ceil(total / per_page) if total else 0
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== TRADE STATUS WORKFLOW ====================

VALID_TRADE_STATUSES = ['draft', 'pending', 'approved', 'confirmed', 'shipped', 'delivered', 'settled', 'cancelled']
TRADE_STATUS_FLOW = {
    'draft': ['pending', 'cancelled'],
    'pending': ['approved', 'cancelled'],
    'approved': ['confirmed', 'cancelled'],
    'confirmed': ['shipped', 'cancelled'],
    'shipped': ['delivered', 'cancelled'],
    'delivered': ['settled'],
    'settled': [],
    'cancelled': []
}


@app.route('/api/trades/status', methods=['POST'])

def upsert_trade_status():
    """Create or update a trade status record."""
    try:
        data = request.get_json() or {}
        trade_id = data.get('trade_id', '').strip()
        trade_type = data.get('trade_type', '').strip()
        status = data.get('status', 'draft').strip()

        if not trade_id or not trade_type:
            return jsonify({'error': 'trade_id and trade_type required'}), 400
        if status not in VALID_TRADE_STATUSES:
            return jsonify({'error': f'Invalid status. Must be one of: {", ".join(VALID_TRADE_STATUSES)}'}), 400

        conn = get_crm_db()
        existing = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()

        if existing:
            old_status = existing['status']
            conn.execute('''
                UPDATE trade_status SET status = ?, trade_type = ?, assigned_to = ?,
                       notes = ?, updated_at = datetime('now')
                WHERE trade_id = ?
            ''', (status, trade_type, data.get('assigned_to'), data.get('notes'), trade_id))
            conn.commit()
            row = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
            conn.close()
            _log_audit(g.user, 'trade_status_update', 'trade', trade_id,
                       details=f'Status changed from {old_status} to {status}',
                       old_value=old_status, new_value=status,
                       ip_address=request.remote_addr)
            return jsonify(dict(row))
        else:
            conn.execute('''
                INSERT INTO trade_status (trade_id, trade_type, status, assigned_to, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (trade_id, trade_type, status, data.get('assigned_to'), data.get('notes')))
            conn.commit()
            row = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
            conn.close()
            _log_audit(g.user, 'trade_status_create', 'trade', trade_id,
                       details=f'Trade created with status {status}',
                       new_value=status, ip_address=request.remote_addr)
            return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trades/status/<trade_id>', methods=['GET'])

def get_trade_status(trade_id):
    """Get status for a specific trade."""
    try:
        conn = get_crm_db()
        row = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({'error': 'Trade not found'}), 404
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trades/status', methods=['GET'])

def list_trade_statuses():
    """List all trade statuses with optional filters."""
    try:
        conn = get_crm_db()
        conditions = ['1=1']
        params = []

        status = request.args.get('status')
        if status:
            conditions.append('status = ?')
            params.append(status)

        trade_type = request.args.get('trade_type')
        if trade_type:
            conditions.append('trade_type = ?')
            params.append(trade_type)

        assigned_to = request.args.get('assigned_to')
        if assigned_to:
            conditions.append('assigned_to = ?')
            params.append(assigned_to)

        where = ' AND '.join(conditions)
        rows = conn.execute(
            f'SELECT * FROM trade_status WHERE {where} ORDER BY updated_at DESC LIMIT 500',
            params
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trades/<trade_id>/approve', methods=['POST'])

def approve_trade(trade_id):
    """Approve a trade (admin or senior trader)."""
    try:
        user = g.user.lower()
        if user not in ADMIN_USERS:
            return jsonify({'error': 'Only admin/senior traders can approve trades'}), 403

        conn = get_crm_db()
        row = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Trade not found'}), 404

        if row['status'] != 'pending':
            conn.close()
            return jsonify({'error': f'Cannot approve trade in "{row["status"]}" status. Must be "pending".'}), 400

        conn.execute('''
            UPDATE trade_status SET status = 'approved', approved_by = ?, approved_at = datetime('now'),
                   updated_at = datetime('now')
            WHERE trade_id = ?
        ''', (g.user, trade_id))
        conn.commit()
        updated = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
        conn.close()

        _log_audit(g.user, 'trade_approve', 'trade', trade_id,
                   details=f'Trade approved by {g.user}',
                   old_value='pending', new_value='approved',
                   ip_address=request.remote_addr)
        return jsonify(dict(updated))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/trades/<trade_id>/advance', methods=['POST'])

def advance_trade(trade_id):
    """Advance a trade to the next status in the workflow."""
    try:
        conn = get_crm_db()
        row = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Trade not found'}), 404

        current = row['status']
        allowed_next = TRADE_STATUS_FLOW.get(current, [])
        if not allowed_next:
            conn.close()
            return jsonify({'error': f'Trade in "{current}" status cannot be advanced.'}), 400

        # Use requested status if provided and valid, otherwise take the first allowed
        data = request.get_json() or {}
        requested = data.get('status', '').strip()
        if requested:
            if requested not in allowed_next:
                conn.close()
                return jsonify({
                    'error': f'Cannot move from "{current}" to "{requested}". Allowed: {", ".join(allowed_next)}'
                }), 400
            next_status = requested
        else:
            # Default: first non-cancelled option, or cancelled if that's the only option
            next_status = next((s for s in allowed_next if s != 'cancelled'), allowed_next[0])

        # Approval requires admin
        if next_status == 'approved':
            user = g.user.lower()
            if user not in ADMIN_USERS:
                conn.close()
                return jsonify({'error': 'Only admin/senior traders can approve trades'}), 403
            conn.execute('''
                UPDATE trade_status SET status = ?, approved_by = ?, approved_at = datetime('now'),
                       updated_at = datetime('now'), notes = COALESCE(?, notes)
                WHERE trade_id = ?
            ''', (next_status, g.user, data.get('notes'), trade_id))
        else:
            conn.execute('''
                UPDATE trade_status SET status = ?, updated_at = datetime('now'),
                       notes = COALESCE(?, notes)
                WHERE trade_id = ?
            ''', (next_status, data.get('notes'), trade_id))

        conn.commit()
        updated = conn.execute('SELECT * FROM trade_status WHERE trade_id = ?', (trade_id,)).fetchone()
        conn.close()

        _log_audit(g.user, 'trade_advance', 'trade', trade_id,
                   details=f'Trade advanced from {current} to {next_status}',
                   old_value=current, new_value=next_status,
                   ip_address=request.remote_addr)
        return jsonify(dict(updated))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== TRADE VALIDATION ====================

VALID_PRODUCTS = [
    '2x4#1', '2x6#1', '2x8#1', '2x10#1', '2x12#1',
    '2x4#2', '2x6#2', '2x8#2', '2x10#2', '2x12#2',
    '2x4#3', '2x6#3', '2x8#3', '2x10#3', '2x12#3',
    '2x4#4', '2x6#4', '2x8#4', '2x10#4', '2x12#4',
    '2x4 MSR', '2x6 MSR', '2x8 MSR', '2x10 MSR', '2x12 MSR',
]

@app.route('/api/trades/validate', methods=['POST'])

def validate_trade():
    """Validate a trade before saving."""
    try:
        data = request.get_json() or {}
        errors = []
        warnings = []

        # Required fields
        if not data.get('product'):
            errors.append('Product is required')
        elif data['product'] not in VALID_PRODUCTS:
            warnings.append(f'Product "{data["product"]}" is not in the standard product list')

        volume = data.get('volume')
        if volume is None:
            errors.append('Volume is required')
        else:
            try:
                vol = float(volume)
                if vol <= 0:
                    errors.append('Volume must be greater than 0')
            except (ValueError, TypeError):
                errors.append('Volume must be a number')

        price = data.get('price')
        if price is None:
            errors.append('Price is required')
        else:
            try:
                p = float(price)
                if p <= 0:
                    errors.append('Price must be greater than 0')
                elif p < 100:
                    warnings.append(f'Price ${p} seems unusually low')
                elif p > 1000:
                    warnings.append(f'Price ${p} seems unusually high')
            except (ValueError, TypeError):
                errors.append('Price must be a number')

        if not data.get('date'):
            errors.append('Date is required')

        # Customer/mill CRM validation
        customer = data.get('customer', '').strip()
        mill = data.get('mill', '').strip()

        if customer:
            conn = get_crm_db()
            cust_row = conn.execute(
                'SELECT id FROM customers WHERE UPPER(name) = UPPER(?)', (customer,)
            ).fetchone()
            if not cust_row:
                warnings.append(f'Customer "{customer}" not found in CRM')
            conn.close()

        if mill:
            conn = get_crm_db()
            mill_row = conn.execute(
                'SELECT id FROM mills WHERE UPPER(name) = UPPER(?)', (mill,)
            ).fetchone()
            if not mill_row:
                warnings.append(f'Mill "{mill}" not found in CRM')
            conn.close()

        # Credit check
        if customer and not errors:
            try:
                conn = get_crm_db()
                credit = conn.execute(
                    'SELECT * FROM credit_limits WHERE UPPER(customer_name) = UPPER(?)',
                    (customer,)
                ).fetchone()
                conn.close()
                if credit:
                    trade_value = float(data.get('volume', 0)) * float(data.get('price', 0))
                    new_exposure = (credit['current_exposure'] or 0) + trade_value
                    limit = credit['credit_limit'] or 0
                    if limit > 0 and new_exposure > limit:
                        warnings.append(
                            f'Credit limit exceeded: exposure would be ${new_exposure:,.0f} vs limit ${limit:,.0f}'
                        )
                    elif limit > 0 and new_exposure > limit * 0.8:
                        warnings.append(
                            f'Near credit limit: exposure would be ${new_exposure:,.0f} (limit ${limit:,.0f})'
                        )
            except Exception:
                pass

        return jsonify({
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== CREDIT MANAGEMENT ====================

@app.route('/api/credit/<customer>', methods=['GET'])

def get_credit(customer):
    """Get credit status for a customer."""
    try:
        conn = get_crm_db()
        row = conn.execute(
            'SELECT * FROM credit_limits WHERE UPPER(customer_name) = UPPER(?)',
            (customer,)
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({
                'customer_name': customer,
                'credit_limit': 0,
                'current_exposure': 0,
                'payment_terms': 'Net 30',
                'last_payment_date': None,
                'notes': None,
                'exists': False
            })
        result = dict(row)
        result['exists'] = True
        result['available'] = max(0, (result['credit_limit'] or 0) - (result['current_exposure'] or 0))
        result['utilization'] = round(
            (result['current_exposure'] or 0) / result['credit_limit'] * 100, 1
        ) if result['credit_limit'] else 0
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/credit/<customer>', methods=['PUT'])

def update_credit(customer):
    """Update credit limit and terms for a customer."""
    try:
        data = request.get_json() or {}
        conn = get_crm_db()
        existing = conn.execute(
            'SELECT * FROM credit_limits WHERE UPPER(customer_name) = UPPER(?)',
            (customer,)
        ).fetchone()

        if existing:
            old_limit = existing['credit_limit']
            old_terms = existing['payment_terms']
            # SECURITY: only update fields from hardcoded allowlist
            allowed = ['credit_limit', 'current_exposure', 'payment_terms', 'last_payment_date', 'notes']
            fields = [f for f in allowed if f in data]
            if not fields:
                conn.close()
                return jsonify({'error': 'No fields to update'}), 400
            set_clause = ', '.join(f'{f} = ?' for f in fields) + ', updated_at = datetime(\'now\')'
            values = [data[f] for f in fields] + [existing['id']]
            conn.execute(f'UPDATE credit_limits SET {set_clause} WHERE id = ?', values)
            conn.commit()
            row = conn.execute('SELECT * FROM credit_limits WHERE id = ?', (existing['id'],)).fetchone()
            conn.close()
            _log_audit(g.user, 'credit_update', 'credit', customer, customer,
                       old_value={'limit': old_limit, 'terms': old_terms},
                       new_value={'limit': data.get('credit_limit', old_limit),
                                  'terms': data.get('payment_terms', old_terms)},
                       ip_address=request.remote_addr)
            return jsonify(dict(row))
        else:
            conn.execute('''
                INSERT INTO credit_limits (customer_name, credit_limit, current_exposure,
                                           payment_terms, last_payment_date, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                customer,
                data.get('credit_limit', 0),
                data.get('current_exposure', 0),
                data.get('payment_terms', 'Net 30'),
                data.get('last_payment_date'),
                data.get('notes')
            ))
            conn.commit()
            row = conn.execute(
                'SELECT * FROM credit_limits WHERE UPPER(customer_name) = UPPER(?)',
                (customer,)
            ).fetchone()
            conn.close()
            _log_audit(g.user, 'credit_create', 'credit', customer, customer,
                       new_value={'limit': data.get('credit_limit', 0),
                                  'terms': data.get('payment_terms', 'Net 30')},
                       ip_address=request.remote_addr)
            return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/credit/summary', methods=['GET'])

def credit_summary():
    """List all customers with their credit exposure vs limit."""
    try:
        conn = get_crm_db()
        rows = conn.execute(
            'SELECT * FROM credit_limits ORDER BY customer_name'
        ).fetchall()
        conn.close()
        result = []
        for r in rows:
            d = dict(r)
            d['available'] = max(0, (d['credit_limit'] or 0) - (d['current_exposure'] or 0))
            d['utilization'] = round(
                (d['current_exposure'] or 0) / d['credit_limit'] * 100, 1
            ) if d['credit_limit'] else 0
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== FREIGHT RECONCILIATION ====================

@app.route('/api/freight/reconcile', methods=['POST'])

def freight_reconcile():
    """Compare estimated vs actual freight for a trade or set of trades."""
    try:
        data = request.get_json() or {}
        trades = data.get('trades', [data] if 'trade_id' in data else [])
        results = []

        for trade in trades:
            trade_id = trade.get('trade_id', '')
            estimated = trade.get('estimated_freight')
            actual = trade.get('actual_freight')

            if estimated is None or actual is None:
                results.append({
                    'trade_id': trade_id,
                    'error': 'Both estimated_freight and actual_freight required'
                })
                continue

            try:
                est = float(estimated)
                act = float(actual)
            except (ValueError, TypeError):
                results.append({
                    'trade_id': trade_id,
                    'error': 'Freight values must be numbers'
                })
                continue

            variance = act - est
            variance_pct = round((variance / est * 100), 2) if est != 0 else 0

            entry = {
                'trade_id': trade_id,
                'estimated_freight': est,
                'actual_freight': act,
                'variance': round(variance, 2),
                'variance_pct': variance_pct,
                'status': 'over' if variance > 0 else 'under' if variance < 0 else 'match'
            }
            results.append(entry)

            _log_audit(g.user, 'freight_reconcile', 'freight', trade_id,
                       details=f'Variance: ${variance:+.2f} ({variance_pct:+.1f}%)',
                       old_value=str(est), new_value=str(act),
                       ip_address=request.remote_addr)

        total_estimated = sum(r.get('estimated_freight', 0) for r in results if 'error' not in r)
        total_actual = sum(r.get('actual_freight', 0) for r in results if 'error' not in r)
        total_variance = round(total_actual - total_estimated, 2)

        return jsonify({
            'trades': results,
            'summary': {
                'count': len([r for r in results if 'error' not in r]),
                'total_estimated': round(total_estimated, 2),
                'total_actual': round(total_actual, 2),
                'total_variance': total_variance,
                'total_variance_pct': round(
                    (total_variance / total_estimated * 100), 2
                ) if total_estimated else 0
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/freight/variance', methods=['GET'])

def freight_variance():
    """List freight variances from audit log entries."""
    try:
        conn = get_crm_db()
        conditions = ["action = 'freight_reconcile'"]
        params = []

        date_from = request.args.get('from')
        if date_from:
            conditions.append('timestamp >= ?')
            params.append(date_from)

        date_to = request.args.get('to')
        if date_to:
            conditions.append('timestamp <= ?')
            params.append(date_to)

        where = ' AND '.join(conditions)
        rows = conn.execute(
            f'SELECT * FROM audit_log WHERE {where} ORDER BY timestamp DESC LIMIT 200',
            params
        ).fetchall()
        conn.close()

        variances = []
        for r in rows:
            try:
                est = float(r['old_value']) if r['old_value'] else 0
                act = float(r['new_value']) if r['new_value'] else 0
                variance = act - est
                variances.append({
                    'trade_id': r['entity_id'],
                    'timestamp': r['timestamp'],
                    'user': r['user'],
                    'estimated': est,
                    'actual': act,
                    'variance': round(variance, 2),
                    'variance_pct': round((variance / est * 100), 2) if est else 0,
                    'details': r['details']
                })
            except (ValueError, TypeError):
                continue

        return jsonify(variances)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== AUTO-OFFERING SYSTEM ====================

def _compute_offering_products(profile, destination):
    """Generate offering line items for a profile using best-cost sourcing + freight + seasonal adjustment."""
    products = json.loads(profile['products']) if isinstance(profile['products'], str) else profile['products']
    preferred = json.loads(profile['preferred_mills']) if profile['preferred_mills'] else []
    margin_target = float(profile['margin_target'] or 25)
    result_products = []

    # Only consider quotes from the last 2 business days
    quote_cutoff = business_day_cutoff(2)

    for product in products:
        # 1. Get latest mill quotes for this product (only recent)
        conn = get_mi_db()
        mill_rows = conn.execute("""
            SELECT mill_name, price, date FROM mill_quotes
            WHERE product=? AND price > 0 AND date >= ?
            ORDER BY date DESC
        """, (product, quote_cutoff)).fetchall()
        conn.close()

        # Deduplicate: latest price per mill
        seen_mills = {}
        for r in mill_rows:
            mill = r['mill_name']
            if mill not in seen_mills:
                seen_mills[mill] = {'mill': mill, 'fob': float(r['price']), 'date': r['date']}

        if not seen_mills:
            result_products.append({'product': product, 'error': 'No pricing available within 30 days'})
            continue

        # 2. If preferred mills set, filter to those (but fall back to all if none match)
        if preferred:
            preferred_matches = {k: v for k, v in seen_mills.items()
                                 if any(p.lower() in k.lower() for p in preferred)}
            if preferred_matches:
                seen_mills = preferred_matches

        # 3. Calculate freight for each mill
        candidates = []
        for mill_name, mill_data in seen_mills.items():
            origin = ''
            dir_entry = MILL_DIRECTORY.get(mill_name)
            if dir_entry:
                origin = f"{dir_entry[0]}, {dir_entry[1]}"
            else:
                for k, v in MILL_DIRECTORY.items():
                    if mill_name.startswith(k.split(' - ')[0]):
                        origin = f"{v[0]}, {v[1]}"
                        break

            if not origin or not destination:
                continue

            try:
                freight_per_mbf = 0
                miles = None
                mi_conn = get_mi_db()
                lane = mi_conn.execute(
                    "SELECT miles FROM lanes WHERE origin=? AND destination=?",
                    (origin, destination)
                ).fetchone()
                mi_conn.close()

                if lane:
                    miles = float(lane['miles'])
                else:
                    try:
                        coords_o = geocode_location(origin)
                        coords_d = geocode_location(destination)
                        if coords_o and coords_d:
                            miles = get_distance(coords_o, coords_d)
                    except Exception:
                        pass

                if miles:
                    base = 450
                    rate = 2.25
                    mbf_per_tl = 23
                    freight_per_mbf = round((base + miles * rate) / mbf_per_tl)
                else:
                    freight_per_mbf = 20
            except Exception:
                freight_per_mbf = 20

            landed = mill_data['fob'] + freight_per_mbf
            candidates.append({
                'mill': mill_name,
                'fob': mill_data['fob'],
                'freight': freight_per_mbf,
                'landed': landed,
                'date': mill_data['date']
            })

        if not candidates:
            result_products.append({'product': product, 'error': 'Could not calculate freight'})
            continue

        candidates.sort(key=lambda c: c['landed'])
        best = candidates[0]

        # 4. Seasonal margin adjustment
        seasonal_adj = 0
        seasonal_note = ''
        try:
            mi_conn = get_mi_db()
            cutoff_5y = (datetime.now() - timedelta(days=365 * 5)).strftime('%Y-%m-%d')
            s_rows = mi_conn.execute(
                "SELECT date, price FROM rl_prices WHERE product=? AND region='west' AND length='RL' AND date>=?",
                (product, cutoff_5y)
            ).fetchall()
            mi_conn.close()

            if s_rows:
                current_month = datetime.now().month
                month_prices = [float(r['price']) for r in s_rows if int(r['date'][5:7]) == current_month]
                if month_prices:
                    latest = float(s_rows[-1]['price'])
                    pct = round(sum(1 for p in month_prices if p <= latest) / len(month_prices) * 100)
                    if pct < 30:
                        seasonal_adj = -5
                        seasonal_note = f"Below seasonal norm ({pct}th %ile) — tighter margin"
                    elif pct > 70:
                        seasonal_adj = 5
                        seasonal_note = f"Above seasonal norm ({pct}th %ile) — wider margin"
                    else:
                        seasonal_note = f"Near seasonal norm ({pct}th %ile)"
        except Exception:
            pass

        adjusted_margin = margin_target + seasonal_adj
        recommended_sell = best['landed'] + adjusted_margin

        # Build alternatives list (top 3 after best)
        alts = []
        for c in candidates[1:4]:
            alts.append({'mill': c['mill'], 'fob': c['fob'], 'freight': c['freight'],
                         'landed': c['landed'], 'diff': round(c['landed'] - best['landed'])})

        result_products.append({
            'product': product,
            'mill': best['mill'],
            'fob': best['fob'],
            'freight': best['freight'],
            'landed': best['landed'],
            'margin': adjusted_margin,
            'price': recommended_sell,
            'seasonalNote': seasonal_note,
            'quoteDate': best['date'],
            'alternatives': alts
        })

    return result_products


# --- Offering Profile CRUD ---

@app.route('/api/offerings/profiles', methods=['GET'])
def list_offering_profiles():
    """List all offering profiles, optionally filtered by trader."""
    try:
        conn = get_crm_db()
        trader = request.args.get('trader', '')
        if trader:
            rows = conn.execute('SELECT * FROM offering_profiles WHERE trader=? ORDER BY customer_name', (trader,)).fetchall()
        else:
            rows = conn.execute('SELECT * FROM offering_profiles ORDER BY customer_name').fetchall()
        conn.close()

        profiles = []
        for r in rows:
            d = dict(r)
            d['products'] = json.loads(d['products']) if d['products'] else []
            d['preferred_mills'] = json.loads(d['preferred_mills']) if d['preferred_mills'] else []
            profiles.append(d)

        return jsonify(profiles)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/profiles', methods=['POST'])
def create_offering_profile():
    """Create a new offering profile for a customer."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'JSON body required'}), 400

        required = ['customer_id', 'customer_name', 'destination', 'products', 'trader']
        for f in required:
            if not data.get(f):
                return jsonify({'error': f'{f} required'}), 400

        products_json = json.dumps(data['products']) if isinstance(data['products'], list) else data['products']
        preferred_json = json.dumps(data.get('preferred_mills', [])) if isinstance(data.get('preferred_mills', []), list) else data.get('preferred_mills', '[]')

        conn = get_crm_db()
        cur = conn.execute("""
            INSERT INTO offering_profiles (customer_id, customer_name, destination, products, margin_target,
                frequency, preferred_mills, day_of_week, active, notes, trader)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data['customer_id'], data['customer_name'], data['destination'], products_json,
            float(data.get('margin_target', 25)), data.get('frequency', 'weekly'),
            preferred_json, int(data.get('day_of_week', 1)),
            1 if data.get('active', True) else 0, data.get('notes', ''), data['trader']
        ))
        conn.commit()
        profile_id = cur.lastrowid
        conn.close()

        return jsonify({'id': profile_id, 'message': 'Profile created'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/profiles/<int:pid>', methods=['PUT'])
def update_offering_profile(pid):
    """Update an existing offering profile."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'JSON body required'}), 400

        conn = get_crm_db()
        existing = conn.execute('SELECT * FROM offering_profiles WHERE id=?', (pid,)).fetchone()
        if not existing:
            conn.close()
            return jsonify({'error': 'Profile not found'}), 404

        fields = []
        vals = []
        for col in ['customer_name', 'destination', 'notes', 'frequency', 'trader']:
            if col in data:
                fields.append(f'{col}=?')
                vals.append(data[col])
        if 'products' in data:
            fields.append('products=?')
            vals.append(json.dumps(data['products']) if isinstance(data['products'], list) else data['products'])
        if 'preferred_mills' in data:
            fields.append('preferred_mills=?')
            vals.append(json.dumps(data['preferred_mills']) if isinstance(data['preferred_mills'], list) else data['preferred_mills'])
        if 'margin_target' in data:
            fields.append('margin_target=?')
            vals.append(float(data['margin_target']))
        if 'day_of_week' in data:
            fields.append('day_of_week=?')
            vals.append(int(data['day_of_week']))
        if 'active' in data:
            fields.append('active=?')
            vals.append(1 if data['active'] else 0)

        if fields:
            fields.append("updated_at=datetime('now')")
            vals.append(pid)
            conn.execute(f"UPDATE offering_profiles SET {', '.join(fields)} WHERE id=?", vals)
            conn.commit()

        conn.close()
        return jsonify({'message': 'Profile updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/profiles/<int:pid>', methods=['DELETE'])
def delete_offering_profile(pid):
    """Delete an offering profile."""
    try:
        conn = get_crm_db()
        conn.execute('DELETE FROM offering_profiles WHERE id=?', (pid,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Profile deleted'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- Offering Generation ---

@app.route('/api/offerings/generate', methods=['POST'])
def generate_offerings():
    """Generate draft offerings for due profiles (or specific profile_id)."""
    try:
        data = request.get_json() or {}
        profile_id = data.get('profile_id')
        force = data.get('force', False)  # bypass schedule check

        conn = get_crm_db()

        if profile_id:
            profiles = conn.execute('SELECT * FROM offering_profiles WHERE id=? AND active=1', (profile_id,)).fetchall()
        else:
            profiles = conn.execute('SELECT * FROM offering_profiles WHERE active=1').fetchall()

        now = datetime.now()
        today_dow = now.weekday()  # 0=Mon
        generated = []

        for profile in profiles:
            p = dict(profile)

            # Check if schedule is due (skip if force=True)
            if not force and not profile_id:
                freq = p.get('frequency', 'weekly')
                target_dow = p.get('day_of_week', 1)

                if freq == 'daily':
                    pass  # always due
                elif freq == 'weekly':
                    if today_dow != target_dow:
                        continue
                elif freq == 'biweekly':
                    if today_dow != target_dow:
                        continue
                    # Check if we already generated this week
                    week_start = (now - timedelta(days=today_dow)).strftime('%Y-%m-%d')
                    existing = conn.execute(
                        "SELECT COUNT(*) FROM offerings WHERE profile_id=? AND generated_at>=?",
                        (p['id'], week_start)
                    ).fetchone()[0]
                    if existing > 0:
                        continue

            # Check if we already have a draft for this profile today
            today_str = now.strftime('%Y-%m-%d')
            existing_today = conn.execute(
                "SELECT COUNT(*) FROM offerings WHERE profile_id=? AND status='draft' AND DATE(generated_at)=?",
                (p['id'], today_str)
            ).fetchone()[0]
            if existing_today > 0 and not force:
                continue

            # Generate offering products
            result_products = _compute_offering_products(p, p['destination'])

            # Calculate total margin
            total_margin = sum(item.get('margin', 0) for item in result_products if 'error' not in item)

            # Set expiration: 3 days from now
            expires = (now + timedelta(days=3)).strftime('%Y-%m-%d %H:%M:%S')

            # Insert offering
            cur = conn.execute("""
                INSERT INTO offerings (profile_id, customer_id, customer_name, destination, status,
                    products, margin_target, total_margin, expires_at, trader)
                VALUES (?, ?, ?, ?, 'draft', ?, ?, ?, ?, ?)
            """, (
                p['id'], p['customer_id'], p['customer_name'], p['destination'],
                json.dumps(result_products), p['margin_target'], total_margin,
                expires, p['trader']
            ))
            conn.commit()

            generated.append({
                'offering_id': cur.lastrowid,
                'customer': p['customer_name'],
                'products_count': len([x for x in result_products if 'error' not in x]),
                'total_margin': total_margin
            })

        conn.close()
        return jsonify({'generated': len(generated), 'offerings': generated})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- Offering Management ---

@app.route('/api/offerings', methods=['GET'])
def list_offerings():
    """List offerings with optional filters."""
    try:
        conn = get_crm_db()
        status = request.args.get('status', '')
        customer_id = request.args.get('customer_id', '')
        trader = request.args.get('trader', '')
        limit_n = int(request.args.get('limit', 50))

        sql = 'SELECT * FROM offerings WHERE 1=1'
        params = []

        if status:
            sql += ' AND status=?'
            params.append(status)
        if customer_id:
            sql += ' AND customer_id=?'
            params.append(int(customer_id))
        if trader:
            sql += ' AND trader=?'
            params.append(trader)

        sql += ' ORDER BY generated_at DESC LIMIT ?'
        params.append(limit_n)

        rows = conn.execute(sql, params).fetchall()
        conn.close()

        offerings = []
        for r in rows:
            d = dict(r)
            d['products'] = json.loads(d['products']) if isinstance(d['products'], str) else d['products']
            offerings.append(d)

        return jsonify(offerings)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/<int:oid>', methods=['GET'])
def get_offering(oid):
    """Get a single offering by ID."""
    try:
        conn = get_crm_db()
        row = conn.execute('SELECT * FROM offerings WHERE id=?', (oid,)).fetchone()
        conn.close()
        if not row:
            return jsonify({'error': 'Offering not found'}), 404
        d = dict(row)
        d['products'] = json.loads(d['products']) if isinstance(d['products'], str) else d['products']
        return jsonify(d)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/<int:oid>', methods=['PUT'])
def update_offering(oid):
    """Update an offering (edit prices, notes, status)."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'JSON body required'}), 400

        conn = get_crm_db()
        existing = conn.execute('SELECT * FROM offerings WHERE id=?', (oid,)).fetchone()
        if not existing:
            conn.close()
            return jsonify({'error': 'Offering not found'}), 404

        fields = []
        vals = []

        if 'products' in data:
            products = data['products']
            fields.append('products=?')
            vals.append(json.dumps(products) if isinstance(products, list) else products)
            # Recalculate total margin
            if isinstance(products, list):
                total_margin = sum(item.get('margin', 0) for item in products if isinstance(item, dict) and 'error' not in item)
                fields.append('total_margin=?')
                vals.append(total_margin)

        if 'status' in data:
            fields.append('status=?')
            vals.append(data['status'])
            if data['status'] == 'approved':
                fields.append("approved_at=datetime('now')")
                fields.append('approved_by=?')
                vals.append(data.get('approved_by', 'Ian'))
            elif data['status'] == 'sent':
                fields.append("sent_at=datetime('now')")

        if 'edit_notes' in data:
            fields.append('edit_notes=?')
            vals.append(data['edit_notes'])

        if fields:
            vals.append(oid)
            # Build SET clause carefully (some fields have no placeholder)
            set_parts = []
            final_vals = []
            for f, v in zip(fields, vals[:-1]):
                set_parts.append(f)
                if '=?' in f:
                    final_vals.append(v)
            final_vals.append(oid)
            conn.execute(f"UPDATE offerings SET {', '.join(set_parts)} WHERE id=?", final_vals)
            conn.commit()

        conn.close()
        return jsonify({'message': 'Offering updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/<int:oid>/approve', methods=['PUT'])
def approve_offering(oid):
    """Quick-approve an offering."""
    try:
        data = request.get_json() or {}
        conn = get_crm_db()
        existing = conn.execute('SELECT * FROM offerings WHERE id=?', (oid,)).fetchone()
        if not existing:
            conn.close()
            return jsonify({'error': 'Offering not found'}), 404

        conn.execute("""
            UPDATE offerings SET status='approved', approved_at=datetime('now'),
                approved_by=?, edit_notes=?
            WHERE id=?
        """, (data.get('approved_by', 'Ian'), data.get('notes', ''), oid))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Offering approved'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/history/<int:customer_id>', methods=['GET'])
def offering_history(customer_id):
    """Get offering history for a specific customer."""
    try:
        conn = get_crm_db()
        limit_n = int(request.args.get('limit', 20))
        rows = conn.execute(
            'SELECT * FROM offerings WHERE customer_id=? ORDER BY generated_at DESC LIMIT ?',
            (customer_id, limit_n)
        ).fetchall()
        conn.close()

        offerings = []
        for r in rows:
            d = dict(r)
            d['products'] = json.loads(d['products']) if isinstance(d['products'], str) else d['products']
            offerings.append(d)

        return jsonify(offerings)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/offerings/pending-count', methods=['GET'])
def offerings_pending_count():
    """Get count of pending draft offerings."""
    try:
        conn = get_crm_db()
        trader = request.args.get('trader', '')
        if trader:
            cnt = conn.execute("SELECT COUNT(*) FROM offerings WHERE status='draft' AND trader=?", (trader,)).fetchone()[0]
        else:
            cnt = conn.execute("SELECT COUNT(*) FROM offerings WHERE status='draft'").fetchone()[0]
        conn.close()
        return jsonify({'count': cnt})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== ENHANCED DASHBOARD ====================

@app.route('/api/dashboard/summary', methods=['GET'])

def dashboard_summary():
    """Aggregated KPI data for the enterprise dashboard."""
    try:
        conn = get_crm_db()
        now = datetime.now()
        today = now.strftime('%Y-%m-%d')
        week_ago = (now - timedelta(days=7)).strftime('%Y-%m-%d')
        month_ago = (now - timedelta(days=30)).strftime('%Y-%m-%d')
        year_start = f'{now.year}-01-01'

        # Trade status counts
        status_counts = {}
        for s in VALID_TRADE_STATUSES:
            cnt = conn.execute(
                'SELECT COUNT(*) FROM trade_status WHERE status = ?', (s,)
            ).fetchone()[0]
            status_counts[s] = cnt

        pending_approvals = status_counts.get('pending', 0)
        open_positions = sum(status_counts.get(s, 0) for s in ['approved', 'confirmed', 'shipped'])

        # Recent audit entries
        recent_audit = conn.execute(
            'SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 10'
        ).fetchall()

        # Credit exposure summary
        credit_rows = conn.execute('SELECT * FROM credit_limits').fetchall()
        total_credit_limit = sum((r['credit_limit'] or 0) for r in credit_rows)
        total_exposure = sum((r['current_exposure'] or 0) for r in credit_rows)
        over_limit = [dict(r) for r in credit_rows
                      if (r['credit_limit'] or 0) > 0 and (r['current_exposure'] or 0) > (r['credit_limit'] or 0)]

        # CRM stats
        total_customers = conn.execute('SELECT COUNT(*) FROM customers').fetchone()[0]
        total_mills = conn.execute('SELECT COUNT(*) FROM mills').fetchone()[0]
        total_prospects = conn.execute(
            "SELECT COUNT(*) FROM prospects WHERE status IN ('prospect', 'qualified')"
        ).fetchone()[0]

        # Audit activity counts
        audit_today = conn.execute(
            "SELECT COUNT(*) FROM audit_log WHERE timestamp >= ?", (today,)
        ).fetchone()[0]
        audit_week = conn.execute(
            "SELECT COUNT(*) FROM audit_log WHERE timestamp >= ?", (week_ago,)
        ).fetchone()[0]

        conn.close()

        # Mill Intel stats
        mi_stats = {}
        try:
            mi_conn = get_mi_db()
            mi_stats['total_quotes'] = mi_conn.execute('SELECT COUNT(*) FROM mill_quotes').fetchone()[0]
            mi_stats['quotes_today'] = mi_conn.execute(
                'SELECT COUNT(*) FROM mill_quotes WHERE date = ?', (today,)
            ).fetchone()[0]
            mi_stats['quotes_this_week'] = mi_conn.execute(
                'SELECT COUNT(*) FROM mill_quotes WHERE date >= ?', (week_ago,)
            ).fetchone()[0]
            mi_stats['active_mills'] = mi_conn.execute(
                'SELECT COUNT(DISTINCT mill_name) FROM mill_quotes WHERE date >= ?', (week_ago,)
            ).fetchone()[0]

            # Top movers (biggest price changes in last 7 days)
            top_movers = mi_conn.execute('''
                SELECT product,
                       ROUND(AVG(CASE WHEN date >= ? THEN price END), 2) as current_avg,
                       ROUND(AVG(CASE WHEN date < ? AND date >= ? THEN price END), 2) as prev_avg,
                       COUNT(CASE WHEN date >= ? THEN 1 END) as recent_quotes
                FROM mill_quotes
                WHERE date >= ?
                GROUP BY product
                HAVING current_avg IS NOT NULL AND prev_avg IS NOT NULL
                ORDER BY ABS(current_avg - prev_avg) DESC
                LIMIT 5
            ''', (week_ago, week_ago, month_ago, week_ago, month_ago)).fetchall()
            mi_stats['top_movers'] = []
            for m in top_movers:
                d = dict(m)
                if d['prev_avg'] and d['current_avg']:
                    d['change'] = round(d['current_avg'] - d['prev_avg'], 2)
                    d['change_pct'] = round(d['change'] / d['prev_avg'] * 100, 1) if d['prev_avg'] else 0
                    mi_stats['top_movers'].append(d)

            mi_conn.close()
        except Exception:
            pass

        return jsonify({
            'trade_statuses': status_counts,
            'pending_approvals': pending_approvals,
            'open_positions': open_positions,
            'credit': {
                'total_limit': total_credit_limit,
                'total_exposure': total_exposure,
                'available': max(0, total_credit_limit - total_exposure),
                'utilization': round(total_exposure / total_credit_limit * 100, 1) if total_credit_limit else 0,
                'over_limit_count': len(over_limit),
                'over_limit': over_limit[:5]
            },
            'crm': {
                'customers': total_customers,
                'mills': total_mills,
                'active_prospects': total_prospects
            },
            'mill_intel': mi_stats,
            'audit': {
                'today': audit_today,
                'this_week': audit_week,
                'recent': [dict(r) for r in recent_audit]
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== OFFERING SCHEDULER ====================
import threading

def _offering_scheduler_loop():
    """Background thread: runs daily at 6 AM to generate offerings for due profiles."""
    import time as _time
    while True:
        try:
            now = datetime.now()
            # Next 6 AM
            target = now.replace(hour=6, minute=0, second=0, microsecond=0)
            if now >= target:
                target += timedelta(days=1)
            wait_seconds = (target - now).total_seconds()
            _time.sleep(wait_seconds)

            # Generate offerings for all due profiles
            with app.app_context():
                conn = get_crm_db()
                profiles = conn.execute('SELECT * FROM offering_profiles WHERE active=1').fetchall()
                today_dow = datetime.now().weekday()
                today_str = datetime.now().strftime('%Y-%m-%d')

                for profile in profiles:
                    p = dict(profile)
                    freq = p.get('frequency', 'weekly')
                    target_dow = p.get('day_of_week', 1)

                    if freq == 'daily':
                        pass
                    elif freq == 'weekly' and today_dow != target_dow:
                        continue
                    elif freq == 'biweekly' and today_dow != target_dow:
                        continue

                    # Skip if already generated today
                    existing = conn.execute(
                        "SELECT COUNT(*) FROM offerings WHERE profile_id=? AND DATE(generated_at)=?",
                        (p['id'], today_str)
                    ).fetchone()[0]
                    if existing > 0:
                        continue

                    result_products = _compute_offering_products(p, p['destination'])
                    total_margin = sum(item.get('margin', 0) for item in result_products if 'error' not in item)
                    expires = (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d %H:%M:%S')

                    conn.execute("""
                        INSERT INTO offerings (profile_id, customer_id, customer_name, destination, status,
                            products, margin_target, total_margin, expires_at, trader)
                        VALUES (?, ?, ?, ?, 'draft', ?, ?, ?, ?, ?)
                    """, (
                        p['id'], p['customer_id'], p['customer_name'], p['destination'],
                        json.dumps(result_products), p['margin_target'], total_margin,
                        expires, p['trader']
                    ))

                conn.commit()
                conn.close()
                print(f"[Scheduler] Generated offerings at {datetime.now()}")
        except Exception as e:
            print(f"[Scheduler] Error: {e}")
            import time as _time
            _time.sleep(3600)  # retry in 1 hour on error

# Start scheduler thread
_scheduler_thread = threading.Thread(target=_offering_scheduler_loop, daemon=True)
_scheduler_thread.start()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
